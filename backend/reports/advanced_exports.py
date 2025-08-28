import os
import zipfile
import tempfile
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone
from django.conf import settings
from django.db import transaction

# For enhanced PDF formatting
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .models import SavedReport, ReportExecution
from .services_new.summary_service import SummaryReportService
from .services_new.sales_service import SalesReportService
from .services_new.payments_service import PaymentsReportService
from .services_new.operations_service import OperationsReportService
from .services_new.export_service import ExportService
from .services_new.products_service import ProductsReportService
from users.models import User

logger = logging.getLogger(__name__)


class AdvancedExportService:
    """
    Advanced export service with bulk operations, compression, and enhanced formatting.
    """

    # Export priorities
    PRIORITY_LOW = 1
    PRIORITY_NORMAL = 2
    PRIORITY_HIGH = 3
    PRIORITY_URGENT = 4

    # File size limits (in bytes)
    MAX_SINGLE_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    MAX_ARCHIVE_SIZE = 200 * 1024 * 1024  # 200MB

    @classmethod
    def create_bulk_export(
        cls,
        user: User,
        report_configs: List[Dict[str, Any]],
        export_format: str = "xlsx",
        compress: bool = True,
        custom_template: Optional[Dict[str, Any]] = None,
        priority: int = PRIORITY_NORMAL,
    ) -> Tuple[str, int]:
        """
        Create a bulk export operation for multiple reports.

        Args:
            user: User requesting the export
            report_configs: List of report configurations
                [{"type": "sales", "start_date": "2024-01-01", "end_date": "2024-01-31", "filters": {}}]
            export_format: Format for export (csv, xlsx, pdf)
            compress: Whether to compress multiple files into ZIP
            custom_template: Custom template configuration
            priority: Export priority level

        Returns:
            Tuple of (operation_id, estimated_completion_time_seconds)
        """
        try:
            operation_id = cls._generate_operation_id()

            # Validate report configs
            validated_configs = cls._validate_report_configs(report_configs)

            # Estimate completion time
            estimated_time = cls._estimate_completion_time(
                validated_configs, export_format
            )

            # Create export metadata
            export_metadata = {
                "operation_id": operation_id,
                "user_id": user.id,
                "report_configs": validated_configs,
                "export_format": export_format,
                "compress": compress,
                "custom_template": custom_template,
                "priority": priority,
                "status": "queued",
                "created_at": timezone.now().isoformat(),
                "estimated_completion": estimated_time,
            }

            # Store metadata (you could use a separate model for this)
            cls._store_export_metadata(operation_id, export_metadata)

            logger.info(
                f"Bulk export created: {operation_id} with {len(validated_configs)} reports"
            )

            return operation_id, estimated_time

        except Exception as e:
            logger.error(f"Error creating bulk export: {e}")
            raise

    @classmethod
    def process_bulk_export(
        cls, operation_id: str, progress_callback: Optional[callable] = None
    ) -> Dict[str, Any]:
        """
        Process a bulk export operation.

        Args:
            operation_id: Unique operation identifier
            progress_callback: Optional callback for progress updates

        Returns:
            Export result metadata
        """
        try:
            # Load export metadata
            metadata = cls._load_export_metadata(operation_id)
            if not metadata:
                raise ValueError(f"Export operation {operation_id} not found")

            metadata["status"] = "processing"
            metadata["started_at"] = timezone.now().isoformat()
            cls._store_export_metadata(operation_id, metadata)

            user = User.objects.get(id=metadata["user_id"])
            report_configs = metadata["report_configs"]
            export_format = metadata["export_format"]
            compress = metadata["compress"]
            custom_template = metadata.get("custom_template")

            generated_files = []
            total_reports = len(report_configs)

            # Process each report
            for i, config in enumerate(report_configs):
                if progress_callback:
                    progress_callback(f"Generating report {i+1} of {total_reports}")

                try:
                    # Generate report data
                    report_data = cls._generate_single_report(user, config)

                    # Create export file
                    file_content, filename = cls._create_export_file(
                        report_data, config["type"], export_format, custom_template
                    )

                    generated_files.append(
                        {
                            "filename": filename,
                            "content": file_content,
                            "report_type": config["type"],
                            "size": len(file_content),
                        }
                    )

                except Exception as e:
                    logger.error(f"Error generating report {config['type']}: {e}")
                    # Continue with other reports
                    continue

            if not generated_files:
                raise Exception("No reports were successfully generated")

            # Create final export package
            final_file, final_filename = cls._create_export_package(
                generated_files, compress, operation_id
            )

            # Store the final file
            file_path = cls._store_export_file(final_file, final_filename, operation_id)

            # Update metadata
            metadata["status"] = "completed"
            metadata["completed_at"] = timezone.now().isoformat()
            metadata["file_path"] = file_path
            metadata["file_size"] = len(final_file)
            metadata["reports_generated"] = len(generated_files)
            cls._store_export_metadata(operation_id, metadata)

            logger.info(f"Bulk export completed: {operation_id}")

            return {
                "operation_id": operation_id,
                "status": "completed",
                "file_path": file_path,
                "file_size": len(final_file),
                "reports_generated": len(generated_files),
                "completion_time": metadata["completed_at"],
            }

        except Exception as e:
            # Update metadata with error
            metadata = cls._load_export_metadata(operation_id) or {}
            metadata["status"] = "failed"
            metadata["error"] = str(e)
            metadata["failed_at"] = timezone.now().isoformat()
            cls._store_export_metadata(operation_id, metadata)

            logger.error(f"Bulk export failed: {operation_id} - {e}")
            raise

    @classmethod
    def create_custom_template(
        cls, template_name: str, template_config: Dict[str, Any], user: User
    ) -> str:
        """
        Create a custom export template.

        Args:
            template_name: Name for the template
            template_config: Template configuration
            user: User creating the template

        Returns:
            Template ID
        """
        template_id = f"tpl_{int(timezone.now().timestamp())}_{user.id}"

        template_data = {
            "id": template_id,
            "name": template_name,
            "config": template_config,
            "created_by": user.id,
            "created_at": timezone.now().isoformat(),
        }

        # Store template (you might want a separate model for this)
        cls._store_template(template_id, template_data)

        return template_id

    @classmethod
    def get_export_status(cls, operation_id: str) -> Dict[str, Any]:
        """Get the status of an export operation."""
        metadata = cls._load_export_metadata(operation_id)
        if not metadata:
            return {"status": "not_found"}

        return {
            "operation_id": operation_id,
            "status": metadata.get("status", "unknown"),
            "progress": metadata.get("progress", 0),
            "created_at": metadata.get("created_at"),
            "estimated_completion": metadata.get("estimated_completion"),
            "file_path": metadata.get("file_path"),
            "file_size": metadata.get("file_size"),
            "reports_generated": metadata.get("reports_generated", 0),
            "error": metadata.get("error"),
        }

    @classmethod
    def cleanup_old_exports(cls, days_to_keep: int = 7) -> int:
        """
        Clean up old export files and metadata.

        Args:
            days_to_keep: Number of days to keep exports

        Returns:
            Number of exports cleaned up
        """
        cutoff_date = timezone.now() - timedelta(days=days_to_keep)
        cleaned_count = 0

        try:
            # This would need to be implemented based on your storage method
            # For now, it's a placeholder
            logger.info(f"Cleaned up {cleaned_count} old exports")
            return cleaned_count

        except Exception as e:
            logger.error(f"Error cleaning up old exports: {e}")
            return 0

    # Private helper methods
    @classmethod
    def _generate_operation_id(cls) -> str:
        """Generate a unique operation ID."""
        timestamp = int(timezone.now().timestamp())
        return f"export_{timestamp}_{os.urandom(4).hex()}"

    @classmethod
    def _validate_report_configs(
        cls, configs: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Validate report configurations."""
        validated = []
        valid_types = ["summary", "sales", "products", "payments", "operations"]

        for config in configs:
            if config.get("type") not in valid_types:
                logger.warning(f"Invalid report type: {config.get('type')}")
                continue

            # Validate dates
            try:
                start_date = datetime.strptime(config["start_date"], "%Y-%m-%d").date()
                end_date = datetime.strptime(config["end_date"], "%Y-%m-%d").date()

                if start_date > end_date:
                    logger.warning(
                        f"Invalid date range: {config['start_date']} to {config['end_date']}"
                    )
                    continue

            except (ValueError, KeyError) as e:
                logger.warning(f"Invalid date format in config: {e}")
                continue

            validated.append(config)

        return validated

    @classmethod
    def _estimate_completion_time(
        cls, configs: List[Dict[str, Any]], format: str
    ) -> int:
        """Estimate completion time in seconds."""
        base_time_per_report = {"csv": 2, "xlsx": 5, "pdf": 8}

        base_time = base_time_per_report.get(format, 5)
        total_time = len(configs) * base_time

        # Add compression time if multiple reports
        if len(configs) > 1:
            total_time += 3

        return total_time

    @classmethod
    def _generate_single_report(
        cls, user: User, config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Generate a single report based on configuration."""
        report_type = config["type"]
        start_date = datetime.strptime(config["start_date"], "%Y-%m-%d").date()
        end_date = datetime.strptime(config["end_date"], "%Y-%m-%d").date()
        filters = config.get("filters", {})

        # Use existing ReportService methods
        if report_type == "summary":
            return SummaryReportService.generate_summary_report(
                start_date, end_date, use_cache=False
            )
        elif report_type == "sales":
            return SalesReportService.generate_sales_report(
                start_date, end_date, use_cache=False
            )
        elif report_type == "products":
            return ProductsReportService.generate_products_report(
                start_date,
                end_date,
                category_id=filters.get("category_id"),
                limit=filters.get("limit", 50),
                trend_period=filters.get("trend_period", "auto"),
                use_cache=False,
            )
        elif report_type == "payments":
            return PaymentsReportService.generate_payments_report(
                start_date, end_date, use_cache=False
            )
        elif report_type == "operations":
            return OperationsReportService.generate_operations_report(
                start_date, end_date, use_cache=False
            )
        else:
            raise ValueError(f"Unknown report type: {report_type}")

    @classmethod
    def _create_export_file(
        cls,
        report_data: Dict[str, Any],
        report_type: str,
        format: str,
        custom_template: Optional[Dict[str, Any]] = None,
    ) -> Tuple[bytes, str]:
        """Create an export file for a single report."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{timestamp}.{format}"

        if format == "csv":
            # Use specific service for refactored reports, original service for others
            if report_type == "products":
                content = ProductsReportService.export_products_to_csv(report_data)
            else:
                content = ExportService.export_to_csv(report_data, report_type)
        elif format == "xlsx":
            if report_type == "products":
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Products Report"
                
                # Styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Call the new ProductsReportService export method
                ProductsReportService.export_products_to_xlsx(report_data, ws, header_font, header_fill, header_alignment)
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                content = output.getvalue()
            else:
                content = ExportService.export_to_xlsx(report_data, report_type)
        elif format == "pdf":
            if report_type == "products":
                import io
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.lib.enums import TA_CENTER
                
                # Create PDF in landscape mode for better table display
                output = io.BytesIO()
                doc = SimpleDocTemplate(output, pagesize=landscape(letter), 
                                      leftMargin=0.5*inch, rightMargin=0.5*inch,
                                      topMargin=0.5*inch, bottomMargin=0.5*inch)
                
                # Create story and styles
                story = []
                styles = getSampleStyleSheet()
                
                # Add custom title style
                styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], 
                                        alignment=TA_CENTER, fontSize=18, spaceAfter=20))
                
                # Call the new ProductsReportService export method
                ProductsReportService.export_products_to_pdf(report_data, story, styles)
                
                # Build PDF
                doc.build(story)
                content = output.getvalue()
            else:
                content = ExportService.export_to_pdf(report_data, report_type)
        else:
            raise ValueError(f"Unsupported export format: {format}")

        return content, filename

    @classmethod
    def _create_export_package(
        cls, files: List[Dict[str, Any]], compress: bool, operation_id: str
    ) -> Tuple[bytes, str]:
        """Create the final export package (single file or ZIP)."""
        if len(files) == 1 and not compress:
            # Single file, no compression needed
            file_info = files[0]
            return file_info["content"], file_info["filename"]

        # Multiple files or compression requested - create ZIP
        zip_buffer = BytesIO()
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"reports_export_{timestamp}.zip"

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for file_info in files:
                zip_file.writestr(file_info["filename"], file_info["content"])

            # Add a manifest file
            manifest = {
                "operation_id": operation_id,
                "created_at": timezone.now().isoformat(),
                "files": [
                    {
                        "filename": f["filename"],
                        "report_type": f["report_type"],
                        "size": f["size"],
                    }
                    for f in files
                ],
            }
            zip_file.writestr("manifest.json", json.dumps(manifest, indent=2))

        zip_buffer.seek(0)
        return zip_buffer.read(), zip_filename

    @classmethod
    def _store_export_file(
        cls, file_content: bytes, filename: str, operation_id: str
    ) -> str:
        """Store the export file and return the file path."""
        # Create a path with operation_id for organization
        file_path = f"exports/{operation_id}/{filename}"

        # Store using Django's default storage
        saved_path = default_storage.save(file_path, ContentFile(file_content))

        return saved_path

    @classmethod
    def _store_export_metadata(
        cls, operation_id: str, metadata: Dict[str, Any]
    ) -> None:
        """Store export operation metadata."""
        # For now, store in cache or a simple file
        # In production, you might want a dedicated model
        from django.core.cache import cache

        cache.set(
            f"export_metadata_{operation_id}", metadata, timeout=86400 * 7
        )  # 7 days

    @classmethod
    def _load_export_metadata(cls, operation_id: str) -> Optional[Dict[str, Any]]:
        """Load export operation metadata."""
        from django.core.cache import cache

        return cache.get(f"export_metadata_{operation_id}")

    @classmethod
    def _store_template(cls, template_id: str, template_data: Dict[str, Any]) -> None:
        """Store export template."""
        from django.core.cache import cache

        cache.set(
            f"export_template_{template_id}", template_data, timeout=86400 * 30
        )  # 30 days


class ExportQueue:
    """
    Export queue management for handling multiple export operations.
    """

    @classmethod
    def add_to_queue(
        cls, operation_id: str, priority: int = AdvancedExportService.PRIORITY_NORMAL
    ) -> None:
        """Add an export operation to the processing queue."""
        from django.core.cache import cache

        queue_key = f"export_queue_p{priority}"
        queue = cache.get(queue_key, [])

        queue_item = {
            "operation_id": operation_id,
            "added_at": timezone.now().isoformat(),
            "priority": priority,
        }

        queue.append(queue_item)
        cache.set(queue_key, queue, timeout=86400)  # 24 hours

    @classmethod
    def get_next_operation(cls) -> Optional[str]:
        """Get the next operation to process from the queue."""
        from django.core.cache import cache

        # Check queues in priority order (highest to lowest)
        priorities = [
            AdvancedExportService.PRIORITY_URGENT,
            AdvancedExportService.PRIORITY_HIGH,
            AdvancedExportService.PRIORITY_NORMAL,
            AdvancedExportService.PRIORITY_LOW,
        ]

        for priority in priorities:
            queue_key = f"export_queue_p{priority}"
            queue = cache.get(queue_key, [])

            if queue:
                # Get the first item (FIFO within priority)
                operation = queue.pop(0)
                cache.set(queue_key, queue, timeout=86400)
                return operation["operation_id"]

        return None

    @classmethod
    def get_queue_status(cls) -> Dict[str, Any]:
        """Get the current status of all export queues."""
        from django.core.cache import cache

        status = {"total_operations": 0, "by_priority": {}}

        priorities = [
            ("urgent", AdvancedExportService.PRIORITY_URGENT),
            ("high", AdvancedExportService.PRIORITY_HIGH),
            ("normal", AdvancedExportService.PRIORITY_NORMAL),
            ("low", AdvancedExportService.PRIORITY_LOW),
        ]

        for priority_name, priority_value in priorities:
            queue_key = f"export_queue_p{priority_value}"
            queue = cache.get(queue_key, [])

            status["by_priority"][priority_name] = {
                "count": len(queue),
                "operations": [op["operation_id"] for op in queue],
            }
            status["total_operations"] += len(queue)

        return status
