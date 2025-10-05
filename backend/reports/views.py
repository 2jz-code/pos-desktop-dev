import logging
import io
from datetime import datetime
from rest_framework import viewsets, permissions, status, filters
from core_backend.base import BaseViewSet, ReadOnlyBaseViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.http import Http404

from core_backend.pagination import StandardPagination
from users.permissions import IsManagerOrHigher

from .models import SavedReport, ReportTemplate, ReportExecution, ReportCache
from .serializers import (
    ReportParameterSerializer,
    ProductReportParameterSerializer,
    SavedReportSerializer,
    SavedReportCreateSerializer,
    ReportTemplateSerializer,
    ReportExecutionSerializer,
    ReportCacheSerializer,
    ReportExportRequestSerializer,
    BulkExportRequestSerializer,
    BulkExportStatusSerializer,
    ExportQueueStatusSerializer,
    CustomTemplateSerializer,
)
# Import from the new modular services
from .services_new.sales_service import SalesReportService  # New modular service
from .services_new.summary_service import SummaryReportService  # New modular service
from .services_new.payments_service import PaymentsReportService  # New modular service
from .services_new.products_service import ProductsReportService  # New modular service
from .services_new.operations_service import OperationsReportService  # New modular service
from .services_new.saved_reports_service import SavedReportService  # New modular service
from .services_new.export_service import ExportService  # New modular service
from .advanced_exports import AdvancedExportService, ExportQueue
from .tasks import create_bulk_export_async, process_export_queue

logger = logging.getLogger(__name__)

class ReportViewSet(viewsets.ViewSet):
    """
    Comprehensive report generation ViewSet following existing patterns
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        """Generate summary report matching the frontend UI requirements"""
        serializer = ReportParameterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = serializer.validated_data["start_date"]
            end_date = serializer.validated_data["end_date"]

            # Use cache by default, but allow bypassing with ?use_cache=false
            use_cache = request.query_params.get("use_cache", "true").lower() != "false"

            report_data = SummaryReportService.generate_summary_report(
                tenant=request.tenant,
                start_date=start_date, end_date=end_date, use_cache=use_cache
            )

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Summary report generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate summary report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="sales")
    def sales(self, request):
        """Generate sales report matching the frontend UI requirements"""
        serializer = ReportParameterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = serializer.validated_data["start_date"]
            end_date = serializer.validated_data["end_date"]
            group_by = request.query_params.get("group_by", "day")
            use_cache = request.query_params.get("use_cache", "true").lower() != "false"

            report_data = SalesReportService.generate_sales_report(
                tenant=request.tenant,
                start_date=start_date, end_date=end_date, group_by=group_by, use_cache=use_cache
            )

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Sales report generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate sales report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="products")
    def products(self, request):
        """Generate products report matching the frontend UI requirements"""
        serializer = ProductReportParameterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = serializer.validated_data["start_date"]
            end_date = serializer.validated_data["end_date"]
            category_id = serializer.validated_data.get("category_id")
            limit = serializer.validated_data.get("limit", 10)
            trend_period = request.query_params.get("trend_period", "auto")
            use_cache = request.query_params.get("use_cache", "true").lower() != "false"

            report_data = ProductsReportService.generate_products_report(
                tenant=request.tenant,
                start_date=start_date,
                end_date=end_date,
                category_id=category_id,
                limit=limit,
                trend_period=trend_period,
                use_cache=use_cache,
            )

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Products report generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate products report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="payments")
    def payments(self, request):
        """Generate payments report matching the frontend UI requirements"""
        serializer = ReportParameterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = serializer.validated_data["start_date"]
            end_date = serializer.validated_data["end_date"]
            use_cache = request.query_params.get("use_cache", "true").lower() != "false"

            report_data = PaymentsReportService.generate_payments_report(
                tenant=request.tenant,
                start_date=start_date, end_date=end_date, use_cache=use_cache
            )

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Payments report generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate payments report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="operations")
    def operations(self, request):
        """Generate operations report matching the frontend UI requirements"""
        serializer = ReportParameterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = serializer.validated_data["start_date"]
            end_date = serializer.validated_data["end_date"]
            use_cache = request.query_params.get("use_cache", "true").lower() != "false"

            report_data = OperationsReportService.generate_operations_report(
                tenant=request.tenant,
                start_date=start_date, end_date=end_date, use_cache=use_cache
            )

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Operations report generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate operations report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="quick-metrics")
    def quick_metrics(self, request):
        """Get today/MTD/YTD quick metrics for dashboard"""
        try:
            metrics_data = SummaryReportService.get_quick_metrics()
            return Response(metrics_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Quick metrics generation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate quick metrics", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="export")
    def export(self, request):
        """Export report to file"""
        serializer = ReportExportRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            report_type = serializer.validated_data["report_type"]
            parameters = serializer.validated_data["parameters"]
            format_type = serializer.validated_data["format"]
            
            # Extract date parameters
            start_date = parameters.get("start_date")
            end_date = parameters.get("end_date")
            
            if not start_date or not end_date:
                return Response(
                    {"error": "start_date and end_date are required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Convert date strings to datetime objects
            try:
                start_date = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                end_date = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use ISO 8601 format."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Generate the report data first
            if report_type == "summary":
                report_data = SummaryReportService.generate_summary_report(start_date, end_date)
            elif report_type == "sales":
                report_data = SalesReportService.generate_sales_report(start_date, end_date)
            elif report_type == "products":
                category_id = parameters.get("category_id")
                limit = parameters.get("limit", 10)
                trend_period = parameters.get("trend_period", "auto")
                report_data = ProductsReportService.generate_products_report(
                    start_date, end_date, category_id, limit, trend_period
                )
            elif report_type == "payments":
                report_data = PaymentsReportService.generate_payments_report(start_date, end_date)
            elif report_type == "operations":
                report_data = OperationsReportService.generate_operations_report(start_date, end_date)
            else:
                return Response(
                    {"error": f"Unknown report type: {report_type}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Export to the requested format
            format_type = format_type.lower()
            if format_type == "csv":
                # Use specific service for refactored reports, original service for others
                if report_type == "sales":
                    file_data = SalesReportService.export_sales_to_csv(report_data)
                elif report_type == "payments":
                    file_data = PaymentsReportService.export_payments_to_csv(report_data)
                elif report_type == "products":
                    file_data = ProductsReportService.export_products_to_csv(report_data)
                elif report_type == "operations":
                    file_data = OperationsReportService.export_operations_to_csv(report_data)
                else:
                    file_data = ExportService.export_to_csv(report_data, report_type)
                content_type = "text/csv"
                file_extension = "csv"
            elif format_type == "xlsx" or format_type == "excel":
                # Use specific service for sales and payments reports, original service for others
                if report_type == "sales":
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sales Report"
                    
                    # Styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Call the new SalesReportService export method
                    SalesReportService.export_sales_to_xlsx(report_data, ws, header_font, header_fill, header_alignment)
                    
                    # Save to bytes
                    output = io.BytesIO()
                    wb.save(output)
                    file_data = output.getvalue()
                elif report_type == "payments":
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Payments Report"
                    
                    # Styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Call the new PaymentsReportService export method
                    PaymentsReportService.export_payments_to_xlsx(report_data, ws, header_font, header_fill, header_alignment)
                    
                    # Save to bytes
                    output = io.BytesIO()
                    wb.save(output)
                    file_data = output.getvalue()
                elif report_type == "products":
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Products Report"
                    
                    # Styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Call the new ProductsReportService export method
                    ProductsReportService.export_products_to_xlsx(report_data, ws, header_font, header_fill, header_alignment)
                    
                    # Save to bytes
                    output = io.BytesIO()
                    wb.save(output)
                    file_data = output.getvalue()
                elif report_type == "operations":
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Operations Report"
                    
                    # Styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Call the new OperationsReportService export method
                    OperationsReportService.export_operations_to_xlsx(report_data, ws, header_font, header_fill, header_alignment)
                    
                    # Save to bytes
                    output = io.BytesIO()
                    wb.save(output)
                    file_data = output.getvalue()
                else:
                    file_data = ExportService.export_to_xlsx(report_data, report_type)
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                file_extension = "xlsx"
            elif format_type == "pdf":
                # Use specific service for sales and payments reports, original service for others
                if report_type == "sales":
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.lib.enums import TA_CENTER
                    
                    # Create PDF in landscape mode for better table display
                    output = io.BytesIO()
                    doc = SimpleDocTemplate(output, pagesize=landscape(letter), 
                                          leftMargin=0.5*inch, rightMargin=0.5*inch,
                                          topMargin=0.5*inch, bottomMargin=0.5*inch)
                    
                    # Create story and styles
                    story = []
                    styles = getSampleStyleSheet()
                    
                    # Add custom title style
                    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], 
                                            alignment=TA_CENTER, fontSize=18, spaceAfter=20))
                    
                    # Call the new SalesReportService export method
                    SalesReportService.export_sales_to_pdf(report_data, story, styles)
                    
                    # Build PDF
                    doc.build(story)
                    file_data = output.getvalue()
                elif report_type == "payments":
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.lib.enums import TA_CENTER
                    
                    # Create PDF
                    output = io.BytesIO()
                    doc = SimpleDocTemplate(output, pagesize=letter, 
                                          leftMargin=0.5*inch, rightMargin=0.5*inch,
                                          topMargin=0.5*inch, bottomMargin=0.5*inch)
                    
                    # Create story and styles
                    story = []
                    styles = getSampleStyleSheet()
                    
                    # Add custom title style
                    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], 
                                            alignment=TA_CENTER, fontSize=18, spaceAfter=20))
                    
                    # Call the new PaymentsReportService export method
                    PaymentsReportService.export_payments_to_pdf(report_data, story, styles)
                    
                    # Build PDF
                    doc.build(story)
                    file_data = output.getvalue()
                elif report_type == "products":
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.lib.enums import TA_CENTER
                    
                    # Create PDF in landscape mode for better table display
                    output = io.BytesIO()
                    doc = SimpleDocTemplate(output, pagesize=landscape(letter), 
                                          leftMargin=0.5*inch, rightMargin=0.5*inch,
                                          topMargin=0.5*inch, bottomMargin=0.5*inch)
                    
                    # Create story and styles
                    story = []
                    styles = getSampleStyleSheet()
                    
                    # Add custom title style
                    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], 
                                            alignment=TA_CENTER, fontSize=18, spaceAfter=20))
                    
                    # Call the new ProductsReportService export method
                    ProductsReportService.export_products_to_pdf(report_data, story, styles)
                    
                    # Build PDF
                    doc.build(story)
                    file_data = output.getvalue()
                elif report_type == "operations":
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import inch
                    from reportlab.lib.enums import TA_CENTER
                    
                    # Create PDF in portrait mode for concise operations report
                    output = io.BytesIO()
                    doc = SimpleDocTemplate(output, pagesize=letter, 
                                          leftMargin=0.75*inch, rightMargin=0.75*inch,
                                          topMargin=0.75*inch, bottomMargin=0.75*inch)
                    
                    # Create story and styles
                    story = []
                    styles = getSampleStyleSheet()
                    
                    # Add custom title style
                    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], 
                                            alignment=TA_CENTER, fontSize=18, spaceAfter=20))
                    
                    # Call the new OperationsReportService export method
                    OperationsReportService.export_operations_to_pdf(story, report_data, styles)
                    
                    # Build PDF
                    doc.build(story)
                    file_data = output.getvalue()
                else:
                    file_data = ExportService.export_to_pdf(report_data, report_type)
                content_type = "application/pdf"
                file_extension = "pdf"
            else:
                return Response(
                    {"error": f"Unknown format: {format_type}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Create filename
            filename = f"{report_type}-report-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.{file_extension}"
            
            # Return file response
            from django.http import HttpResponse
            response = HttpResponse(file_data, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            logger.error(f"Report export failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to export report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="cache-stats")
    def cache_stats(self, request):
        """Get cache statistics (admin only)"""
        if not request.user.is_staff:
            return Response(
                {"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN
            )

        try:
            from .services_new.base import BaseReportService
            stats = BaseReportService.get_cache_stats()
            return Response(stats, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Cache stats retrieval failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to retrieve cache statistics", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="clear-cache")
    def clear_cache(self, request):
        """Clear expired cache entries (admin only)"""
        if not request.user.is_staff:
            return Response(
                {"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN
            )

        try:
            from .services_new.base import BaseReportService
            deleted_count = BaseReportService.cleanup_expired_cache()
            return Response(
                {
                    "message": f"Cleaned up {deleted_count} expired cache entries",
                    "deleted_count": deleted_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.error(f"Cache cleanup failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to clean cache", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class SavedReportViewSet(BaseViewSet):
    """
    Saved reports management following existing patterns
    """

    queryset = SavedReport.objects.all()
    serializer_class = SavedReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["report_type", "status", "schedule", "format"]
    search_fields = ["name", "report_type"]
    ordering_fields = ["created_at", "last_run", "name"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Filter by user - users can only see their own reports"""
        queryset = super().get_queryset()

        # Staff can see all reports, regular users only their own
        if self.request.user.is_staff:
            return queryset

        return queryset.filter(user=self.request.user)

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == "create":
            return SavedReportCreateSerializer
        return SavedReportSerializer

    def perform_create(self, serializer):
        """Set user on creation"""
        serializer.save(user=self.request.user)

    @action(detail=True, methods=["post"], url_path="run")
    def run(self, request, pk=None):
        """Run a saved report"""
        saved_report = self.get_object()

        try:
            # Use the new SavedReportService to run the report
            report_data = SavedReportService.run_saved_report(
                saved_report=saved_report,
                user=request.user
            )
            
            return Response(report_data, status=status.HTTP_200_OK)

        except PermissionError as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_403_FORBIDDEN
            )
        except ValueError as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Saved report execution failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to run saved report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["post"], url_path="duplicate")
    def duplicate(self, request, pk=None):
        """Duplicate a saved report"""
        saved_report = self.get_object()

        try:
            # Use the new SavedReportService to duplicate the report
            new_name = request.data.get("name")  # Allow optional custom name
            duplicated_report = SavedReportService.duplicate_saved_report(
                saved_report=saved_report,
                user=request.user,
                new_name=new_name
            )

            serializer = self.get_serializer(duplicated_report)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except PermissionError as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            logger.error(f"Report duplication failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to duplicate report", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class ReportTemplateViewSet(BaseViewSet):
    """
    Report templates management
    """

    queryset = ReportTemplate.objects.all()
    serializer_class = ReportTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["report_type", "is_system_template"]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "created_at"]
    ordering = ["name"]

    def perform_create(self, serializer):
        """Set created_by on creation"""
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="create-report")
    def create_report(self, request, pk=None):
        """Create a saved report from template"""
        template = self.get_object()

        # Get custom name from request or use template name
        report_name = request.data.get("name", f"Report from {template.name}")
        
        # Get optional custom parameters
        custom_parameters = request.data.get("parameters")

        try:
            # Use the new SavedReportService to create from template
            saved_report = SavedReportService.create_from_template(
                template=template,
                user=request.user,
                report_name=report_name,
                parameters=custom_parameters
            )

            serializer = SavedReportSerializer(saved_report)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Report creation from template failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to create report from template", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class ReportCacheViewSet(ReadOnlyBaseViewSet):
    """
    Report cache management (admin only)
    """

    queryset = ReportCache.objects.all()
    serializer_class = ReportCacheSerializer
    permission_classes = [IsManagerOrHigher]
    filterset_fields = ["report_type"]
    search_fields = ["parameters_hash"]
    ordering_fields = ["generated_at", "expires_at"]
    ordering = ["-generated_at"]

    @action(detail=False, methods=["post"], url_path="cleanup")
    def cleanup(self, request):
        """Clean up expired cache entries"""
        try:
            from .services_new.base import BaseReportService
            deleted_count = BaseReportService.cleanup_expired_cache()
            return Response(
                {
                    "message": f"Cleaned up {deleted_count} expired cache entries",
                    "deleted_count": deleted_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.error(f"Cache cleanup failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to clean cache", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class ReportExecutionViewSet(ReadOnlyBaseViewSet):
    """
    Report execution history (read-only)
    """

    queryset = ReportExecution.objects.all()
    serializer_class = ReportExecutionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["status", "saved_report__report_type"]
    search_fields = ["saved_report__name"]
    ordering_fields = ["started_at", "completed_at"]
    ordering = ["-started_at"]

    def get_queryset(self):
        """Filter by user - users can only see their own execution history"""
        queryset = super().get_queryset()

        # Staff can see all executions, regular users only their own
        if self.request.user.is_staff:
            return queryset

        return queryset.filter(saved_report__user=self.request.user)

# Legacy view for backward compatibility
class SalesSummaryViewSet(viewsets.ViewSet):
    """
    Legacy sales summary view for backward compatibility
    """

    permission_classes = [IsAuthenticated]

    def list(self, request):
        """Generate basic sales summary (legacy format)"""
        try:
            # Use the new ReportViewSet internally
            report_viewset = ReportViewSet()
            report_viewset.request = request

            # Call the summary action
            response = report_viewset.summary(request)

            if response.status_code == 200:
                # Transform new format to legacy format for backward compatibility
                data = response.data
                legacy_data = {
                    "start_date": data.get("date_range", {}).get("start", ""),
                    "end_date": data.get("date_range", {}).get("end", ""),
                    "gross_sales": data.get("total_sales", 0),
                    "net_sales": data.get("total_sales", 0) - data.get("total_tax", 0),
                    "total_discounts": data.get("total_discounts", 0),
                    "total_tax": data.get("total_tax", 0),
                    "total_orders": data.get("total_transactions", 0),
                }
                return Response(legacy_data, status=status.HTTP_200_OK)
            else:
                return response

        except Exception as e:
            logger.error(f"Legacy sales summary failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to generate sales summary", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class BulkExportViewSet(viewsets.ViewSet):
    """
    Advanced bulk export operations for Phase 3
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["post"], url_path="create")
    def create_bulk_export(self, request):
        """
        Create a bulk export operation for multiple reports.

        Body example:
        {
            "report_configs": [
                {
                    "type": "sales",
                    "start_date": "2024-01-01",
                    "end_date": "2024-01-31",
                    "filters": {}
                },
                {
                    "type": "products",
                    "start_date": "2024-01-01",
                    "end_date": "2024-01-31",
                    "filters": {"category_id": 1}
                }
            ],
            "export_format": "xlsx",
            "compress": true,
            "priority": 2
        }
        """
        serializer = BulkExportRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Create the bulk export asynchronously
            task_result = create_bulk_export_async.delay(
                user_id=request.user.id,
                report_configs=serializer.validated_data["report_configs"],
                export_format=serializer.validated_data["export_format"],
                compress=serializer.validated_data["compress"],
                priority=serializer.validated_data["priority"],
            )

            # Return operation details
            return Response(
                {
                    "message": "Bulk export operation created successfully",
                    "task_id": task_result.id,
                    "status": "queued",
                    "estimated_time": serializer.validated_data.get(
                        "estimated_time", "unknown"
                    ),
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            logger.error(f"Bulk export creation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to create bulk export", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="status/(?P<operation_id>[^/.]+)")
    def get_export_status(self, request, operation_id=None):
        """Get the status of a bulk export operation"""
        try:
            status_data = AdvancedExportService.get_export_status(operation_id)
            serializer = BulkExportStatusSerializer(data=status_data)

            if serializer.is_valid():
                return Response(serializer.validated_data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"error": "Invalid status data", "detail": serializer.errors},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as e:
            logger.error(f"Export status retrieval failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to get export status", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="queue-status")
    def get_queue_status(self, request):
        """Get the current status of the export queue (admin only)"""
        if not request.user.is_staff:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        try:
            queue_status = ExportQueue.get_queue_status()
            serializer = ExportQueueStatusSerializer(data=queue_status)

            if serializer.is_valid():
                return Response(serializer.validated_data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"error": "Invalid queue status data", "detail": serializer.errors},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as e:
            logger.error(f"Queue status retrieval failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to get queue status", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="process-queue")
    def process_queue(self, request):
        """Manually trigger export queue processing (admin only)"""
        if not request.user.is_staff:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        try:
            # Trigger the queue processing task
            task_result = process_export_queue.delay()

            return Response(
                {
                    "message": "Export queue processing triggered",
                    "task_id": task_result.id,
                    "status": "processing",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            logger.error(f"Queue processing trigger failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to process queue", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="templates")
    def create_template(self, request):
        """Create a custom export template"""
        serializer = CustomTemplateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            template_config = {
                "branding": serializer.validated_data.get("branding", {}),
                "formatting": serializer.validated_data.get("formatting", {}),
                "layout": serializer.validated_data.get("layout", {}),
            }

            template_id = AdvancedExportService.create_custom_template(
                template_name=serializer.validated_data["template_name"],
                template_config=template_config,
                user=request.user,
            )

            return Response(
                {
                    "message": "Custom template created successfully",
                    "template_id": template_id,
                    "template_name": serializer.validated_data["template_name"],
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            logger.error(f"Template creation failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to create template", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["delete"], url_path="cleanup")
    def cleanup_old_exports(self, request):
        """Clean up old export files (admin only)"""
        if not request.user.is_staff:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        try:
            days_to_keep = int(request.query_params.get("days", 7))
            cleaned_count = AdvancedExportService.cleanup_old_exports(days_to_keep)

            return Response(
                {
                    "message": f"Cleaned up {cleaned_count} old export files",
                    "cleaned_count": cleaned_count,
                    "days_kept": days_to_keep,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.error(f"Export cleanup failed: {e}", exc_info=True)
            return Response(
                {"error": "Failed to cleanup exports", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
