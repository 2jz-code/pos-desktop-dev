import time
import hashlib
import json
import logging
import csv
import io
from decimal import Decimal
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional

from django.db import transaction
from django.db.models import (
    Sum,
    Count,
    Avg,
    F,
    Q,
    Max,
    Min,
    ExpressionWrapper,
    DecimalField,
    Value,
)
from django.db.models.functions import (
    TruncDate,
    TruncHour,
    TruncWeek,
    TruncMonth,
    Extract,
    Coalesce,
)
from django.utils import timezone
from django.conf import settings
import pytz
from django.core.cache import cache
from core_backend.infrastructure.cache_utils import (
    cache_static_data,
    cache_dynamic_data,
    cache_session_data,
)

# Export functionality imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from orders.models import Order, OrderItem
from payments.models import Payment, PaymentTransaction
from products.models import Product, Category
from users.models import User
from .models import ReportCache, SavedReport, ReportExecution

logger = logging.getLogger(__name__)


class ReportService:
    """Central service for all report operations following existing patterns"""

    # Cache TTL in hours for different report types
    CACHE_TTL = {
        "summary": 1,
        "sales": 2,
        "products": 4,
        "payments": 2,
        "operations": 1,
    }

    # Phase 3C: Advanced caching methods

    @staticmethod
    @cache_static_data(timeout=3600 * 8)  # 8 hours - business metrics change slowly
    def get_cached_business_kpis():
        """Cache core business KPIs that don't change frequently"""
        try:
            # Calculate key business metrics
            thirty_days_ago = timezone.now() - timedelta(days=30)
            seven_days_ago = timezone.now() - timedelta(days=7)
            today = timezone.now()

            # Monthly performance
            monthly_orders = Order.objects.filter(
                created_at__gte=thirty_days_ago, status=Order.OrderStatus.COMPLETED
            )

            monthly_revenue = monthly_orders.aggregate(total=Sum("grand_total"))[
                "total"
            ] or Decimal("0.00")

            # Weekly performance
            weekly_orders = Order.objects.filter(
                created_at__gte=seven_days_ago, status=Order.OrderStatus.COMPLETED
            )

            weekly_revenue = weekly_orders.aggregate(total=Sum("grand_total"))[
                "total"
            ] or Decimal("0.00")

            # Average order value
            avg_order_value = monthly_orders.aggregate(avg=Avg("grand_total"))[
                "avg"
            ] or Decimal("0.00")

            # Top products (last 30 days)
            top_products = (
                OrderItem.objects.filter(
                    order__created_at__gte=thirty_days_ago,
                    order__status=Order.OrderStatus.COMPLETED,
                )
                .values("product_id", "product__name")
                .annotate(
                    total_quantity=Sum("quantity"),
                    total_revenue=Sum(F("quantity") * F("price_at_sale")),
                )
                .order_by("-total_revenue")[:10]
            )

            return {
                "monthly_revenue": float(monthly_revenue),
                "weekly_revenue": float(weekly_revenue),
                "monthly_order_count": monthly_orders.count(),
                "weekly_order_count": weekly_orders.count(),
                "average_order_value": float(avg_order_value),
                "top_products": list(top_products),
                "period": {
                    "start_date": thirty_days_ago.isoformat(),
                    "end_date": today.isoformat(),
                },
                "last_updated": timezone.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Failed to generate business KPIs: {e}")
            return {
                "error": f"Failed to generate KPIs: {str(e)}",
                "monthly_revenue": 0,
                "weekly_revenue": 0,
                "monthly_order_count": 0,
                "weekly_order_count": 0,
                "average_order_value": 0,
                "top_products": [],
            }

    @staticmethod
    @cache_dynamic_data(timeout=1800)  # 30 minutes - current day changes frequently
    def get_real_time_sales_summary():
        """Cache current day sales summary for dashboard"""
        try:
            # Today's performance
            today_start = timezone.now().replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            today_end = timezone.now()

            today_orders = Order.objects.filter(
                created_at__gte=today_start,
                created_at__lte=today_end,
                status=Order.OrderStatus.COMPLETED,
            )

            today_revenue = today_orders.aggregate(total=Sum("grand_total"))[
                "total"
            ] or Decimal("0.00")

            # Hourly breakdown for today
            hourly_sales = (
                today_orders.annotate(hour=Extract("created_at", "hour"))
                .values("hour")
                .annotate(count=Count("id"), revenue=Sum("grand_total"))
                .order_by("hour")
            )

            # Payment method breakdown
            payment_methods = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(
                    payment__order__created_at__gte=today_start,
                    payment__order__status=Order.OrderStatus.COMPLETED,
                    status="completed",
                )
                .values("method")
                .annotate(count=Count("id"), total=Sum("amount"))
                .order_by("-total")
            )

            return {
                "today_revenue": float(today_revenue),
                "today_order_count": today_orders.count(),
                "hourly_breakdown": list(hourly_sales),
                "payment_methods": list(payment_methods),
                "current_hour": today_end.hour,
                "last_updated": timezone.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Failed to generate real-time sales summary: {e}")
            return {
                "error": f"Failed to generate real-time summary: {str(e)}",
                "today_revenue": 0,
                "today_order_count": 0,
                "hourly_breakdown": [],
                "payment_methods": [],
            }

    @staticmethod
    def calculate_net_revenue(subtotal, tips, discounts, refunds=0):
        """
        Calculate net revenue using the correct business formula.
        
        Net Revenue = Subtotal + Tips - Discounts - Refunds
        
        This represents the actual business profit, excluding:
        - Tax (goes to government, not business profit)
        - Surcharges (covers processing fees, not business profit)
        
        Args:
            subtotal (float): Base order subtotal
            tips (float): Customer tips collected
            discounts (float): Discount amounts applied
            refunds (float): Total refunds issued (default: 0)
            
        Returns:
            float: Calculated net revenue
        """
        return subtotal + tips - discounts - refunds

    @staticmethod
    def get_revenue_breakdown(subtotal, tips, discounts, tax, surcharges, refunds=0):
        """
        Get detailed breakdown of revenue components for frontend display.
        
        Returns a dictionary with clear categorization of revenue vs non-revenue items.
        """
        net_revenue = ReportService.calculate_net_revenue(subtotal, tips, discounts, refunds)
        
        return {
            # Revenue components (contribute to business profit)
            "revenue_components": {
                "subtotal": subtotal,
                "tips": tips,
                "discounts_applied": -discounts,  # Negative because it reduces revenue
                "refunds": -refunds,  # Negative because it reduces revenue
                "net_revenue": net_revenue
            },
            # Non-revenue components (informational only)
            "non_revenue_components": {
                "tax": tax,  # Goes to government
                "surcharges": surcharges  # Covers processing fees
            },
            # Total customer payment
            "customer_totals": {
                "grand_total": subtotal + tax + surcharges,
                "total_collected": subtotal + tax + surcharges + tips
            }
        }

    @staticmethod
    @cache_static_data(timeout=3600 * 24)  # 24 hours - historical trends change rarely
    def get_historical_trends_data():
        """Cache monthly/yearly trends that rarely change"""
        try:
            # Last 12 months performance
            twelve_months_ago = timezone.now() - timedelta(days=365)

            monthly_trends = (
                Order.objects.filter(
                    created_at__gte=twelve_months_ago,
                    status=Order.OrderStatus.COMPLETED,
                )
                .annotate(month=TruncMonth("created_at"))
                .values("month")
                .annotate(
                    order_count=Count("id"),
                    revenue=Sum("grand_total"),
                    avg_order_value=Avg("grand_total"),
                )
                .order_by("month")
            )

            # Year-over-year comparison (if we have data)
            this_year = timezone.now().year
            last_year = this_year - 1

            this_year_revenue = Order.objects.filter(
                created_at__year=this_year, status=Order.OrderStatus.COMPLETED
            ).aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")

            last_year_revenue = Order.objects.filter(
                created_at__year=last_year, status=Order.OrderStatus.COMPLETED
            ).aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")

            # Calculate growth rate
            growth_rate = 0
            if last_year_revenue > 0:
                growth_rate = (
                    (this_year_revenue - last_year_revenue) / last_year_revenue
                ) * 100

            return {
                "monthly_trends": list(monthly_trends),
                "year_over_year": {
                    "this_year_revenue": float(this_year_revenue),
                    "last_year_revenue": float(last_year_revenue),
                    "growth_rate": float(growth_rate),
                },
                "trend_period": {
                    "start_date": twelve_months_ago.isoformat(),
                    "end_date": timezone.now().isoformat(),
                },
                "last_updated": timezone.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Failed to generate historical trends: {e}")
            return {
                "error": f"Failed to generate trends: {str(e)}",
                "monthly_trends": [],
                "year_over_year": {
                    "this_year_revenue": 0,
                    "last_year_revenue": 0,
                    "growth_rate": 0,
                },
            }

    @staticmethod
    @cache_dynamic_data(
        timeout=3600 * 2
    )  # 2 hours - payment patterns change moderately
    def get_payment_analytics():
        """Cache payment method analytics and trends"""
        try:
            # Last 30 days payment analytics
            thirty_days_ago = timezone.now() - timedelta(days=30)

            # Payment method performance
            payment_performance = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(created_at__gte=thirty_days_ago, status="completed")
                .values("method")
                .annotate(
                    transaction_count=Count("id"),
                    total_amount=Sum("amount"),
                    avg_amount=Avg("amount"),
                    success_rate=Count("id", filter=Q(status="completed"))
                    * 100.0
                    / Count("id"),
                )
                .order_by("-total_amount")
            )

            # Daily payment trends
            daily_payments = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(created_at__gte=thirty_days_ago, status="completed")
                .annotate(date=TruncDate("created_at"))
                .values("date")
                .annotate(count=Count("id"), total=Sum("amount"))
                .order_by("date")
            )

            # Payment failure analysis
            failed_payments = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(
                    created_at__gte=thirty_days_ago,
                    status__in=["failed", "declined", "cancelled"],
                )
                .values("method", "status")
                .annotate(count=Count("id"))
                .order_by("-count")
            )

            # Average transaction times (if available)
            avg_processing_times = {
                "card": "2.5s",  # Placeholder - would need actual timing data
                "cash": "0.1s",
                "gift_card": "1.2s",
            }

            return {
                "payment_methods": list(payment_performance),
                "daily_trends": list(daily_payments),
                "failed_payments": list(failed_payments),
                "processing_times": avg_processing_times,
                "total_processed": sum(p["total_amount"] for p in payment_performance),
                "period": {
                    "start_date": thirty_days_ago.isoformat(),
                    "end_date": timezone.now().isoformat(),
                },
                "last_updated": timezone.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Failed to generate payment analytics: {e}")
            return {
                "error": f"Failed to generate payment analytics: {str(e)}",
                "payment_methods": [],
                "daily_trends": [],
                "failed_payments": [],
                "processing_times": {},
                "total_processed": 0,
            }

    @staticmethod
    @cache_session_data(
        timeout=900
    )  # 15 minutes - performance monitoring changes frequently
    def get_performance_monitoring_cache():
        """Cache comprehensive system performance metrics"""
        try:
            from core_backend.infrastructure.cache import CacheMonitor

            # Get cache health stats
            cache_health = CacheMonitor.health_check()
            cache_stats = CacheMonitor.get_all_cache_stats()

            # Database performance indicators
            recent_orders = Order.objects.filter(
                created_at__gte=timezone.now() - timedelta(hours=1)
            ).count()

            recent_transactions = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(created_at__gte=timezone.now() - timedelta(hours=1))
                .count()
            )

            # System health indicators
            performance_metrics = {
                "cache_health": cache_health,
                "cache_statistics": cache_stats,
                "recent_activity": {
                    "orders_last_hour": recent_orders,
                    "transactions_last_hour": recent_transactions,
                    "system_load": "normal",  # Placeholder - would integrate with system monitoring
                },
                "response_times": {
                    "database_avg": "15ms",  # Placeholder - would need actual monitoring
                    "cache_avg": "2ms",
                    "api_avg": "45ms",
                },
                "uptime": "99.9%",  # Placeholder - would integrate with uptime monitoring
                "last_updated": timezone.now().isoformat(),
            }

            return performance_metrics

        except Exception as e:
            logger.error(f"Failed to generate performance monitoring data: {e}")
            return {
                "error": f"Failed to generate performance metrics: {str(e)}",
                "cache_health": {},
                "cache_statistics": {},
                "recent_activity": {},
                "response_times": {},
                "uptime": "unknown",
            }

    @staticmethod
    def get_local_timezone():
        """Get the configured local timezone"""
        return pytz.timezone(settings.TIME_ZONE)

    @staticmethod
    def trunc_date_local(field_name):
        """Truncate date field to local timezone instead of UTC"""
        from django.db.models import DateTimeField
        from django.db.models.functions import Cast

        # Convert to local timezone, then truncate to date
        local_tz = ReportService.get_local_timezone()
        return TruncDate(Cast(field_name, DateTimeField()), tzinfo=local_tz)

    @staticmethod
    @transaction.atomic
    def generate_summary_report(
        start_date: datetime, end_date: datetime, use_cache: bool = True
    ) -> Dict[str, Any]:
        """Generate summary report with intelligent caching"""

        # Generate cache key
        cache_key = ReportService._generate_cache_key(
            "summary", {"start_date": start_date, "end_date": end_date}
        )

        # Check cache first
        if use_cache:
            cached_data = ReportService._get_cached_report(cache_key)
            if cached_data:
                logger.info(f"Summary report served from cache: {cache_key[:8]}...")
                return cached_data

        logger.info(f"Generating summary report for {start_date} to {end_date}")
        start_time = time.time()

        # Base queryset with optimization
        orders_queryset = (
            Order.objects.filter(
                status=Order.OrderStatus.COMPLETED,
                created_at__range=(start_date, end_date),
                subtotal__gt=0,  # Exclude orders with $0.00 subtotals
            )
            .select_related("cashier", "customer")
            .prefetch_related("items__product")
        )

        # Core metrics aggregation (WITHOUT items to avoid JOIN duplication)
        summary_data = orders_queryset.aggregate(
            total_sales=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
            total_transactions=Count("id"),
            total_tax=Coalesce(Sum("tax_total"), Value(Decimal("0.00"))),
            total_discounts=Coalesce(
                Sum("total_discounts_amount"), Value(Decimal("0.00"))
            ),
        )

        # Convert Decimal to float for JSON serialization
        summary_data = {
            k: float(v) if isinstance(v, Decimal) else v
            for k, v in summary_data.items()
        }

        # Calculate total_items separately to avoid ORDER duplication from JOINs
        summary_data["total_items"] = orders_queryset.aggregate(
            total_items=Coalesce(Sum("items__quantity"), Value(0))
        )["total_items"]

        # Calculate average ticket
        summary_data["average_ticket"] = (
            summary_data["total_sales"] / summary_data["total_transactions"]
            if summary_data["total_transactions"] > 0
            else 0
        )

        # Growth metrics (compare with previous period)
        previous_period_days = (end_date - start_date).days
        previous_start = start_date - timedelta(days=previous_period_days)
        previous_end = start_date

        previous_data = Order.objects.filter(
            status=Order.OrderStatus.COMPLETED,
            created_at__range=(previous_start, previous_end),
            subtotal__gt=0,  # Exclude orders with $0.00 subtotals
        ).aggregate(
            prev_sales=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
            prev_transactions=Count("id"),
        )

        # Calculate growth percentages
        if previous_data["prev_sales"] > 0:
            summary_data["sales_growth"] = round(
                (
                    (summary_data["total_sales"] - float(previous_data["prev_sales"]))
                    / float(previous_data["prev_sales"])
                )
                * 100,
                2,
            )
        else:
            summary_data["sales_growth"] = 0

        if previous_data["prev_transactions"] > 0:
            summary_data["transaction_growth"] = round(
                (
                    (
                        summary_data["total_transactions"]
                        - previous_data["prev_transactions"]
                    )
                    / previous_data["prev_transactions"]
                )
                * 100,
                2,
            )
        else:
            summary_data["transaction_growth"] = 0

        # Top product calculation (single product by quantity)
        top_product = (
            OrderItem.objects.filter(order__in=orders_queryset)
            .values("product__name")
            .annotate(total_sold=Sum("quantity"))
            .order_by("-total_sold")
            .first()
        )

        summary_data["top_product"] = (
            top_product["product__name"] if top_product else "N/A"
        )

        # Top products by revenue (for charts)
        top_products_by_revenue = (
            OrderItem.objects.filter(order__in=orders_queryset)
            .values("product__name", "product__id")
            .annotate(
                revenue=Sum(F("quantity") * F("price_at_sale")),
                quantity_sold=Sum("quantity"),
            )
            .order_by("-revenue")[:5]
        )

        summary_data["top_products_by_revenue"] = [
            {
                "name": item["product__name"],
                "revenue": float(item["revenue"] or 0),
                "quantity": item["quantity_sold"] or 0,
            }
            for item in top_products_by_revenue
        ]

        # Sales trend data (daily breakdown)
        daily_sales = (
            orders_queryset.annotate(date=ReportService.trunc_date_local("created_at"))
            .values("date")
            .annotate(sales=Sum("grand_total"), transactions=Count("id"))
            .order_by("date")
        )

        summary_data["sales_trend"] = [
            {
                "date": item["date"].strftime("%Y-%m-%d"),
                "sales": float(item["sales"] or 0),
                "transactions": item["transactions"],
            }
            for item in daily_sales
        ]

        # Payment method distribution (match legacy system behavior - amount only)
        payment_methods = (
            PaymentTransaction.objects.filter(
                payment__order__in=orders_queryset,
            )
            .values("method")
            .annotate(
                amount=Sum("amount"),
                count=Count("id"),  # Use amount only to match legacy
            )
            .order_by("-amount")
        )

        total_payment_amount = sum(float(pm["amount"] or 0) for pm in payment_methods)

        summary_data["payment_distribution"] = [
            {
                "method": item["method"],
                "amount": float(item["amount"] or 0),
                "count": item["count"],
                "percentage": (
                    round((float(item["amount"] or 0) / total_payment_amount * 100), 2)
                    if total_payment_amount > 0
                    else 0
                ),
            }
            for item in payment_methods
        ]

        # Hourly performance
        hourly_data = (
            orders_queryset.annotate(hour=Extract("created_at", "hour"))
            .values("hour")
            .annotate(sales=Sum("grand_total"), orders=Count("id"))
            .order_by("hour")
        )

        summary_data["hourly_performance"] = [
            {
                "hour": f"{item['hour']:02d}:00",
                "sales": float(item["sales"] or 0),
                "orders": item["orders"],
            }
            for item in hourly_data
        ]

        # Add metadata
        summary_data["generated_at"] = timezone.now().isoformat()
        summary_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Cache the result
        generation_time = time.time() - start_time
        ReportService._cache_report(
            cache_key, summary_data, ttl_hours=ReportService.CACHE_TTL["summary"]
        )

        logger.info(f"Summary report generated in {generation_time:.2f}s")
        return summary_data

    @staticmethod
    @transaction.atomic
    def generate_sales_report(
        start_date: datetime,
        end_date: datetime,
        group_by: str = "day",
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate detailed sales report"""
        
        # Import PaymentTransaction at function scope to avoid scoping issues
        from payments.models import PaymentTransaction

        cache_key = ReportService._generate_cache_key(
            "sales",
            {"start_date": start_date, "end_date": end_date, "group_by": group_by},
        )

        if use_cache:
            cached_data = ReportService._get_cached_report(cache_key)
            if cached_data:
                return cached_data

        logger.info(f"Generating sales report for {start_date} to {end_date}")
        start_time = time.time()

        # Base queryset with optimization
        orders_queryset = (
            Order.objects.filter(
                status=Order.OrderStatus.COMPLETED,
                created_at__range=(start_date, end_date),
                subtotal__gt=0,  # Exclude orders with $0.00 subtotals
            )
            .select_related("cashier", "customer", "payment_details")
            .prefetch_related("items__product")
        )

        # Calculate total refunds separately
        total_refunds = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__in=orders_queryset,
                status=PaymentTransaction.TransactionStatus.REFUNDED,
            )
            .aggregate(
                total_refunds=Coalesce(Sum("refunded_amount"), Value(Decimal("0.00")))
            )["total_refunds"]
        )

        # Core sales metrics (WITHOUT items to avoid JOIN duplication)
        sales_data = orders_queryset.aggregate(
            total_revenue=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
            total_subtotal=Coalesce(Sum("subtotal"), Value(Decimal("0.00"))),
            total_orders=Count("id"),
            avg_order_value=Coalesce(Avg("grand_total"), Value(Decimal("0.00"))),
            total_tax=Coalesce(Sum("tax_total"), Value(Decimal("0.00"))),
            total_discounts=Coalesce(
                Sum("total_discounts_amount"), Value(Decimal("0.00"))
            ),
        )

        # Calculate payment totals separately to avoid JOIN duplication
        from payments.models import Payment

        payment_totals = Payment.objects.filter(order__in=orders_queryset).aggregate(
            total_surcharges=Coalesce(Sum("total_surcharges"), Value(Decimal("0.00"))),
            total_tips=Coalesce(Sum("total_tips"), Value(Decimal("0.00"))),
        )

        # Add payment totals to sales_data
        sales_data.update(payment_totals)

        # Convert Decimal to float
        sales_data = {
            k: float(v) if isinstance(v, Decimal) else v for k, v in sales_data.items()
        }

        # Add refunds to the sales data
        sales_data["total_refunds"] = float(total_refunds)
        
        # Calculate net revenue using the centralized method
        sales_data["net_revenue"] = ReportService.calculate_net_revenue(
            sales_data["total_subtotal"],
            sales_data["total_tips"],
            sales_data["total_discounts"],
            sales_data["total_refunds"]
        )
        
        # Add detailed revenue breakdown for frontend
        sales_data["revenue_breakdown"] = ReportService.get_revenue_breakdown(
            sales_data["total_subtotal"],
            sales_data["total_tips"],
            sales_data["total_discounts"],
            sales_data["total_tax"],
            sales_data["total_surcharges"],
            sales_data["total_refunds"]
        )

        # Calculate total_items separately to avoid ORDER duplication from JOINs
        sales_data["total_items"] = orders_queryset.aggregate(
            total_items=Coalesce(Sum("items__quantity"), Value(0))
        )["total_items"]

        # Sales by period (daily, weekly, or monthly breakdown)
        if group_by == "week":
            trunc_period = TruncWeek("created_at")
        elif group_by == "month":
            trunc_period = TruncMonth("created_at")
        else:
            trunc_period = TruncDate("created_at")

        # Step 1: Aggregate revenue and orders. This is correct as it doesn't involve joins that cause duplication.
        sales_agg = (
            orders_queryset.annotate(period=trunc_period)
            .values("period")
            .annotate(
                revenue=Sum("grand_total"),
                orders=Count("id"),
            )
            .order_by("period")
        )

        # Step 2: Aggregate item quantities separately to avoid inflating the sales/order counts.
        # This query introduces a JOIN on OrderItem, so it's done independently.
        items_agg = (
            orders_queryset.annotate(period=trunc_period)
            .values("period")
            .annotate(items=Sum("items__quantity"))
            .order_by("period")
        )

        # Step 3: Merge the two aggregations in memory for the final result.
        items_dict = {item["period"]: item["items"] for item in items_agg}

        # Get detailed transaction data for each period
        transaction_details_by_period = {}
        from calendar import monthrange

        for period_item in sales_agg:
            period_start = period_item["period"]

            # Create timezone-aware datetime objects for the period range
            local_tz = ReportService.get_local_timezone()
            start_dt = timezone.make_aware(
                datetime.combine(period_start, datetime.min.time()), local_tz
            )

            if group_by == "week":
                end_dt = start_dt + timedelta(days=7)
            elif group_by == "month":
                days_in_month = monthrange(start_dt.year, start_dt.month)[1]
                end_dt = start_dt + timedelta(days=days_in_month)
            else:  # day
                end_dt = start_dt + timedelta(days=1)

            # Get transactions for this period using the precise datetime range
            period_transactions = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(
                    payment__order__in=orders_queryset,
                    payment__created_at__gte=start_dt,
                    payment__created_at__lt=end_dt,
                    status=PaymentTransaction.TransactionStatus.SUCCESSFUL,
                )
                .select_related("payment", "payment__order")
                .order_by("-payment__created_at")
            )

            # Get payment totals for this period
            period_payments = Payment.objects.filter(
                order__in=orders_queryset,
                created_at__gte=start_dt,
                created_at__lt=end_dt,
            ).aggregate(
                total_tips=Coalesce(Sum("total_tips"), Value(Decimal("0.00"))),
                total_surcharges=Coalesce(
                    Sum("total_surcharges"), Value(Decimal("0.00"))
                ),
                total_collected=Coalesce(
                    Sum("total_collected"), Value(Decimal("0.00"))
                ),
            )

            # Group transactions by payment method
            method_breakdown = {}
            for transaction in period_transactions:
                method = transaction.method
                if method not in method_breakdown:
                    method_breakdown[method] = {
                        "count": 0,
                        "total_amount": 0,
                        "total_tips": 0,
                        "total_surcharges": 0,
                    }
                method_breakdown[method]["count"] += 1
                method_breakdown[method]["total_amount"] += float(
                    transaction.amount or 0
                )
                method_breakdown[method]["total_tips"] += float(transaction.tip or 0)
                method_breakdown[method]["total_surcharges"] += float(
                    transaction.surcharge or 0
                )

            transaction_details_by_period[period_item["period"]] = {
                "transactions": [
                    {
                        "order_number": trans.payment.order.order_number,
                        "created_at": trans.payment.order.created_at.isoformat(),
                        "amount": float(trans.amount or 0),
                        "tip": float(trans.tip or 0),
                        "surcharge": float(trans.surcharge or 0),
                        "method": trans.method,
                        "transaction_id": trans.transaction_id,
                        "card_brand": trans.card_brand,
                        "card_last4": trans.card_last4,
                    }
                    for trans in period_transactions
                ],
                "payment_totals": {
                    "total_tips": float(period_payments["total_tips"]),
                    "total_surcharges": float(period_payments["total_surcharges"]),
                    "total_collected": float(period_payments["total_collected"]),
                },
                "method_breakdown": method_breakdown,
            }

        sales_data["sales_by_period"] = [
            {
                "date": item["period"].strftime("%Y-%m-%d"),
                "revenue": float(item["revenue"] or 0),
                "orders": item["orders"],
                "items": items_dict.get(item["period"], 0) or 0,
                "transaction_details": transaction_details_by_period.get(
                    item["period"],
                    {
                        "transactions": [],
                        "payment_totals": {
                            "total_tips": 0,
                            "total_surcharges": 0,
                            "total_collected": 0,
                        },
                        "method_breakdown": {},
                    },
                ),
            }
            for item in sales_agg
        ]

        # Sales by category
        category_sales = (
            OrderItem.objects.filter(order__in=orders_queryset)
            .values("product__category__name")
            .annotate(
                revenue=Sum(F("quantity") * F("price_at_sale")),
                quantity=Sum("quantity"),
            )
            .order_by("-revenue")
        )

        sales_data["sales_by_category"] = [
            {
                "category": item["product__category__name"] or "Uncategorized",
                "revenue": float(item["revenue"] or 0),
                "quantity": item["quantity"] or 0,
            }
            for item in category_sales
        ]

        # Top performing hours
        hourly_sales = (
            orders_queryset.annotate(hour=Extract("created_at", "hour"))
            .values("hour")
            .annotate(revenue=Sum("grand_total"), orders=Count("id"))
            .order_by("-revenue")[:10]
        )

        sales_data["top_hours"] = [
            {
                "hour": f"{item['hour']:02d}:00",
                "revenue": float(item["revenue"] or 0),
                "orders": item["orders"],
            }
            for item in hourly_sales
        ]

        # Add metadata
        # Add comprehensive order vs payment reconciliation
        # Calculate what customers actually paid (matching payment processing approach)
        
        # Get all payment transactions for completed, canceled, AND voided orders (matching payments report)
        all_order_transactions = PaymentTransaction.objects.filter(
            payment__order__created_at__range=(start_date, end_date),
            payment__order__status__in=[Order.OrderStatus.COMPLETED, Order.OrderStatus.CANCELLED, Order.OrderStatus.VOID],
            payment__order__subtotal__gt=0
        )
        
        # Debug: Log transaction counts for diagnosis
        total_transactions_debug = all_order_transactions.count()
        logger.info(f"=== SALES REPORT TRANSACTION DEBUG ===")
        logger.info(f"Total transactions found: {total_transactions_debug}")
        logger.info(f"Date range: {start_date} to {end_date}")
        
        payment_reconciliation = all_order_transactions.aggregate(
            successful_payments=Coalesce(
                Sum("amount", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)),
                Value(Decimal("0.00"))
            ),
            successful_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)),
            refunded_payments=Coalesce(
                Sum("amount", filter=Q(status=PaymentTransaction.TransactionStatus.REFUNDED)),
                Value(Decimal("0.00"))
            ),
            refunded_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.REFUNDED)),
            failed_payments=Coalesce(
                Sum("amount", filter=Q(status=PaymentTransaction.TransactionStatus.FAILED)),
                Value(Decimal("0.00"))
            ),
            failed_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.FAILED)),
            canceled_payments=Coalesce(
                Sum("amount", filter=Q(status=PaymentTransaction.TransactionStatus.CANCELED)),
                Value(Decimal("0.00"))
            ),
            canceled_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.CANCELED))
        )
        
        # Debug: Log individual payment status totals
        logger.info(f"Successful payments: ${payment_reconciliation['successful_payments']} (count: {payment_reconciliation['successful_count']})")
        logger.info(f"Refunded payments: ${payment_reconciliation['refunded_payments']} (count: {payment_reconciliation['refunded_count']})")
        logger.info(f"Failed payments: ${payment_reconciliation['failed_payments']} (count: {payment_reconciliation['failed_count']})")
        logger.info(f"Canceled payments: ${payment_reconciliation['canceled_payments']} (count: {payment_reconciliation['canceled_count']})")
        logger.info(f"=== END TRANSACTION DEBUG ===")
        
        # Calculate voided orders total separately
        voided_orders_total = Order.objects.filter(
            status=Order.OrderStatus.VOID,
            created_at__range=(start_date, end_date),
            subtotal__gt=0
        ).aggregate(
            total=Coalesce(Sum("grand_total"), Value(Decimal("0.00")))
        )["total"]
        
        # Calculate comprehensive totals
        total_orders_value = float(sales_data["total_revenue"])  # What customers ordered ($16,782.72)
        voided_orders_value = float(voided_orders_total or 0)
        total_payment_attempts = float(
            payment_reconciliation["successful_payments"] + 
            payment_reconciliation["refunded_payments"] + 
            payment_reconciliation["failed_payments"] + 
            payment_reconciliation["canceled_payments"]
        )  # What was attempted to be processed ($16,771.24)
        
        sales_data["order_vs_payment_reconciliation"] = {
            "total_orders_value": total_orders_value,
            "total_payment_attempts": total_payment_attempts,
            "successfully_processed": float(payment_reconciliation["successful_payments"]),
            "voided_orders_value": voided_orders_value,
            "lost_revenue": total_orders_value - total_payment_attempts,
            "order_completion_rate": round((total_payment_attempts / total_orders_value * 100), 2) if total_orders_value > 0 else 0,
            "payment_breakdown": {
                "successful": {
                    "amount": float(payment_reconciliation["successful_payments"]),
                    "count": payment_reconciliation["successful_count"]
                },
                "refunded": {
                    "amount": float(payment_reconciliation["refunded_payments"]),
                    "count": payment_reconciliation["refunded_count"]
                },
                "failed": {
                    "amount": float(payment_reconciliation["failed_payments"]),
                    "count": payment_reconciliation["failed_count"]
                },
                "canceled": {
                    "amount": float(payment_reconciliation["canceled_payments"]),
                    "count": payment_reconciliation["canceled_count"]
                }
            }
        }

        sales_data["generated_at"] = timezone.now().isoformat()
        sales_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Cache the result
        generation_time = time.time() - start_time
        ReportService._cache_report(
            cache_key, sales_data, ttl_hours=ReportService.CACHE_TTL["sales"]
        )

        logger.info(f"Sales report generated in {generation_time:.2f}s")
        return sales_data

    @staticmethod
    @transaction.atomic
    def generate_products_report(
        start_date: datetime,
        end_date: datetime,
        category_id: Optional[int] = None,
        limit: int = 10,
        trend_period: str = "auto",  # "daily", "weekly", "monthly", "auto"
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate product performance report"""

        cache_key = ReportService._generate_cache_key(
            "products",
            {
                "start_date": start_date,
                "end_date": end_date,
                "category_id": category_id,
                "limit": limit,
                "trend_period": trend_period,
            },
        )

        if use_cache:
            cached_data = ReportService._get_cached_report(cache_key)
            if cached_data:
                return cached_data

        logger.info(f"Generating products report for {start_date} to {end_date}")
        start_time = time.time()

        # Base queryset
        order_items = OrderItem.objects.filter(
            order__status=Order.OrderStatus.COMPLETED,
            order__created_at__range=(start_date, end_date),
            order__subtotal__gt=0,  # Exclude orders with $0.00 subtotals
        ).select_related("product", "product__category")

        # Filter by category if specified
        if category_id:
            order_items = order_items.filter(product__category_id=category_id)

        # Top products by revenue
        top_products = (
            order_items.annotate(revenue=F("quantity") * F("price_at_sale"))
            .values("product__name", "product__id")
            .annotate(
                total_revenue=Sum("revenue"),
                total_sold=Sum("quantity"),
                avg_price=Avg("price_at_sale"),
            )
            .order_by("-total_revenue")[:limit]
        )

        products_data = {
            "top_products": [
                {
                    "name": item["product__name"],
                    "id": item["product__id"],
                    "revenue": float(item["total_revenue"] or 0),
                    "sold": item["total_sold"],
                    "avg_price": float(item["avg_price"] or 0),
                }
                for item in top_products
            ]
        }

        # Best sellers by quantity
        best_sellers = (
            order_items.values("product__name", "product__id")
            .annotate(
                total_sold=Sum("quantity"),
                total_revenue=Sum(F("quantity") * F("price_at_sale")),
            )
            .order_by("-total_sold")[:limit]
        )

        products_data["best_sellers"] = [
            {
                "name": item["product__name"],
                "id": item["product__id"],
                "sold": item["total_sold"],
                "revenue": float(item["total_revenue"] or 0),
            }
            for item in best_sellers
        ]

        # Category performance
        category_performance = (
            order_items.values("product__category__name")
            .annotate(
                revenue=Sum(F("quantity") * F("price_at_sale")),
                units_sold=Sum("quantity"),
                unique_products=Count("product__id", distinct=True),
            )
            .order_by("-revenue")
        )

        products_data["category_performance"] = [
            {
                "category": item["product__category__name"] or "Uncategorized",
                "revenue": float(item["revenue"] or 0),
                "units_sold": item["units_sold"] or 0,
                "unique_products": item["unique_products"],
            }
            for item in category_performance
        ]

        # Product trends with proper period handling
        date_diff = (end_date - start_date).days

        # Determine trend period
        if trend_period == "auto":
            if date_diff <= 7:
                actual_period = "daily"
            elif date_diff <= 60:
                actual_period = "weekly"
            else:
                actual_period = "monthly"
        else:
            actual_period = trend_period

        # Apply correct grouping based on period
        if actual_period == "weekly":
            from django.db.models.functions import TruncWeek

            product_trends = (
                order_items.filter(
                    product__in=[p["product__id"] for p in top_products[:5]]
                )
                .annotate(period=TruncWeek("order__created_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )
        elif actual_period == "monthly":
            from django.db.models.functions import TruncMonth

            product_trends = (
                order_items.filter(
                    product__in=[p["product__id"] for p in top_products[:5]]
                )
                .annotate(period=TruncMonth("order__created_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )
        else:  # daily
            product_trends = (
                order_items.filter(
                    product__in=[p["product__id"] for p in top_products[:5]]
                )
                .annotate(period=ReportService.trunc_date_local("order__created_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )

        # Group trends by product
        trends_by_product = {}
        for trend in product_trends:
            product_name = trend["product__name"]
            if product_name not in trends_by_product:
                trends_by_product[product_name] = []

            trends_by_product[product_name].append(
                {"date": trend["period"].strftime("%Y-%m-%d"), "sold": trend["sold"]}
            )

        products_data["product_trends"] = trends_by_product

        # Summary stats
        products_data["summary"] = {
            "total_products": order_items.values("product__id").distinct().count(),
            "total_revenue": float(
                order_items.aggregate(total=Sum(F("quantity") * F("price_at_sale")))[
                    "total"
                ]
                or 0
            ),
            "total_units_sold": order_items.aggregate(total=Sum("quantity"))["total"]
            or 0,
        }

        # Add metadata
        products_data["generated_at"] = timezone.now().isoformat()
        products_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }
        products_data["filters"] = {
            "category_id": category_id,
            "limit": limit,
            "trend_period": trend_period,
            "actual_period": actual_period,
        }

        # Cache the result
        generation_time = time.time() - start_time
        ReportService._cache_report(
            cache_key, products_data, ttl_hours=ReportService.CACHE_TTL["products"]
        )

        logger.info(f"Products report generated in {generation_time:.2f}s")
        return products_data

    @staticmethod
    @transaction.atomic
    def generate_payments_report(
        start_date: datetime, end_date: datetime, use_cache: bool = True
    ) -> Dict[str, Any]:
        """Generate payment methods report"""

        cache_key = ReportService._generate_cache_key(
            "payments", {"start_date": start_date, "end_date": end_date}
        )

        if use_cache:
            cached_data = ReportService._get_cached_report(cache_key)
            if cached_data:
                return cached_data

        logger.info(f"Generating payments report for {start_date} to {end_date}")
        start_time = time.time()

        # Base queryset for successful transactions only
        successful_transactions = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__status=Order.OrderStatus.COMPLETED,
                payment__order__created_at__range=(start_date, end_date),
                payment__order__subtotal__gt=0,
                status=PaymentTransaction.TransactionStatus.SUCCESSFUL,
            )
            .select_related("payment", "payment__order")
        )

        # Base queryset for refunded transactions (include both completed and canceled orders)
        refunded_transactions = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__status__in=[Order.OrderStatus.COMPLETED, Order.OrderStatus.CANCELLED],
                payment__order__created_at__range=(start_date, end_date),
                payment__order__subtotal__gt=0,
                status=PaymentTransaction.TransactionStatus.REFUNDED,
            )
            .select_related("payment", "payment__order")
        )

        # Track failed and canceled transactions for complete transparency
        failed_transactions = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__status=Order.OrderStatus.COMPLETED,
                payment__order__created_at__range=(start_date, end_date),
                payment__order__subtotal__gt=0,
                status=PaymentTransaction.TransactionStatus.FAILED,
            )
            .select_related("payment", "payment__order")
        )

        canceled_transactions = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__status=Order.OrderStatus.COMPLETED,
                payment__order__created_at__range=(start_date, end_date),
                payment__order__subtotal__gt=0,
                status=PaymentTransaction.TransactionStatus.CANCELED,
            )
            .select_related("payment", "payment__order")
        )

        # Get payments for the filtered orders
        payments = Payment.objects.filter(
            order__status=Order.OrderStatus.COMPLETED,
            order__created_at__range=(start_date, end_date),
            order__subtotal__gt=0,
        )

        # Get payment method breakdown from SUCCESSFUL transactions only
        payment_methods_agg = (
            successful_transactions.values("method")
            .annotate(
                amount=Sum("amount"),
                count=Count("id"),
                processing_fees=Sum("surcharge"),
            )
            .order_by("-amount")
        )

        # Get refunded amounts by payment method for display
        refunded_methods_agg = refunded_transactions.values("method").annotate(
            refunded_amount=Sum("amount"),
            refunded_count=Count("id"),
        )

        # Create a lookup dict for refunded amounts
        refunded_by_method = {
            item["method"]: {
                "amount": float(item["refunded_amount"] or 0),
                "count": item["refunded_count"],
            }
            for item in refunded_methods_agg
        }

        # Calculate totals from successful transactions
        total_processed_from_transactions = sum(
            float(item["amount"] or 0) for item in payment_methods_agg
        )
        total_fees_from_transactions = sum(
            float(item["processing_fees"] or 0) for item in payment_methods_agg
        )

        # Get authoritative totals from the Payment model for validation
        payment_totals = payments.aggregate(
            total_collected=Coalesce(Sum("total_collected"), Value(Decimal("0.00"))),
            total_surcharges=Coalesce(Sum("total_surcharges"), Value(Decimal("0.00"))),
        )
        payment_total_collected = float(payment_totals["total_collected"])
        payment_total_fees = float(payment_totals["total_surcharges"])

        # Calculate trends by comparing with the previous period's transaction data
        previous_period_days = (end_date - start_date).days
        previous_start = start_date - timedelta(days=previous_period_days)
        previous_end = start_date

        previous_successful_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status=Order.OrderStatus.COMPLETED,
            payment__order__created_at__range=(previous_start, previous_end),
            payment__order__subtotal__gt=0,
            status=PaymentTransaction.TransactionStatus.SUCCESSFUL,
        )

        # Get the previous period's amounts per method for trend calculation
        previous_amounts = {
            item["method"]: float(item["amount"] or 0)
            for item in previous_successful_transactions.values("method").annotate(
                amount=Sum("amount")
            )
        }

        # Build the final payment_methods list using successful transaction amounts
        final_payment_methods = []
        for item in payment_methods_agg:
            method_amount = float(item["amount"] or 0)
            method_fees = float(item["processing_fees"] or 0)
            method_name = item["method"]

            # Get refunded amounts for this method
            refunded_info = refunded_by_method.get(
                method_name, {"amount": 0, "count": 0}
            )
            refunded_amount = refunded_info["amount"]
            refunded_count = refunded_info["count"]

            # Calculate total processed (successful + refunded) for this method
            total_method_processed = method_amount + refunded_amount

            # Calculate percentage based on successful transaction totals
            amount_percentage = (
                (method_amount / total_processed_from_transactions * 100)
                if total_processed_from_transactions > 0
                else 0
            )

            # Trend calculation
            previous_amount = previous_amounts.get(method_name, 0)
            trend = (
                round(
                    ((method_amount - previous_amount) / (previous_amount or 1)) * 100,
                    2,
                )
                if previous_amount > 0
                else 0
            )

            final_payment_methods.append(
                {
                    "method": method_name,
                    "amount": method_amount,  # Successful transactions only
                    "count": item["count"],
                    "avg_amount": (
                        method_amount / item["count"] if item["count"] > 0 else 0
                    ),
                    "processing_fees": method_fees,
                    "percentage": round(amount_percentage, 2),
                    "trend": trend,
                    # New fields for complete picture
                    "refunded_amount": refunded_amount,
                    "refunded_count": refunded_count,
                    "total_processed": total_method_processed,  # For reconciliation
                    "net_amount": method_amount,  # Successful amount (refunds already excluded)
                }
            )

        payments_data = {"payment_methods": final_payment_methods}

        # Daily payment volume (use Payment model for accuracy)
        daily_payments = (
            payments.annotate(date=ReportService.trunc_date_local("order__created_at"))
            .values("date")
            .annotate(amount=Sum("total_collected"), count=Count("id"))
            .order_by("date")
        )

        payments_data["daily_volume"] = [
            {
                "date": item["date"].strftime("%Y-%m-%d"),
                "amount": float(item["amount"] or 0),
                "count": item["count"],
            }
            for item in daily_payments
        ]

        # Daily breakdown by payment method (successful transactions only)
        daily_breakdown_data = (
            successful_transactions.annotate(
                date=ReportService.trunc_date_local("created_at")
            )
            .values("date", "method")
            .annotate(total=Sum("amount"))
            .order_by("date", "method")
        )

        daily_breakdown = {}
        for item in daily_breakdown_data:
            date_str = item["date"].strftime("%Y-%m-%d")
            if date_str not in daily_breakdown:
                daily_breakdown[date_str] = {"date": date_str, "total": 0}
            method_key = item["method"].lower().replace("_", "")
            daily_breakdown[date_str][method_key] = float(item["total"] or 0)

        for date_str, breakdown in daily_breakdown.items():
            breakdown["total"] = sum(
                v for k, v in breakdown.items() if k not in ["date", "total"]
            )

        payments_data["daily_breakdown"] = sorted(
            list(daily_breakdown.values()), key=lambda x: x["date"]
        )

        # Calculate comprehensive totals first (before using them)
        total_refunds = float(
            refunded_transactions.aggregate(total=Sum("amount"))["total"] or 0
        )

        # Calculate failed and canceled transaction totals for complete transparency
        total_failed = float(
            failed_transactions.aggregate(total=Sum("amount"))["total"] or 0
        )

        total_canceled = float(
            canceled_transactions.aggregate(total=Sum("amount"))["total"] or 0
        )

        # Calculate total attempted (ALL transactions regardless of status)
        total_attempted = (
            total_processed_from_transactions
            + total_refunds
            + total_failed
            + total_canceled
        )

        # Calculate processing issues (non-successful transactions)
        total_processing_issues = total_refunds + total_failed + total_canceled

        # Temporary diagnostic logging to compare with diagnostic script
        logger.info(f"=== PAYMENT TOTALS BREAKDOWN ===")
        logger.info(f"Successful transactions: ${total_processed_from_transactions}")
        logger.info(f"Refunded transactions: ${total_refunds}")
        logger.info(f"Failed transactions: ${total_failed}")
        logger.info(f"Canceled transactions: ${total_canceled}")
        logger.info(f"TOTAL ATTEMPTED: ${total_attempted}")
        logger.info(f"Expected from diagnostic: $16771.24")
        logger.info(f"Difference: ${16771.24 - total_attempted}")
        logger.info(f"=== END BREAKDOWN ===")

        # Order totals for comparison
        completed_orders = Order.objects.filter(
            status=Order.OrderStatus.COMPLETED,
            created_at__range=(start_date, end_date),
            subtotal__gt=0,
        ).aggregate(
            order_total=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
            order_count=Count("id"),
        )

        payments_data["order_totals_comparison"] = {
            "order_grand_total": float(completed_orders["order_total"] or 0),
            "order_count": completed_orders["order_count"],
            "payment_transaction_total": total_attempted,  # Use total attempted (all transactions)
            "difference": float(completed_orders["order_total"] or 0)
            - total_attempted,  # Compare against total attempted amount
        }

        payments_data["summary"] = {
            # Primary metrics showing complete payment processing picture
            "total_attempted": total_attempted,  # ALL transactions regardless of status
            "successfully_processed": total_processed_from_transactions,  # Successful only
            "processing_issues": total_processing_issues,  # Failed + Canceled + Refunded
            # Detailed breakdown by status
            "breakdown": {
                "successful": {
                    "amount": total_processed_from_transactions,
                    "count": successful_transactions.count(),
                },
                "refunded": {
                    "amount": total_refunds,
                    "count": refunded_transactions.count(),
                },
                "failed": {
                    "amount": total_failed,
                    "count": failed_transactions.count(),
                },
                "canceled": {
                    "amount": total_canceled,
                    "count": canceled_transactions.count(),
                },
            },
            # Legacy fields for backward compatibility
            "total_processed": total_processed_from_transactions,  # Successful transactions only
            "total_transactions": successful_transactions.count(),
            "total_refunds": total_refunds,
            "total_refunded_transactions": refunded_transactions.count(),
            "net_revenue": total_processed_from_transactions,  # Note: This is payment-focused net revenue (successful transactions)
            # Calculated rates
            "processing_success_rate": (
                round((total_processed_from_transactions / total_attempted * 100), 2)
                if total_attempted > 0
                else 0
            ),
            "processing_issues_rate": (
                round((total_processing_issues / total_attempted * 100), 2)
                if total_attempted > 0
                else 0
            ),
        }

        # Processing statistics (use all transactions for complete picture)
        all_transactions_for_stats = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status=Order.OrderStatus.COMPLETED,
            payment__order__created_at__range=(start_date, end_date),
            payment__order__subtotal__gt=0,
        )

        processing_stats = all_transactions_for_stats.aggregate(
            total_attempts=Count("id"),
            successful=Count(
                "id", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)
            ),
            failed=Count(
                "id", filter=Q(status=PaymentTransaction.TransactionStatus.FAILED)
            ),
            refunded=Count(
                "id", filter=Q(status=PaymentTransaction.TransactionStatus.REFUNDED)
            ),
        )

        success_rate = (
            (processing_stats["successful"] / processing_stats["total_attempts"]) * 100
            if processing_stats["total_attempts"] > 0
            else 0
        )

        payments_data["processing_stats"] = {
            "total_attempts": processing_stats["total_attempts"],
            "successful": processing_stats["successful"],
            "failed": processing_stats["failed"],
            "refunded": processing_stats["refunded"],
            "success_rate": round(success_rate, 2),
        }

        # Add metadata
        payments_data["generated_at"] = timezone.now().isoformat()
        payments_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Cache the result
        generation_time = time.time() - start_time
        ReportService._cache_report(
            cache_key, payments_data, ttl_hours=ReportService.CACHE_TTL["payments"]
        )

        logger.info(f"Payments report generated in {generation_time:.2f}s")
        return payments_data

    @staticmethod
    @transaction.atomic
    def generate_operations_report(
        start_date: datetime, end_date: datetime, use_cache: bool = True
    ) -> Dict[str, Any]:
        """Generate operations report"""

        cache_key = ReportService._generate_cache_key(
            "operations", {"start_date": start_date, "end_date": end_date}
        )

        if use_cache:
            cached_data = ReportService._get_cached_report(cache_key)
            if cached_data:
                return cached_data

        logger.info(f"Generating operations report for {start_date} to {end_date}")
        start_time = time.time()

        # Base queryset
        orders = Order.objects.filter(
            status=Order.OrderStatus.COMPLETED,
            created_at__range=(start_date, end_date),
            subtotal__gt=0,  # Exclude orders with $0.00 subtotals
        ).select_related("cashier")

        # Hourly patterns
        hourly_patterns = (
            orders.annotate(hour=Extract("created_at", "hour"))
            .values("hour")
            .annotate(
                orders=Count("id"),
                revenue=Sum("grand_total"),
                avg_order_value=Avg("grand_total"),
            )
            .order_by("hour")
        )

        operations_data = {
            "hourly_patterns": [
                {
                    "hour": f"{item['hour']:02d}:00",
                    "orders": item["orders"],
                    "revenue": float(item["revenue"] or 0),
                    "avg_order_value": float(item["avg_order_value"] or 0),
                }
                for item in hourly_patterns
            ]
        }

        # Peak hours (top 5 by order volume)
        peak_hours = list(hourly_patterns.order_by("-orders")[:5])
        operations_data["peak_hours"] = [
            {
                "hour": f"{item['hour']:02d}:00",
                "orders": item["orders"],
                "revenue": float(item["revenue"] or 0),
            }
            for item in peak_hours
        ]

        # Daily order volume
        daily_volume = (
            orders.annotate(date=ReportService.trunc_date_local("created_at"))
            .values("date")
            .annotate(orders=Count("id"), revenue=Sum("grand_total"))
            .order_by("date")
        )

        operations_data["daily_volume"] = [
            {
                "date": item["date"].strftime("%Y-%m-%d"),
                "orders": item["orders"],
                "revenue": float(item["revenue"] or 0),
            }
            for item in daily_volume
        ]

        # Staff performance (if cashier data available)
        staff_performance = (
            orders.filter(cashier__isnull=False)
            .values("cashier__username", "cashier__first_name", "cashier__last_name")
            .annotate(
                orders_processed=Count("id"),
                total_revenue=Sum("grand_total"),
                avg_order_value=Avg("grand_total"),
            )
            .order_by("-orders_processed")
        )

        operations_data["staff_performance"] = [
            {
                "cashier": f"{item['cashier__first_name'] or ''} {item['cashier__last_name'] or ''}".strip()
                or item["cashier__username"],
                "orders_processed": item["orders_processed"],
                "revenue": float(item["total_revenue"] or 0),
                "avg_order_value": float(item["avg_order_value"] or 0),
            }
            for item in staff_performance
        ]

        # Order volume summary
        total_orders = orders.count()
        total_days = (end_date - start_date).days + 1

        operations_data["summary"] = {
            "total_orders": total_orders,
            "avg_orders_per_day": (
                round(total_orders / total_days, 2) if total_days > 0 else 0
            ),
            "peak_day": None,
            "slowest_day": None,
        }

        # Find peak and slowest days
        if daily_volume:
            peak_day = max(daily_volume, key=lambda x: x["orders"])
            slowest_day = min(daily_volume, key=lambda x: x["orders"])

            operations_data["summary"]["peak_day"] = {
                "date": peak_day["date"],
                "orders": peak_day["orders"],
            }
            operations_data["summary"]["slowest_day"] = {
                "date": slowest_day["date"],
                "orders": slowest_day["orders"],
            }

        # Add metadata
        operations_data["generated_at"] = timezone.now().isoformat()
        operations_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Cache the result
        generation_time = time.time() - start_time
        ReportService._cache_report(
            cache_key, operations_data, ttl_hours=ReportService.CACHE_TTL["operations"]
        )

        logger.info(f"Operations report generated in {generation_time:.2f}s")
        return operations_data

    @staticmethod
    def _generate_cache_key(report_type: str, parameters: Dict[str, Any]) -> str:
        """Generate consistent cache keys"""
        # Sort parameters for consistent hashing
        sorted_params = json.dumps(parameters, sort_keys=True, default=str)
        hash_input = f"{report_type}:{sorted_params}"
        return hashlib.sha256(hash_input.encode()).hexdigest()

    @staticmethod
    def _get_cached_report(cache_key: str) -> Optional[Dict[str, Any]]:
        """Retrieve cached report if valid"""
        try:
            cached = ReportCache.objects.get(
                parameters_hash=cache_key, expires_at__gt=timezone.now()
            )
            return cached.data
        except ReportCache.DoesNotExist:
            return None

    @staticmethod
    def _cache_report(cache_key: str, data: Dict[str, Any], ttl_hours: int = 1) -> None:
        """Cache report data with TTL"""
        expires_at = timezone.now() + timedelta(hours=ttl_hours)

        # Extract report type from data or use default
        report_type = data.get("report_type", "unknown")

        # Get the parameters from cache key (simplified)
        parameters = {"cached_at": timezone.now().isoformat()}

        try:
            ReportCache.objects.update_or_create(
                parameters_hash=cache_key,
                defaults={
                    "report_type": report_type,
                    "parameters": parameters,
                    "data": data,
                    "expires_at": expires_at,
                    "generated_at": timezone.now(),
                },
            )
            logger.info(f"Report cached: {cache_key[:8]}... (expires in {ttl_hours}h)")
        except Exception as e:
            logger.error(f"Failed to cache report: {e}")

    @staticmethod
    def cleanup_expired_cache() -> int:
        """Remove expired cache entries"""
        deleted_count = ReportCache.objects.filter(
            expires_at__lt=timezone.now()
        ).delete()[0]

        if deleted_count > 0:
            logger.info(f"Cleaned up {deleted_count} expired cache entries")

        return deleted_count

    @staticmethod
    def invalidate_cache_for_report_type(report_type: str) -> int:
        """Invalidate all cache entries for a specific report type"""
        deleted_count = ReportCache.objects.filter(report_type=report_type).delete()[0]

        if deleted_count > 0:
            logger.info(f"Invalidated {deleted_count} cache entries for {report_type}")

        return deleted_count

    @staticmethod
    def get_cache_stats() -> Dict[str, Any]:
        """Get cache statistics"""
        total_entries = ReportCache.objects.count()
        expired_entries = ReportCache.objects.filter(
            expires_at__lt=timezone.now()
        ).count()

        cache_by_type = (
            ReportCache.objects.values("report_type")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        return {
            "total_entries": total_entries,
            "expired_entries": expired_entries,
            "valid_entries": total_entries - expired_entries,
            "cache_by_type": list(cache_by_type),
        }

    @staticmethod
    @transaction.atomic
    def get_quick_metrics() -> Dict[str, Any]:
        """Get today/MTD/YTD metrics for dashboard"""
        
        # Import PaymentTransaction at function scope to avoid scoping issues
        from payments.models import PaymentTransaction

        # Get local timezone for date calculations
        local_tz = ReportService.get_local_timezone()
        now = timezone.now().astimezone(local_tz)

        # Calculate date ranges
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = now

        # Month to date
        mtd_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mtd_end = now

        # Year to date
        ytd_start = now.replace(
            month=1, day=1, hour=0, minute=0, second=0, microsecond=0
        )
        ytd_end = now

        # Base queryset for completed orders
        def get_metrics_for_period(start_time, end_time):
            # Get order-level metrics WITHOUT JOINs to avoid duplication
            orders_queryset_inner = Order.objects.filter(
                status=Order.OrderStatus.COMPLETED,
                created_at__range=(start_time, end_time),
                subtotal__gt=0,  # Exclude orders with $0.00 subtotals
            )
            
            orders = orders_queryset_inner.aggregate(
                total_sales=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
                total_subtotal=Coalesce(Sum("subtotal"), Value(Decimal("0.00"))),
                total_discounts=Coalesce(Sum("total_discounts_amount"), Value(Decimal("0.00"))),
                total_orders=Count("id"),
            )
            
            # Get payment totals for tips
            from payments.models import Payment
            payment_totals = Payment.objects.filter(order__in=orders_queryset_inner).aggregate(
                total_tips=Coalesce(Sum("total_tips"), Value(Decimal("0.00"))),
            )

            # Calculate total items separately to avoid JOIN duplication
            total_items = OrderItem.objects.filter(
                order__status=Order.OrderStatus.COMPLETED,
                order__created_at__range=(start_time, end_time),
                order__subtotal__gt=0,  # Exclude orders with $0.00 subtotals
            ).aggregate(total_items=Coalesce(Sum("quantity"), Value(0)))["total_items"]

            # Add comprehensive payment reconciliation (matching sales report approach)
            
            orders_queryset = Order.objects.filter(
                status=Order.OrderStatus.COMPLETED,
                created_at__range=(start_time, end_time),
                subtotal__gt=0,
            )
            
            all_order_transactions = PaymentTransaction.objects.filter(
                payment__order__in=orders_queryset
            ).aggregate(
                successful_payments=Coalesce(
                    Sum("amount", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)),
                    Value(Decimal("0.00"))
                ),
                total_attempted_payments=Coalesce(Sum("amount"), Value(Decimal("0.00")))
            )

            total_orders_value = float(orders["total_sales"] or 0)  # What customers ordered
            total_payment_attempts = float(all_order_transactions["total_attempted_payments"] or 0)  # What was processed
            successful_payments = float(all_order_transactions["successful_payments"] or 0)
            
            # Calculate net revenue using the centralized method
            net_revenue = ReportService.calculate_net_revenue(
                float(orders["total_subtotal"] or 0),
                float(payment_totals["total_tips"] or 0),
                float(orders["total_discounts"] or 0),
                0  # No refunds calculation in quick metrics for performance
            )

            return {
                "sales": total_payment_attempts,  # Keep for backward compatibility
                "gross_revenue": successful_payments,  # Total successfully processed payments
                "net_revenue": net_revenue,  # New proper net revenue calculation
                "orders": orders["total_orders"] or 0,
                "items": total_items or 0,
                "avg_order_value": (
                    total_orders_value / (orders["total_orders"] or 1)
                    if orders["total_orders"] > 0
                    else 0
                ),
                # Revenue breakdown components
                "subtotal": float(orders["total_subtotal"] or 0),
                "tips": float(payment_totals["total_tips"] or 0),
                "discounts": float(orders["total_discounts"] or 0),
                # Existing comprehensive metrics
                "total_orders_value": total_orders_value,
                "total_payment_attempts": total_payment_attempts,
                "successful_payments": successful_payments,
                "lost_revenue": total_orders_value - total_payment_attempts,
                "order_completion_rate": round((total_payment_attempts / total_orders_value * 100), 2) if total_orders_value > 0 else 0,
            }

        # Get metrics for each period
        today_metrics = get_metrics_for_period(today_start, today_end)
        mtd_metrics = get_metrics_for_period(mtd_start, mtd_end)
        ytd_metrics = get_metrics_for_period(ytd_start, ytd_end)

        return {
            "today": {
                **today_metrics,
                "date_range": {
                    "start": today_start.isoformat(),
                    "end": today_end.isoformat(),
                },
            },
            "month_to_date": {
                **mtd_metrics,
                "date_range": {
                    "start": mtd_start.isoformat(),
                    "end": mtd_end.isoformat(),
                },
            },
            "year_to_date": {
                **ytd_metrics,
                "date_range": {
                    "start": ytd_start.isoformat(),
                    "end": ytd_end.isoformat(),
                },
            },
            "generated_at": timezone.now().isoformat(),
        }

    # Export functionality methods
    @staticmethod
    def export_to_csv(report_data: Dict[str, Any], report_type: str) -> bytes:
        """Export report data to CSV format"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header with report info
            writer.writerow([f"{report_type.title()} Report"])
            writer.writerow([f"Generated: {report_data.get('generated_at', 'N/A')}"])
            writer.writerow(
                [
                    f"Date Range: {report_data.get('date_range', {}).get('start', 'N/A')} to {report_data.get('date_range', {}).get('end', 'N/A')}"
                ]
            )
            writer.writerow([])  # Empty row

            # Export based on report type
            if report_type == "summary":
                ReportService._export_summary_to_csv(writer, report_data)
            elif report_type == "sales":
                ReportService._export_sales_to_csv(writer, report_data)
            elif report_type == "products":
                ReportService._export_products_to_csv(writer, report_data)
            elif report_type == "payments":
                ReportService._export_payments_to_csv(writer, report_data)
            elif report_type == "operations":
                ReportService._export_operations_to_csv(writer, report_data)

            return output.getvalue().encode("utf-8")

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    @staticmethod
    def export_to_xlsx(report_data: Dict[str, Any], report_type: str) -> bytes:
        """Export report data to Excel format"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{report_type.title()} Report"

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Report header
            ws.merge_cells("A1:D1")
            ws["A1"] = f"{report_type.title()} Report"
            ws["A1"].font = Font(bold=True, size=16)
            ws["A1"].alignment = header_alignment

            ws["A2"] = f"Generated: {report_data.get('generated_at', 'N/A')}"
            ws["A3"] = (
                f"Date Range: {report_data.get('date_range', {}).get('start', 'N/A')} to {report_data.get('date_range', {}).get('end', 'N/A')}"
            )

            # Export based on report type
            if report_type == "summary":
                ReportService._export_summary_to_xlsx(
                    ws, report_data, header_font, header_fill, header_alignment
                )
            elif report_type == "sales":
                ReportService._export_sales_to_xlsx(
                    ws, report_data, header_font, header_fill, header_alignment
                )
            elif report_type == "products":
                ReportService._export_products_to_xlsx(
                    ws, report_data, header_font, header_fill, header_alignment
                )
            elif report_type == "payments":
                ReportService._export_payments_to_xlsx(
                    ws, report_data, header_font, header_fill, header_alignment
                )
            elif report_type == "operations":
                ReportService._export_operations_to_xlsx(
                    ws, report_data, header_font, header_fill, header_alignment
                )

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise

    @staticmethod
    def export_to_pdf(
        report_data: Dict[str, Any], report_type: str, report_title: str = None
    ) -> bytes:
        """Export report data to PDF format"""
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(
                output,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18,
            )

            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
            )

            # Build story
            story = []

            # Title
            title = report_title or f"{report_type.title()} Report"
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))

            # Report info
            story.append(
                Paragraph(
                    f"Generated: {report_data.get('generated_at', 'N/A')}",
                    styles["Normal"],
                )
            )
            story.append(
                Paragraph(
                    f"Date Range: {report_data.get('date_range', {}).get('start', 'N/A')} to {report_data.get('date_range', {}).get('end', 'N/A')}",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 20))

            # Export based on report type
            if report_type == "summary":
                ReportService._export_summary_to_pdf(story, report_data, styles)
            elif report_type == "sales":
                ReportService._export_sales_to_pdf(story, report_data, styles)
            elif report_type == "products":
                ReportService._export_products_to_pdf(story, report_data, styles)
            elif report_type == "payments":
                ReportService._export_payments_to_pdf(story, report_data, styles)
            elif report_type == "operations":
                ReportService._export_operations_to_pdf(story, report_data, styles)

            # Build PDF
            doc.build(story)
            output.seek(0)
            return output.read()

        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            raise

    # CSV export helpers
    @staticmethod
    def _export_summary_to_csv(writer, report_data: Dict[str, Any]):
        """Export summary report to CSV"""
        writer.writerow(["Summary Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Sales", f"${report_data.get('total_sales', 0):,.2f}"])
        writer.writerow(
            ["Total Transactions", report_data.get("total_transactions", 0)]
        )
        writer.writerow(
            ["Average Ticket", f"${report_data.get('average_ticket', 0):,.2f}"]
        )
        writer.writerow(["Total Tax", f"${report_data.get('total_tax', 0):,.2f}"])
        writer.writerow(
            ["Total Discounts", f"${report_data.get('total_discounts', 0):,.2f}"]
        )
        writer.writerow(["Sales Growth", f"{report_data.get('sales_growth', 0):+.2f}%"])
        writer.writerow(
            ["Transaction Growth", f"{report_data.get('transaction_growth', 0):+.2f}%"]
        )
        writer.writerow(["Top Product", report_data.get("top_product", "N/A")])
        writer.writerow([])

        # Sales trend
        writer.writerow(["Daily Sales Trend"])
        writer.writerow(["Date", "Sales", "Transactions"])
        for item in report_data.get("sales_trend", []):
            writer.writerow(
                [
                    item.get("date"),
                    f"${item.get('sales', 0):,.2f}",
                    item.get("transactions", 0),
                ]
            )
        writer.writerow([])

        # Payment distribution
        writer.writerow(["Payment Method Distribution"])
        writer.writerow(["Method", "Amount", "Count", "Percentage"])
        for item in report_data.get("payment_distribution", []):
            writer.writerow(
                [
                    item.get("method"),
                    f"${item.get('amount', 0):,.2f}",
                    item.get("count", 0),
                    f"{item.get('percentage', 0):.2f}%",
                ]
            )

    @staticmethod
    def _export_sales_to_csv(writer, report_data: Dict[str, Any]):
        """Export sales report to CSV"""
        writer.writerow(["Sales Summary"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(
            ["Total Revenue", f"${report_data.get('total_revenue', 0):,.2f}"]
        )
        writer.writerow(["Total Orders", report_data.get("total_orders", 0)])
        writer.writerow(
            [
                "Average Order Value",
                f"${report_data.get('average_order_value', 0):,.2f}",
            ]
        )
        writer.writerow([])

        # Daily trends
        writer.writerow(["Daily Sales Trends"])
        writer.writerow(["Date", "Revenue", "Orders", "AOV"])
        for item in report_data.get("daily_trends", []):
            writer.writerow(
                [
                    item.get("date"),
                    f"${item.get('revenue', 0):,.2f}",
                    item.get("orders", 0),
                    f"${item.get('aov', 0):,.2f}",
                ]
            )

    @staticmethod
    def _export_products_to_csv(writer, report_data: Dict[str, Any]):
        """Export products report to CSV"""
        writer.writerow(["Product Performance"])
        writer.writerow(["Product", "Quantity Sold", "Revenue", "Average Price"])
        for item in report_data.get("top_products", []):
            writer.writerow(
                [
                    item.get("name"),
                    item.get("quantity_sold", 0),
                    f"${item.get('revenue', 0):,.2f}",
                    f"${item.get('avg_price', 0):,.2f}",
                ]
            )

    @staticmethod
    def _export_payments_to_csv(writer, report_data: Dict[str, Any]):
        """Export payments report to CSV"""
        writer.writerow(["Payment Method Performance"])
        writer.writerow(["Method", "Amount", "Count", "Percentage"])
        for item in report_data.get("payment_methods", []):
            writer.writerow(
                [
                    item.get("method"),
                    f"${item.get('amount', 0):,.2f}",
                    item.get("count", 0),
                    f"{item.get('percentage', 0):.2f}%",
                ]
            )

    @staticmethod
    def _export_operations_to_csv(writer, report_data: Dict[str, Any]):
        """Export operations report to CSV"""
        writer.writerow(["Hourly Performance"])
        writer.writerow(["Hour", "Orders", "Revenue", "Avg Order Value"])
        for item in report_data.get("hourly_patterns", []):
            writer.writerow(
                [
                    item.get("hour"),
                    item.get("orders", 0),
                    f"${item.get('revenue', 0):,.2f}",
                    f"${item.get('avg_order_value', 0):,.2f}",
                ]
            )

    # Excel export helpers
    @staticmethod
    def _export_summary_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export summary report to Excel"""
        row = 5

        # Summary metrics
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Summary Metrics"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = header_alignment
        row += 1

        metrics = [
            ("Total Sales", f"${report_data.get('total_sales', 0):,.2f}"),
            ("Total Transactions", report_data.get("total_transactions", 0)),
            ("Average Ticket", f"${report_data.get('average_ticket', 0):,.2f}"),
            ("Total Tax", f"${report_data.get('total_tax', 0):,.2f}"),
            ("Total Discounts", f"${report_data.get('total_discounts', 0):,.2f}"),
            ("Sales Growth", f"{report_data.get('sales_growth', 0):+.2f}%"),
            ("Transaction Growth", f"{report_data.get('transaction_growth', 0):+.2f}%"),
            ("Top Product", report_data.get("top_product", "N/A")),
        ]

        for metric, value in metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            row += 1

    @staticmethod
    def _export_sales_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export sales report to Excel"""
        row = 5

        # Sales summary
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Sales Summary"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = header_alignment
        row += 1

        metrics = [
            ("Total Revenue", f"${report_data.get('total_revenue', 0):,.2f}"),
            ("Total Orders", report_data.get("total_orders", 0)),
            (
                "Average Order Value",
                f"${report_data.get('average_order_value', 0):,.2f}",
            ),
        ]

        for metric, value in metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            row += 1

    @staticmethod
    def _export_products_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export products report to Excel"""
        row = 5

        # Product performance
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Product Performance"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = header_alignment
        row += 1

        # Headers
        headers = ["Product", "Quantity Sold", "Revenue", "Average Price"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = header_alignment
        row += 1

        # Data
        for item in report_data.get("top_products", []):
            ws.cell(row=row, column=1, value=item.get("name"))
            ws.cell(row=row, column=2, value=item.get("quantity_sold", 0))
            ws.cell(row=row, column=3, value=f"${item.get('revenue', 0):,.2f}")
            ws.cell(row=row, column=4, value=f"${item.get('avg_price', 0):,.2f}")
            row += 1

    @staticmethod
    def _export_payments_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export payments report to Excel"""
        row = 5

        # Payment methods
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Payment Method Performance"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = header_alignment
        row += 1

        # Headers
        headers = ["Method", "Amount", "Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = header_alignment
        row += 1

        # Data
        for item in report_data.get("payment_methods", []):
            ws.cell(row=row, column=1, value=item.get("method"))
            ws.cell(row=row, column=2, value=f"${item.get('amount', 0):,.2f}")
            ws.cell(row=row, column=3, value=item.get("count", 0))
            ws.cell(row=row, column=4, value=f"{item.get('percentage', 0):.2f}%")
            row += 1

    @staticmethod
    def _export_operations_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export operations report to Excel"""
        row = 5

        # Hourly performance
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Hourly Performance"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = header_alignment
        row += 1

        # Headers
        headers = ["Hour", "Orders", "Revenue", "Avg Order Value"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = header_alignment
        row += 1

        # Data
        for item in report_data.get("hourly_patterns", []):
            ws.cell(row=row, column=1, value=item.get("hour"))
            ws.cell(row=row, column=2, value=item.get("orders", 0))
            ws.cell(row=row, column=3, value=f"${item.get('revenue', 0):,.2f}")
            ws.cell(row=row, column=4, value=f"${item.get('avg_order_value', 0):,.2f}")
            row += 1

    # PDF export helpers
    @staticmethod
    def _export_summary_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export summary report to PDF"""
        # Summary metrics table
        story.append(Paragraph("Summary Metrics", styles["Heading2"]))

        summary_data = [
            ["Metric", "Value"],
            ["Total Sales", f"${report_data.get('total_sales', 0):,.2f}"],
            ["Total Transactions", str(report_data.get("total_transactions", 0))],
            ["Average Ticket", f"${report_data.get('average_ticket', 0):,.2f}"],
            ["Total Tax", f"${report_data.get('total_tax', 0):,.2f}"],
            ["Total Discounts", f"${report_data.get('total_discounts', 0):,.2f}"],
            ["Sales Growth", f"{report_data.get('sales_growth', 0):+.2f}%"],
            ["Transaction Growth", f"{report_data.get('transaction_growth', 0):+.2f}%"],
            ["Top Product", report_data.get("top_product", "N/A")],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(summary_table)
        story.append(Spacer(1, 20))

    @staticmethod
    def _export_sales_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export sales report to PDF"""
        story.append(Paragraph("Sales Summary", styles["Heading2"]))

        sales_data = [
            ["Metric", "Value"],
            ["Total Revenue", f"${report_data.get('total_revenue', 0):,.2f}"],
            ["Total Orders", str(report_data.get("total_orders", 0))],
            [
                "Average Order Value",
                f"${report_data.get('average_order_value', 0):,.2f}",
            ],
        ]

        sales_table = Table(sales_data, colWidths=[3 * inch, 2 * inch])
        sales_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(sales_table)

    @staticmethod
    def _export_products_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export products report to PDF"""
        story.append(Paragraph("Product Performance", styles["Heading2"]))

        product_data = [["Product", "Quantity Sold", "Revenue", "Avg Price"]]
        for item in report_data.get("top_products", []):
            product_data.append(
                [
                    item.get("name"),
                    str(item.get("quantity_sold", 0)),
                    f"${item.get('revenue', 0):,.2f}",
                    f"${item.get('avg_price', 0):,.2f}",
                ]
            )

        product_table = Table(
            product_data, colWidths=[2 * inch, 1 * inch, 1 * inch, 1 * inch]
        )
        product_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(product_table)

    @staticmethod
    def _export_payments_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export payments report to PDF"""
        story.append(Paragraph("Payment Method Performance", styles["Heading2"]))

        payment_data = [["Method", "Amount", "Count", "Percentage"]]
        for item in report_data.get("payment_methods", []):
            payment_data.append(
                [
                    item.get("method"),
                    f"${item.get('amount', 0):,.2f}",
                    str(item.get("count", 0)),
                    f"{item.get('percentage', 0):.2f}%",
                ]
            )

        payment_table = Table(
            payment_data, colWidths=[1.5 * inch, 1.5 * inch, 1 * inch, 1 * inch]
        )
        payment_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(payment_table)

    @staticmethod
    def _export_operations_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export operations report to PDF"""
        story.append(Paragraph("Hourly Performance", styles["Heading2"]))

        hourly_data = [["Hour", "Orders", "Revenue", "Avg Order Value"]]
        for item in report_data.get("hourly_patterns", []):
            hourly_data.append(
                [
                    item.get("hour"),
                    str(item.get("orders", 0)),
                    f"${item.get('revenue', 0):,.2f}",
                    f"${item.get('avg_order_value', 0):,.2f}",
                ]
            )

        hourly_table = Table(
            hourly_data, colWidths=[1 * inch, 1 * inch, 1.5 * inch, 1.5 * inch]
        )
        hourly_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(hourly_table)
