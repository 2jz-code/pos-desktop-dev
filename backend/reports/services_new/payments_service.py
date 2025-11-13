"""
Payments report service with generation and export functionality.
Refactored from monolithic reports service for better maintainability.
"""
import time
import logging
import csv
import io
from decimal import Decimal
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

from django.db import transaction
from django.db.models import (
    Sum, Count, Q, Value, F
)
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone

# Export functionality imports
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Table, TableStyle, Spacer, PageBreak
from reportlab.lib import colors

from orders.models import Order
from payments.models import Payment, PaymentTransaction
from .base import BaseReportService
from .timezone_utils import TimezoneUtils

logger = logging.getLogger(__name__)


class PaymentsReportService(BaseReportService):
    """Service for generating and exporting payments reports."""

    @staticmethod
    @transaction.atomic
    def generate_payments_report(
        tenant,
        start_date: datetime,
        end_date: datetime,
        location_id: Optional[int] = None,
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate comprehensive payments report"""

        cache_key = PaymentsReportService._generate_cache_key(
            "payments",
            {"start_date": start_date, "end_date": end_date, "location_id": location_id},
        )

        if use_cache:
            cached_data = PaymentsReportService._get_cached_report(cache_key, tenant)
            if cached_data:
                return cached_data

        logger.info(f"Generating payments report for {start_date} to {end_date}" + (f" at location {location_id}" if location_id else ""))
        start_time = time.time()

        # Multi-location report generation
        if location_id is None:
            return PaymentsReportService._generate_multi_location_payments_report(
                tenant, start_date, end_date, use_cache
            )

        # Get base transaction querysets
        transaction_querysets = PaymentsReportService._get_transaction_querysets(tenant, start_date, end_date, location_id)

        # Calculate payment method breakdown
        payment_methods = PaymentsReportService._calculate_payment_methods(transaction_querysets)

        # Get daily volume data
        daily_volume = PaymentsReportService._calculate_daily_volume(tenant, start_date, end_date, location_id)

        # Get daily breakdown by method
        daily_breakdown = PaymentsReportService._calculate_daily_breakdown(transaction_querysets['successful'])

        # Calculate comprehensive summary
        summary = PaymentsReportService._calculate_payments_summary(transaction_querysets, start_date, end_date, location_id)

        # Get processing statistics
        processing_stats = PaymentsReportService._calculate_processing_stats(tenant, start_date, end_date, location_id)

        # Get reconciliation data
        order_totals_comparison = PaymentsReportService._calculate_order_reconciliation(tenant, start_date, end_date, location_id, summary)

        # Build final report data
        payments_data = {
            "payment_methods": payment_methods,
            "daily_volume": daily_volume,
            "daily_breakdown": daily_breakdown,
            "summary": summary,
            "processing_stats": processing_stats,
            "order_totals_comparison": order_totals_comparison,
            "generated_at": timezone.now().isoformat(),
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
        }

        # Add location metadata
        location_name = "All Locations"
        if location_id is not None:
            from settings.models import StoreLocation
            try:
                location = StoreLocation.objects.get(id=location_id, tenant=tenant)
                location_name = location.name
            except StoreLocation.DoesNotExist:
                location_name = f"Location ID {location_id}"

        payments_data["location_info"] = {
            "location_id": location_id,
            "location_name": location_name,
            "is_multi_location": location_id is None
        }

        # Add tenant_id for export filtering
        payments_data["tenant_id"] = tenant.id

        # Cache the result
        generation_time = time.time() - start_time
        PaymentsReportService._cache_report(
            cache_key, payments_data, tenant, report_type="payments", ttl_hours=PaymentsReportService.CACHE_TTL["payments"]
        )

        logger.info(f"Payments report generated in {generation_time:.2f}s")
        return payments_data

    @staticmethod
    def _get_transaction_querysets(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None) -> Dict[str, Any]:
        """Get base transaction querysets for different statuses."""
        base_filter = {
            "payment__order__tenant": tenant,
            "payment__order__created_at__range": (start_date, end_date),
            "payment__order__subtotal__gt": 0,
        }

        if location_id is not None:
            base_filter["payment__order__store_location_id"] = location_id

        successful_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status=Order.OrderStatus.COMPLETED,
            status=PaymentTransaction.TransactionStatus.SUCCESSFUL,
            **base_filter
        )

        refunded_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status__in=[Order.OrderStatus.COMPLETED, Order.OrderStatus.CANCELLED],
            status=PaymentTransaction.TransactionStatus.REFUNDED,
            **base_filter
        )

        failed_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status=Order.OrderStatus.COMPLETED,
            status=PaymentTransaction.TransactionStatus.FAILED,
            **base_filter
        )

        canceled_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(
            payment__order__status=Order.OrderStatus.COMPLETED,
            status=PaymentTransaction.TransactionStatus.CANCELED,
            **base_filter
        )

        return {
            "successful": successful_transactions,
            "refunded": refunded_transactions,
            "failed": failed_transactions,
            "canceled": canceled_transactions,
        }

    @staticmethod
    def _calculate_payment_methods(transaction_querysets: Dict[str, Any]) -> list:
        """Calculate payment method breakdown with trends."""
        successful_transactions = transaction_querysets['successful']
        refunded_transactions = transaction_querysets['refunded']
        
        # Get payment method aggregation from successful transactions
        # Include tips and surcharges to match total_collected calculation
        payment_methods_agg = (
            successful_transactions.values("method")
            .annotate(
                base_amount=Sum("amount"),
                tips=Sum("tip"),
                surcharges=Sum("surcharge"),
                amount=Sum(F("amount") + F("tip") + F("surcharge")),
                count=Count("id"),
                processing_fees=Sum("surcharge"),
            )
            .order_by("-amount")
        )

        # Get refunded amounts by method
        # Include tips and surcharges to match total_collected calculation
        refunded_methods_agg = refunded_transactions.values("method").annotate(
            refunded_amount=Sum(F("amount") + F("tip") + F("surcharge")),
            refunded_count=Count("id"),
        )

        # Create lookup dict for refunded amounts
        refunded_by_method = {
            item["method"]: {
                "amount": float(item["refunded_amount"] or 0),
                "count": item["refunded_count"],
            }
            for item in refunded_methods_agg
        }

        # Calculate total for percentages
        total_processed = sum(float(item["amount"] or 0) for item in payment_methods_agg)

        # Calculate trends (simplified for now)
        payment_methods = []
        for item in payment_methods_agg:
            method_amount = float(item["amount"] or 0)
            method_fees = float(item["processing_fees"] or 0)
            method_name = item["method"]

            # Get refunded amounts for this method
            refunded_info = refunded_by_method.get(method_name, {"amount": 0, "count": 0})
            refunded_amount = refunded_info["amount"]
            refunded_count = refunded_info["count"]

            # Calculate percentage
            percentage = (method_amount / total_processed * 100) if total_processed > 0 else 0

            payment_methods.append({
                "method": method_name,
                "amount": method_amount,
                "count": item["count"],
                "avg_amount": method_amount / item["count"] if item["count"] > 0 else 0,
                "processing_fees": method_fees,
                "percentage": round(percentage, 2),
                "trend": 0,  # TODO: Calculate actual trend
                "refunded_amount": refunded_amount,
                "refunded_count": refunded_count,
                "total_processed": method_amount + refunded_amount,
                "net_amount": method_amount,
                # Add detailed breakdown for transparency
                "base_amount": float(item["base_amount"] or 0),
                "tips_total": float(item["tips"] or 0), 
                "surcharges_total": float(item["surcharges"] or 0),
            })

        return payment_methods

    @staticmethod
    def _calculate_daily_volume(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None) -> list:
        """Calculate daily payment volume using Payment model."""
        filters = {
            "order__tenant": tenant,
            "order__status": Order.OrderStatus.COMPLETED,
            "order__created_at__range": (start_date, end_date),
            "order__subtotal__gt": 0,
        }

        if location_id is not None:
            filters["order__store_location_id"] = location_id

        payments = Payment.objects.filter(**filters)

        daily_payments = (
            payments.annotate(date=TruncDate("order__created_at"))
            .values("date")
            .annotate(amount=Sum("total_collected"), count=Count("id"))
            .order_by("date")
        )

        return [
            {
                "date": item["date"].strftime("%Y-%m-%d"),
                "amount": float(item["amount"] or 0),
                "count": item["count"],
            }
            for item in daily_payments
        ]

    @staticmethod
    def _calculate_daily_breakdown(successful_transactions) -> list:
        """Calculate daily breakdown by payment method."""
        daily_breakdown_data = (
            successful_transactions.annotate(date=TruncDate("created_at"))
            .values("date", "method")
            .annotate(total=Sum("amount"))
            .order_by("date", "method")
        )

        daily_breakdown = {}
        for item in daily_breakdown_data:
            date_str = item["date"].strftime("%Y-%m-%d")
            if date_str not in daily_breakdown:
                daily_breakdown[date_str] = {"date": date_str, "total": 0}
            method_key = item["method"].lower().replace("_", "")
            daily_breakdown[date_str][method_key] = float(item["total"] or 0)

        # Calculate totals for each day
        for date_str, breakdown in daily_breakdown.items():
            breakdown["total"] = sum(
                v for k, v in breakdown.items() if k not in ["date", "total"]
            )

        return sorted(list(daily_breakdown.values()), key=lambda x: x["date"])

    @staticmethod
    def _calculate_payments_summary(transaction_querysets: Dict[str, Any], start_date: datetime, end_date: datetime, location_id: Optional[int] = None) -> Dict[str, Any]:
        """Calculate comprehensive payments summary."""
        successful = transaction_querysets['successful']
        refunded = transaction_querysets['refunded']
        failed = transaction_querysets['failed']
        canceled = transaction_querysets['canceled']

        # Calculate totals from transactions (include tips and surcharges to match Payment.total_collected)
        total_successful = float(successful.aggregate(total=Sum(F("amount") + F("tip") + F("surcharge")))["total"] or 0)
        total_refunds_amount = float(refunded.aggregate(total=Sum("refunded_amount"))["total"] or 0)
        total_failed = float(failed.aggregate(total=Sum(F("amount") + F("tip") + F("surcharge")))["total"] or 0)
        total_canceled = float(canceled.aggregate(total=Sum(F("amount") + F("tip") + F("surcharge")))["total"] or 0)
        
        # Get the actual total collected from Payment objects for consistency with sales reports
        # Use the same filtering logic as sales service for consistency
        filters = {
            "order__status": Order.OrderStatus.COMPLETED,
            "order__created_at__range": (start_date, end_date),
            "order__subtotal__gt": 0,
        }

        if location_id is not None:
            filters["order__store_location_id"] = location_id

        payment_totals = Payment.objects.filter(**filters).aggregate(
            total_collected_payments=Coalesce(Sum("total_collected"), Value(Decimal("0.00")))
        )
        total_collected_from_payments = float(payment_totals["total_collected_payments"] or 0)

        total_attempted = total_successful + total_refunds_amount + total_failed + total_canceled
        total_processing_issues = total_refunds_amount + total_failed + total_canceled

        # Calculate rates
        processing_success_rate = (
            round((total_successful / total_attempted * 100), 2)
            if total_attempted > 0 else 0
        )
        processing_issues_rate = (
            round((total_processing_issues / total_attempted * 100), 2)
            if total_attempted > 0 else 0
        )

        # Calculate net revenue correctly (use actual collected amounts minus refunds)
        # Use total_collected_from_payments for consistency with sales reports
        net_revenue = total_collected_from_payments - total_refunds_amount
        
        return {
            # Primary metrics
            "total_attempted": total_attempted,
            "successfully_processed": total_collected_from_payments,  # This should be the total collected (before refunds)
            "total_collected": total_collected_from_payments,  # Use Payment.total_collected for consistency
            "processing_issues": total_processing_issues,
            
            # Detailed breakdown
            "breakdown": {
                "successful": {
                    "amount": total_successful,
                    "count": successful.count(),
                },
                "refunded": {
                    "amount": total_refunds_amount,
                    "count": refunded.count(),
                },
                "failed": {
                    "amount": total_failed,
                    "count": failed.count(),
                },
                "canceled": {
                    "amount": total_canceled,
                    "count": canceled.count(),
                },
            },
            
            # Calculated rates
            "processing_success_rate": processing_success_rate,
            "processing_issues_rate": processing_issues_rate,
            
            # Legacy fields for backward compatibility
            "total_processed": total_collected_from_payments,  # Use collected for consistency
            "total_transactions": successful.count(),
            "total_refunds": total_refunds_amount,
            "total_refunded_transactions": refunded.count(),
            "net_revenue": net_revenue,
            "total_after_refunds": net_revenue,  # Explicit field for "total after refunds"
        }

    @staticmethod
    def _calculate_processing_stats(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None) -> Dict[str, Any]:
        """Calculate processing statistics for all transactions."""
        filters = {
            "payment__order__tenant": tenant,
            "payment__order__status": Order.OrderStatus.COMPLETED,
            "payment__order__created_at__range": (start_date, end_date),
            "payment__order__subtotal__gt": 0,
        }

        if location_id is not None:
            filters["payment__order__store_location_id"] = location_id

        all_transactions = PaymentTransaction.objects.select_related(
            "payment", "payment__order"
        ).filter(**filters)

        processing_stats = all_transactions.aggregate(
            total_attempts=Count("id"),
            successful=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)),
            failed=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.FAILED)),
            refunded=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.REFUNDED)),
        )

        success_rate = (
            (processing_stats["successful"] / processing_stats["total_attempts"]) * 100
            if processing_stats["total_attempts"] > 0 else 0
        )

        return {
            "total_attempts": processing_stats["total_attempts"],
            "successful": processing_stats["successful"],
            "failed": processing_stats["failed"],
            "refunded": processing_stats["refunded"],
            "success_rate": round(success_rate, 2),
        }

    @staticmethod
    def _calculate_order_reconciliation(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int], summary: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate order totals comparison for reconciliation."""
        filters = {
            "tenant": tenant,
            "status": Order.OrderStatus.COMPLETED,
            "created_at__range": (start_date, end_date),
            "subtotal__gt": 0,
        }

        if location_id is not None:
            filters["store_location_id"] = location_id

        completed_orders = Order.objects.filter(**filters).aggregate(
            order_total=Coalesce(Sum("grand_total"), Value(Decimal("0.00"))),
            order_count=Count("id"),
        )

        order_total = float(completed_orders["order_total"] or 0)
        payment_total = summary["total_attempted"]

        return {
            "order_grand_total": order_total,
            "order_count": completed_orders["order_count"],
            "payment_transaction_total": payment_total,
            "difference": order_total - payment_total,
        }

    @staticmethod
    def _get_detailed_transaction_data(start_date: datetime, end_date: datetime, tenant_id: Optional[int] = None, location_id: Optional[int] = None) -> list:
        """Get detailed transaction data for CSV export."""
        filters = {
            'payment__order__status': Order.OrderStatus.COMPLETED,
            'payment__order__created_at__range': (start_date, end_date),
            'payment__order__subtotal__gt': 0,
            'status__in': [
                PaymentTransaction.TransactionStatus.SUCCESSFUL,
                PaymentTransaction.TransactionStatus.REFUNDED
            ]
        }

        # Add tenant filter
        if tenant_id:
            filters['payment__order__tenant_id'] = tenant_id

        # Add location filter if specified
        if location_id is not None:
            filters['payment__order__store_location_id'] = location_id

        transactions = PaymentTransaction.objects.select_related(
            'payment', 'payment__order'
        ).filter(**filters).order_by('payment__payment_number', 'created_at')

        transaction_data = []
        for txn in transactions:
            # Calculate total including tip and surcharge
            total_amount = float(txn.amount) + float(txn.tip) + float(txn.surcharge)
            
            transaction_data.append({
                "payment_number": txn.payment.payment_number or str(txn.payment.id)[:8],
                "date": txn.created_at.strftime("%Y-%m-%d %H:%M:%S") if txn.created_at else "",
                "order_id": txn.payment.order.order_number or str(txn.payment.order.id)[:8],
                "method": txn.get_method_display(),
                "status": txn.get_status_display(),
                "amount": float(txn.amount),
                "tip": float(txn.tip),
                "surcharge": float(txn.surcharge),
                "total": total_amount,
                "card_brand": txn.card_brand or "",
                "card_last4": txn.card_last4 or "",
                "transaction_id": txn.transaction_id or "",
                "refunded_amount": float(txn.refunded_amount),
                "refund_reason": txn.refund_reason or "",
            })
        
        return transaction_data

    @staticmethod
    def _generate_multi_location_payments_report(
        tenant, start_date: datetime, end_date: datetime, use_cache: bool = True
    ) -> Dict[str, Any]:
        """Generate payments report for all locations with breakdown per location."""
        from settings.models import StoreLocation

        logger.info(f"Generating multi-location payments report for {start_date} to {end_date}")

        # Get all active locations
        active_locations = StoreLocation.objects.filter(
            tenant=tenant,
            is_active=True
        ).order_by('name')

        # Generate report for each location
        location_reports = []
        for location in active_locations:
            location_report = PaymentsReportService.generate_payments_report(
                tenant=tenant,
                start_date=start_date,
                end_date=end_date,
                location_id=location.id,
                use_cache=use_cache
            )
            location_reports.append({
                "location_id": location.id,
                "location_name": location.name,
                "report_data": location_report
            })

        # Calculate consolidated totals
        consolidated_totals = PaymentsReportService._calculate_consolidated_payments_totals(location_reports)

        return {
            "is_multi_location": True,
            "location_count": len(location_reports),
            "locations": location_reports,
            "consolidated": consolidated_totals,
            "generated_at": timezone.now().isoformat(),
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "tenant_id": tenant.id,
        }

    @staticmethod
    def _calculate_consolidated_payments_totals(location_reports: list) -> Dict[str, Any]:
        """Calculate consolidated payment totals across all locations."""
        from decimal import Decimal
        from collections import defaultdict

        # Initialize aggregation structures
        payment_methods_agg = defaultdict(lambda: {
            'amount': Decimal('0.00'),
            'count': 0,
            'tips_total': Decimal('0.00'),
            'surcharges_total': Decimal('0.00'),
            'processing_fees': Decimal('0.00'),
            'refunded_amount': Decimal('0.00'),
            'refunded_count': 0,
            'base_amount': Decimal('0.00'),
        })

        daily_volume_agg = defaultdict(lambda: {'amount': Decimal('0.00'), 'count': 0})
        daily_breakdown_agg = defaultdict(lambda: defaultdict(Decimal))

        # Totals for summary
        total_successful_amount = Decimal("0.00")
        total_refunded_amount = Decimal("0.00")
        total_failed_amount = Decimal("0.00")
        total_canceled_amount = Decimal("0.00")
        successful_count = 0
        refunded_count = 0
        failed_count = 0
        canceled_count = 0

        total_attempted = Decimal("0.00")
        total_processed = Decimal("0.00")
        total_collected = Decimal("0.00")
        processing_issues = Decimal("0.00")

        order_grand_total = Decimal("0.00")
        order_count = 0
        payment_transaction_total = Decimal("0.00")

        # Aggregate across locations
        for location_data in location_reports:
            loc_report = location_data.get('report_data', {})
            summary = loc_report.get('summary', {})
            breakdown = summary.get('breakdown', {})

            # Aggregate breakdown
            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})
            canceled = breakdown.get('canceled', {})

            total_successful_amount += Decimal(str(successful.get('amount', 0)))
            total_refunded_amount += Decimal(str(refunded.get('amount', 0)))
            total_failed_amount += Decimal(str(failed.get('amount', 0)))
            total_canceled_amount += Decimal(str(canceled.get('amount', 0)))
            successful_count += successful.get('count', 0)
            refunded_count += refunded.get('count', 0)
            failed_count += failed.get('count', 0)
            canceled_count += canceled.get('count', 0)

            # Aggregate summary totals
            total_attempted += Decimal(str(summary.get('total_attempted', 0)))
            total_processed += Decimal(str(summary.get('total_processed', 0)))
            total_collected += Decimal(str(summary.get('total_collected', 0)))
            processing_issues += Decimal(str(summary.get('processing_issues', 0)))

            # Aggregate payment methods
            for method in loc_report.get('payment_methods', []):
                method_name = method['method']
                payment_methods_agg[method_name]['amount'] += Decimal(str(method.get('amount', 0)))
                payment_methods_agg[method_name]['count'] += method.get('count', 0)
                payment_methods_agg[method_name]['tips_total'] += Decimal(str(method.get('tips_total', 0)))
                payment_methods_agg[method_name]['surcharges_total'] += Decimal(str(method.get('surcharges_total', 0)))
                payment_methods_agg[method_name]['processing_fees'] += Decimal(str(method.get('processing_fees', 0)))
                payment_methods_agg[method_name]['refunded_amount'] += Decimal(str(method.get('refunded_amount', 0)))
                payment_methods_agg[method_name]['refunded_count'] += method.get('refunded_count', 0)
                payment_methods_agg[method_name]['base_amount'] += Decimal(str(method.get('base_amount', 0)))

            # Aggregate daily volume
            for daily in loc_report.get('daily_volume', []):
                date = daily['date']
                daily_volume_agg[date]['amount'] += Decimal(str(daily.get('amount', 0)))
                daily_volume_agg[date]['count'] += daily.get('count', 0)

            # Aggregate daily breakdown
            for daily in loc_report.get('daily_breakdown', []):
                date = daily['date']
                for key, value in daily.items():
                    if key != 'date':
                        daily_breakdown_agg[date][key] += Decimal(str(value))

            # Aggregate order totals comparison
            comparison = loc_report.get('order_totals_comparison', {})
            order_grand_total += Decimal(str(comparison.get('order_grand_total', 0)))
            order_count += comparison.get('order_count', 0)
            payment_transaction_total += Decimal(str(comparison.get('payment_transaction_total', 0)))

        # Build payment methods list
        payment_methods = []
        total_payment_amount = float(total_successful_amount)

        for method, data in payment_methods_agg.items():
            amount = float(data['amount'])
            count = data['count']
            payment_methods.append({
                'method': method,
                'amount': amount,
                'count': count,
                'avg_amount': amount / count if count > 0 else 0,
                'processing_fees': float(data['processing_fees']),
                'percentage': (amount / total_payment_amount * 100) if total_payment_amount > 0 else 0,
                'trend': 0,  # Not calculated for multi-location
                'refunded_amount': float(data['refunded_amount']),
                'refunded_count': data['refunded_count'],
                'total_processed': amount,
                'net_amount': amount,
                'tips_total': float(data['tips_total']),
                'base_amount': float(data['base_amount']),
                'surcharges_total': float(data['surcharges_total']),
            })

        # Build daily volume list
        daily_volume = [
            {
                'date': date,
                'amount': float(data['amount']),
                'count': data['count']
            }
            for date, data in sorted(daily_volume_agg.items())
        ]

        # Build daily breakdown list
        daily_breakdown = [
            {
                'date': date,
                **{k: float(v) for k, v in data.items()}
            }
            for date, data in sorted(daily_breakdown_agg.items())
        ]

        # Calculate total transactions
        total_transactions = successful_count + refunded_count + failed_count + canceled_count

        # Build the complete consolidated structure matching the single-location format
        return {
            'payment_methods': payment_methods,
            'daily_volume': daily_volume,
            'daily_breakdown': daily_breakdown,
            'processing_stats': {
                'total_attempts': total_transactions,
                'successful': successful_count,
                'failed': failed_count,
                'refunded': refunded_count,
                'success_rate': (successful_count / total_transactions * 100) if total_transactions > 0 else 0,
            },
            'order_totals_comparison': {
                'order_grand_total': float(order_grand_total),
                'order_count': order_count,
                'payment_transaction_total': float(payment_transaction_total),
                'difference': float(order_grand_total - payment_transaction_total),
            },
            'summary': {
                'breakdown': {
                    'successful': {
                        'amount': float(total_successful_amount),
                        'count': successful_count,
                    },
                    'refunded': {
                        'amount': float(total_refunded_amount),
                        'count': refunded_count,
                    },
                    'failed': {
                        'amount': float(total_failed_amount),
                        'count': failed_count,
                    },
                    'canceled': {
                        'amount': float(total_canceled_amount),
                        'count': canceled_count,
                    },
                },
                'total_attempted': float(total_attempted),
                'successfully_processed': float(total_processed),
                'processing_issues': float(processing_issues),
                'processing_success_rate': (successful_count / total_transactions * 100) if total_transactions > 0 else 0,
                'processing_issues_rate': (failed_count / total_transactions * 100) if total_transactions > 0 else 0,
                'total_processed': float(total_processed),
                'total_transactions': total_transactions,
                'total_refunds': float(total_refunded_amount),
                'total_refunded_transactions': refunded_count,
                'net_revenue': float(total_successful_amount - total_refunded_amount),
                'total_collected': float(total_collected),
            }
        }

    # Export Methods

    @staticmethod
    def export_payments_to_csv(report_data: Dict[str, Any]) -> bytes:
        """Export payments report to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            PaymentsReportService._export_multi_location_payments_to_csv(writer, report_data)
        else:
            PaymentsReportService._export_payments_to_csv(writer, report_data)

        csv_bytes = output.getvalue().encode('utf-8')
        output.close()
        return csv_bytes

    @staticmethod
    def _export_payments_to_csv(writer, report_data: Dict[str, Any]):
        """Export comprehensive payments report to CSV with transaction-level details."""
        from django.utils import timezone
        from datetime import datetime
        
        # Header
        writer.writerow(["Payments Report"])
        writer.writerow(["Generated:", report_data.get("generated_at", timezone.now().isoformat())])
        
        # Extract and format date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            writer.writerow([f"Date Range: {start_date.date()} to {end_date.date()}"])
        except:
            writer.writerow([f"Date Range: {start_str} to {end_str}"])

        # Add location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        writer.writerow(["Location:", location_name])

        writer.writerow([])
        
        # === PAYMENT SUMMARY ===
        writer.writerow(["=== PAYMENT SUMMARY ==="])
        writer.writerow(["Metric", "Value"])
        
        summary = report_data.get("summary", {})
        writer.writerow(["Total Collected", f"${summary.get('total_collected', 0):,.2f}"])
        writer.writerow(["Total Refunds", f"${summary.get('total_refunds', 0):,.2f}"])
        writer.writerow(["Total After Refunds", f"${summary.get('total_after_refunds', 0):,.2f}"])
        writer.writerow(["Net Revenue", f"${summary.get('net_revenue', 0):,.2f}"])
        writer.writerow(["Success Rate", f"{summary.get('processing_success_rate', 0):.2f}%"])
        writer.writerow(["Total Transactions", summary.get('total_transactions', 0)])
        writer.writerow([])
        
        # === PAYMENT METHODS SUMMARY ===
        writer.writerow(["=== PAYMENT METHODS SUMMARY ==="])
        writer.writerow([
            "Method", "Amount", "Count", "Percentage", "Avg Amount", "Processing Fees"
        ])
        
        for method in report_data.get("payment_methods", []):
            writer.writerow([
                method.get("method", ""),
                f"${method.get('amount', 0):,.2f}",
                method.get("count", 0),
                f"{method.get('percentage', 0):.2f}%",
                f"${method.get('avg_amount', 0):,.2f}",
                f"${method.get('processing_fees', 0):,.2f}",
            ])
        
        writer.writerow([])
        
        # === DETAILED TRANSACTION DATA ===
        writer.writerow(["=== COMPLETED PAYMENT TRANSACTIONS ==="])
        
        # Get detailed transaction data
        try:
            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            transaction_data = PaymentsReportService._get_detailed_transaction_data(
                start_date, end_date, tenant_id, location_id
            )
        except:
            # Fallback if datetime parsing fails
            transaction_data = []
        
        # Transaction headers
        writer.writerow([
            "Payment Number", "Date", "Order ID", "Payment Method", "Transaction Status",
            "Amount", "Tip", "Surcharge", "Total", "Card Brand", "Card Last4",
            "Transaction ID", "Refunded Amount", "Refund Reason"
        ])
        
        for transaction in transaction_data:
            writer.writerow([
                transaction.get("payment_number", ""),
                transaction.get("date", ""),
                transaction.get("order_id", ""),
                transaction.get("method", ""),
                transaction.get("status", ""),
                f"${transaction.get('amount', 0):,.2f}",
                f"${transaction.get('tip', 0):,.2f}",
                f"${transaction.get('surcharge', 0):,.2f}",
                f"${transaction.get('total', 0):,.2f}",
                transaction.get("card_brand", ""),
                transaction.get("card_last4", ""),
                transaction.get("transaction_id", ""),
                f"${transaction.get('refunded_amount', 0):,.2f}",
                transaction.get("refund_reason", ""),
            ])
        
        writer.writerow([])
        writer.writerow([f"Total Transactions Exported: {len(transaction_data)}"])

    @staticmethod
    def _export_multi_location_payments_to_csv(writer, report_data: Dict[str, Any]):
        """Export multi-location payments report to CSV with consolidated and per-location sections."""
        from django.utils import timezone
        from datetime import datetime

        # Header
        writer.writerow(["Payments Report - All Locations"])
        writer.writerow(["Generated:", report_data.get("generated_at", timezone.now().isoformat())])

        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            writer.writerow([f"Date Range: {start_date} to {end_date}"])
        except:
            writer.writerow([f"Date Range: {start_str} to {end_str}"])

        # Location count
        location_count = report_data.get('location_count', 0)
        writer.writerow([f"Locations Included: {location_count}"])
        writer.writerow([])

        # === CONSOLIDATED PAYMENT SUMMARY ===
        consolidated = report_data.get('consolidated', {})

        writer.writerow(["=== CONSOLIDATED PAYMENT SUMMARY (ALL LOCATIONS) ==="])
        writer.writerow(["Metric", "Total Value"])
        writer.writerow(["Total Successful Payments", f"${consolidated.get('total_successful_amount', 0):,.2f}"])
        writer.writerow(["Total Refunded", f"${consolidated.get('total_refunded_amount', 0):,.2f}"])
        writer.writerow(["Net Payment Amount", f"${consolidated.get('net_payment_amount', 0):,.2f}"])
        writer.writerow(["Total Tips", f"${consolidated.get('total_tips', 0):,.2f}"])
        writer.writerow(["Total Surcharges", f"${consolidated.get('total_surcharges', 0):,.2f}"])
        writer.writerow([])
        writer.writerow(["Successful Transactions", f"{consolidated.get('successful_count', 0):,}"])
        writer.writerow(["Refunded Transactions", f"{consolidated.get('refunded_count', 0):,}"])
        writer.writerow(["Failed Transactions", f"{consolidated.get('failed_count', 0):,}"])
        writer.writerow(["Total Transactions", f"{consolidated.get('total_transactions', 0):,}"])
        writer.writerow(["Average Transaction", f"${consolidated.get('avg_transaction_amount', 0):,.2f}"])
        writer.writerow([])

        # === LOCATION COMPARISON TABLE ===
        writer.writerow(["=== LOCATION COMPARISON ==="])
        writer.writerow(["Location", "Successful", "Refunded", "Net Amount", "Tips", "Transactions", "Avg Txn", "Failed"])

        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})
            loc_summary = loc_report.get('summary', {})
            breakdown = loc_summary.get('breakdown', {})

            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})

            successful_amount = successful.get('amount', 0)
            successful_count = successful.get('count', 0)
            refunded_amount = refunded.get('amount', 0)
            failed_count = failed.get('count', 0)
            net_revenue = loc_summary.get('net_revenue', 0)

            avg_txn = successful_amount / successful_count if successful_count > 0 else 0

            writer.writerow([
                location_name,
                f"${successful_amount:,.2f}",
                f"${refunded_amount:,.2f}",
                f"${net_revenue:,.2f}",
                f"$0.00",  # Tips (included in successful_amount)
                f"{successful_count:,}",
                f"${avg_txn:,.2f}",
                f"{failed_count:,}",
            ])

        writer.writerow([])
        writer.writerow([])

        # === INDIVIDUAL LOCATION DETAILS ===
        writer.writerow(["=" * 100])
        writer.writerow(["=== DETAILED BREAKDOWN BY LOCATION ==="])
        writer.writerow(["=" * 100])
        writer.writerow([])

        for i, location_data in enumerate(locations, 1):
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Location separator
            writer.writerow(["=" * 100])
            writer.writerow([f"LOCATION {i}: {location_name}"])
            writer.writerow(["=" * 100])
            writer.writerow([])

            # Export this location's summary data (without full transaction details for brevity)
            loc_summary = loc_report.get('summary', {})
            breakdown = loc_summary.get('breakdown', {})
            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})

            successful_amount = successful.get('amount', 0)
            successful_count = successful.get('count', 0)
            refunded_amount = refunded.get('amount', 0)
            refunded_count = refunded.get('count', 0)
            failed_count = failed.get('count', 0)
            net_revenue = loc_summary.get('net_revenue', 0)
            total_collected = loc_summary.get('total_collected', 0)

            # Calculate average
            avg_txn = successful_amount / successful_count if successful_count > 0 else 0

            writer.writerow(["PAYMENT SUMMARY"])
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Total Successful Payments", f"${successful_amount:,.2f}"])
            writer.writerow(["Total Refunded", f"${refunded_amount:,.2f}"])
            writer.writerow(["Net Payment Amount", f"${net_revenue:,.2f}"])
            writer.writerow(["Total Collected", f"${total_collected:,.2f}"])
            writer.writerow([])
            writer.writerow(["Successful Transactions", f"{successful_count:,}"])
            writer.writerow(["Refunded Transactions", f"{refunded_count:,}"])
            writer.writerow(["Failed Transactions", f"{failed_count:,}"])
            writer.writerow(["Average Transaction", f"${avg_txn:,.2f}"])
            writer.writerow([])

            # Payment methods breakdown
            payment_methods = loc_report.get('payment_methods', [])
            if payment_methods:
                writer.writerow(["PAYMENT METHODS BREAKDOWN"])
                writer.writerow(["Method", "Count", "Amount", "Percentage"])
                for method in payment_methods:
                    writer.writerow([
                        method.get('method', 'Unknown'),
                        f"{method.get('count', 0):,}",
                        f"${method.get('amount', 0):,.2f}",
                        f"{method.get('percentage', 0):.1f}%"
                    ])
                writer.writerow([])

            writer.writerow([])

    @staticmethod
    def export_payments_to_xlsx(report_data: Dict[str, Any], ws, header_font, header_fill, header_alignment):
        """Export payments report to Excel format."""
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            # For multi-location reports, ws is actually the workbook
            PaymentsReportService._export_multi_location_payments_to_xlsx(ws, report_data, header_font, header_fill, header_alignment)
        else:
            # Single location report - ws is a worksheet
            PaymentsReportService._export_payments_to_xlsx(ws, report_data, header_font, header_fill, header_alignment)

    @staticmethod
    def _export_payments_to_xlsx(ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export comprehensive payments report to Excel with transaction-level details."""
        from django.utils import timezone
        from datetime import datetime
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        section_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_header_font = Font(bold=True, color="FFFFFF", size=12)
        
        subsection_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subsection_font = Font(bold=True, size=11)
        
        currency_format = '"$"#,##0.00'
        number_format = '#,##0'
        
        row = 1
        
        # Title
        ws.merge_cells(f"A{row}:N{row}")  # Extended for more columns
        ws[f"A{row}"] = "Payments Report"
        ws[f"A{row}"].font = Font(bold=True, size=16)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        ws.merge_cells(f"A{row}:N{row}")
        ws[f"A{row}"] = f"Generated: {generated_at}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1
        
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            ws.merge_cells(f"A{row}:N{row}")
            ws[f"A{row}"] = f"Date Range: {start_date.date()} to {end_date.date()}"
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
        except:
            ws.merge_cells(f"A{row}:N{row}")
            ws[f"A{row}"] = f"Date Range: {start_str} to {end_str}"
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        ws.cell(row=row, column=1, value="Location:")
        ws.cell(row=row, column=2, value=location_name)
        row += 2
        
        # === PAYMENT SUMMARY ===
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "PAYMENT SUMMARY"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Summary headers
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].font = subsection_font
        ws[f"B{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        ws[f"B{row}"].fill = subsection_fill
        row += 1
        
        # Summary metrics
        summary = report_data.get("summary", {})
        summary_metrics = [
            ("Total Collected", summary.get('total_collected', 0), currency_format),
            ("Total Refunds", summary.get('total_refunds', 0), currency_format),
            ("Total After Refunds", summary.get('total_after_refunds', 0), currency_format),
            ("Net Revenue", summary.get('net_revenue', 0), currency_format),
            ("Success Rate", summary.get('processing_success_rate', 0), '"0.00"%'),
            ("Total Transactions", summary.get('total_transactions', 0), number_format),
        ]
        
        for metric, value, fmt in summary_metrics:
            ws[f"A{row}"] = metric
            if fmt == '"0.00"%':
                ws[f"B{row}"] = value / 100  # Convert percentage for Excel
            else:
                ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = fmt
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1
        
        row += 1  # Empty row
        
        # === PAYMENT METHODS SUMMARY ===
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "PAYMENT METHODS SUMMARY"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Payment method headers
        method_headers = ["Method", "Amount", "Count", "Percentage", "Avg Amount", "Processing Fees"]
        
        for col, header in enumerate(method_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        
        # Payment method data
        for method in report_data.get("payment_methods", []):
            method_data = [
                method.get("method", ""),
                float(method.get('amount', 0)),
                method.get('count', 0),
                method.get('percentage', 0) / 100,  # Convert for Excel percentage format
                float(method.get('avg_amount', 0)),
                float(method.get('processing_fees', 0)),
            ]
            
            for col, value in enumerate(method_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Apply number formats
                if col in [2, 5, 6]:  # Currency columns
                    cell.number_format = currency_format
                elif col == 4:  # Percentage column
                    cell.number_format = '0.00%'
                elif col == 3:  # Count column
                    cell.number_format = number_format
            row += 1
        
        row += 2  # Empty rows
        
        # === DETAILED TRANSACTION DATA ===
        ws.merge_cells(f"A{row}:N{row}")
        ws[f"A{row}"] = "COMPLETED PAYMENT TRANSACTIONS"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Get detailed transaction data
        try:
            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            transaction_data = PaymentsReportService._get_detailed_transaction_data(
                start_date, end_date, tenant_id, location_id
            )
        except:
            transaction_data = []
        
        # Transaction headers
        transaction_headers = [
            "Payment Number", "Date", "Order ID", "Payment Method", "Transaction Status",
            "Amount", "Tip", "Surcharge", "Total", "Card Brand", "Card Last4",
            "Transaction ID", "Refunded Amount", "Refund Reason"
        ]
        
        for col, header in enumerate(transaction_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        
        # Transaction data
        for transaction in transaction_data:
            transaction_row_data = [
                transaction.get("payment_number", ""),
                transaction.get("date", ""),
                transaction.get("order_id", ""),
                transaction.get("method", ""),
                transaction.get("status", ""),
                float(transaction.get('amount', 0)),
                float(transaction.get('tip', 0)),
                float(transaction.get('surcharge', 0)),
                float(transaction.get('total', 0)),
                transaction.get("card_brand", ""),
                transaction.get("card_last4", ""),
                transaction.get("transaction_id", ""),
                float(transaction.get('refunded_amount', 0)),
                transaction.get("refund_reason", ""),
            ]
            
            for col, value in enumerate(transaction_row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Apply number formats for currency columns
                if col in [6, 7, 8, 9, 13]:  # Amount, tip, surcharge, total, refunded_amount
                    cell.number_format = currency_format
            row += 1
        
        # Add summary row
        row += 1
        ws[f"A{row}"] = f"Total Transactions Exported: {len(transaction_data)}"
        ws[f"A{row}"].font = Font(bold=True)
        
        # Auto-adjust column widths
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            pass

    @staticmethod
    def _export_multi_location_payments_to_xlsx(wb, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export multi-location payments report to Excel with multiple sheets."""
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.utils import timezone

        # Define styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        section_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_header_font = Font(bold=True, color="FFFFFF", size=12)
        subsection_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subsection_font = Font(bold=True, size=11)
        currency_format = '"$"#,##0.00'
        number_format = '#,##0'

        # Sheet 1: Summary - All Locations
        summary_ws = wb.active
        summary_ws.title = "Summary - All Locations"

        row = 1

        # Title
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = "Payments Report - All Locations Summary"
        summary_ws[f"A{row}"].font = Font(bold=True, size=16)
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = f"Generated: {generated_at}"
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        from datetime import datetime
        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            summary_ws.merge_cells(f"A{row}:F{row}")
            summary_ws[f"A{row}"] = f"Date Range: {start_date} to {end_date}"
            summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        except:
            summary_ws.merge_cells(f"A{row}:F{row}")
            summary_ws[f"A{row}"] = f"Date Range: {start_str} to {end_str}"
            summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Location count
        location_count = report_data.get('location_count', 0)
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = f"Locations Included: {location_count}"
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        summary_ws[f"A{row}"].font = Font(bold=True)
        row += 2

        # === CONSOLIDATED PAYMENT SUMMARY ===
        consolidated = report_data.get('consolidated', {})

        summary_ws.merge_cells(f"A{row}:B{row}")
        summary_ws[f"A{row}"] = "CONSOLIDATED PAYMENT SUMMARY"
        summary_ws[f"A{row}"].font = section_header_font
        summary_ws[f"A{row}"].fill = section_header_fill
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Headers
        summary_ws[f"A{row}"] = "Metric"
        summary_ws[f"B{row}"] = "Total Value"
        summary_ws[f"A{row}"].font = subsection_font
        summary_ws[f"B{row}"].font = subsection_font
        summary_ws[f"A{row}"].fill = subsection_fill
        summary_ws[f"B{row}"].fill = subsection_fill
        row += 1

        # Consolidated metrics
        consolidated_metrics = [
            ("Total Successful Payments", consolidated.get('total_successful_amount', 0), currency_format),
            ("Total Refunded", consolidated.get('total_refunded_amount', 0), currency_format),
            ("Net Payment Amount", consolidated.get('net_payment_amount', 0), currency_format),
            ("Total Tips", consolidated.get('total_tips', 0), currency_format),
            ("Total Surcharges", consolidated.get('total_surcharges', 0), currency_format),
            ("", "", ""),  # Empty row
            ("Successful Transactions", consolidated.get('successful_count', 0), number_format),
            ("Refunded Transactions", consolidated.get('refunded_count', 0), number_format),
            ("Failed Transactions", consolidated.get('failed_count', 0), number_format),
            ("Total Transactions", consolidated.get('total_transactions', 0), number_format),
            ("Average Transaction", consolidated.get('avg_transaction_amount', 0), currency_format),
        ]

        for metric, value, fmt in consolidated_metrics:
            summary_ws[f"A{row}"] = metric
            summary_ws[f"B{row}"] = value if value != "" else ""
            if fmt and value != "":
                summary_ws[f"B{row}"].number_format = fmt
            summary_ws[f"A{row}"].border = thin_border
            summary_ws[f"B{row}"].border = thin_border
            row += 1

        row += 2

        # === LOCATION COMPARISON TABLE ===
        summary_ws.merge_cells(f"A{row}:H{row}")
        summary_ws[f"A{row}"] = "LOCATION COMPARISON"
        summary_ws[f"A{row}"].font = section_header_font
        summary_ws[f"A{row}"].fill = section_header_fill
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Comparison headers
        comparison_headers = ["Location", "Successful", "Refunded", "Net Amount", "Tips", "Txns", "Avg Txn", "Failed"]
        for col, header in enumerate(comparison_headers, 1):
            cell = summary_ws.cell(row=row, column=col, value=header)
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Location comparison rows
        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})
            loc_summary = loc_report.get('summary', {})
            breakdown = loc_summary.get('breakdown', {})

            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})

            successful_amount = successful.get('amount', 0)
            successful_count = successful.get('count', 0)
            refunded_amount = refunded.get('amount', 0)
            failed_count = failed.get('count', 0)
            net_revenue = loc_summary.get('net_revenue', 0)

            avg_txn = successful_amount / successful_count if successful_count > 0 else 0

            comparison_data = [
                location_name,
                successful_amount,
                refunded_amount,
                net_revenue,
                0,  # Tips (included in successful_amount)
                successful_count,
                avg_txn,
                failed_count,
            ]

            for col, value in enumerate(comparison_data, 1):
                cell = summary_ws.cell(row=row, column=col, value=value)
                cell.border = thin_border

                # Apply number formats
                if col == 1:  # Location name
                    cell.alignment = Alignment(horizontal="left")
                elif col in [2, 3, 4, 5, 7]:  # Currency columns
                    cell.number_format = currency_format
                elif col in [6, 8]:  # Number columns
                    cell.number_format = number_format

            row += 1

        # Auto-adjust column widths for summary sheet
        for column in summary_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width

        # === CREATE INDIVIDUAL LOCATION SHEETS ===
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Create a new sheet for this location
            # Sanitize sheet name (Excel sheet names have restrictions)
            safe_sheet_name = location_name[:31]  # Excel sheet name max length is 31
            # Remove invalid characters
            for char in ['\\', '/', '*', '?', ':', '[', ']']:
                safe_sheet_name = safe_sheet_name.replace(char, '')

            location_ws = wb.create_sheet(title=safe_sheet_name)

            # Export this location's data using the existing single-location export logic
            PaymentsReportService._export_payments_to_xlsx(location_ws, loc_report, header_font, header_fill, header_alignment)

    @staticmethod
    def export_payments_to_pdf(report_data: Dict[str, Any], story, styles):
        """Export payments report to PDF format."""
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            PaymentsReportService._export_multi_location_payments_to_pdf(story, report_data, styles)
        else:
            PaymentsReportService._export_payments_to_pdf(story, report_data, styles)

    @staticmethod
    def _export_payments_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export comprehensive payments report to PDF with key transaction details."""
        from django.utils import timezone
        from datetime import datetime
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Spacer, PageBreak
        
        # Report header
        story.append(Paragraph("Payments Report", styles["Title"]))
        story.append(Spacer(1, 12))
        
        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))
        
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            story.append(Paragraph(f"Date Range: {start_date.date()} to {end_date.date()}", styles["Normal"]))
        except:
            story.append(Paragraph(f"Date Range: {start_str} to {end_str}", styles["Normal"]))

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        story.append(Paragraph(f"<b>Location:</b> {location_name}", styles['Normal']))

        story.append(Spacer(1, 20))
        
        # === PAYMENT SUMMARY ===
        story.append(Paragraph("Payment Summary", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        summary = report_data.get("summary", {})
        summary_data = [
            ["Metric", "Value"],
            ["Total Collected", f"${summary.get('total_collected', 0):,.2f}"],
            ["Total Refunds", f"${summary.get('total_refunds', 0):,.2f}"],
            ["Total After Refunds", f"${summary.get('total_after_refunds', 0):,.2f}"],
            ["Net Revenue", f"${summary.get('net_revenue', 0):,.2f}"],
            ["Success Rate", f"{summary.get('processing_success_rate', 0):.2f}%"],
            ["Total Transactions", f"{summary.get('total_transactions', 0):,}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3 * inch, 2.5 * inch])
        summary_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])
        )
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # === PAYMENT METHODS ===
        story.append(Paragraph("Payment Methods Summary", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        payment_methods_data = [["Method", "Amount", "Count", "Percentage", "Processing Fees"]]
        for method in report_data.get("payment_methods", []):
            payment_methods_data.append([
                method.get("method", ""),
                f"${method.get('amount', 0):,.2f}",
                f"{method.get('count', 0):,}",
                f"{method.get('percentage', 0):.1f}%",
                f"${method.get('processing_fees', 0):,.2f}",
            ])
        
        payment_methods_table = Table(payment_methods_data, colWidths=[
            1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1.2*inch
        ])
        payment_methods_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])
        )
        
        story.append(payment_methods_table)
        story.append(Spacer(1, 20))
        
        # === KEY TRANSACTIONS (Sample) ===
        story.append(Paragraph("Recent Transactions Sample", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        # Get sample of transactions for PDF (limit to avoid clutter)
        try:
            # Ensure we have datetime objects for the query
            if isinstance(start_date, datetime):
                query_start = start_date
                query_end = end_date
            else:
                # Parse from strings if needed
                query_start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                query_end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))

            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            transaction_data = PaymentsReportService._get_detailed_transaction_data(
                query_start, query_end, tenant_id, location_id
            )
            # Show max 10 most recent transactions
            sample_transactions = transaction_data[:10] if transaction_data else []
        except Exception as e:
            logger.error(f"Error getting transaction data for PDF: {e}")
            sample_transactions = []
        
        if sample_transactions:
            # Simplified transaction table for PDF
            transaction_table_data = [["Payment #", "Date", "Method", "Amount", "Status"]]
            for txn in sample_transactions:
                transaction_table_data.append([
                    txn.get("payment_number", "")[:12],  # Truncate for space
                    txn.get("date", "")[:10] if txn.get("date") else "",  # Date only
                    txn.get("method", ""),
                    f"${txn.get('total', 0):,.2f}",
                    txn.get("status", ""),
                ])
            
            transaction_table = Table(transaction_table_data, colWidths=[
                1.2*inch, 1.0*inch, 1.2*inch, 1.0*inch, 1.0*inch
            ])
            transaction_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ])
            )
            
            story.append(transaction_table)
            story.append(Spacer(1, 12))
            
            # Add note about full transaction details
            story.append(Paragraph(
                f"<i>Showing {len(sample_transactions)} of {len(transaction_data)} total transactions. "
                f"For complete transaction details, export to CSV or Excel.</i>",
                styles["Normal"]
            ))
        else:
            story.append(Paragraph("No transaction data available for this period.", styles["Normal"]))

        story.append(Spacer(1, 20))

    @staticmethod
    def _export_multi_location_payments_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export multi-location payments report to PDF with multiple pages."""
        from django.utils import timezone
        from datetime import datetime
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        # === EXECUTIVE SUMMARY PAGE ===
        story.append(Paragraph("Payments Report - All Locations", styles["Title"]))
        story.append(Paragraph("Executive Summary", styles["Heading1"]))
        story.append(Spacer(1, 12))

        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))

        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))
        except:
            story.append(Paragraph(f"Date Range: {start_str} to {end_str}", styles["Normal"]))

        # Location count
        location_count = report_data.get('location_count', 0)
        story.append(Paragraph(f"<b>Locations Included: {location_count}</b>", styles["Normal"]))
        story.append(Spacer(1, 20))

        # === CONSOLIDATED PAYMENT SUMMARY ===
        consolidated = report_data.get('consolidated', {})

        story.append(Paragraph("Consolidated Payment Summary (All Locations)", styles["Heading2"]))
        story.append(Spacer(1, 12))

        consolidated_summary_data = [
            ["Metric", "Total Value"],
            ["Total Successful Payments", f"${consolidated.get('total_successful_amount', 0):,.2f}"],
            ["Total Refunded", f"${consolidated.get('total_refunded_amount', 0):,.2f}"],
            ["Net Payment Amount", f"${consolidated.get('net_payment_amount', 0):,.2f}"],
            ["Total Tips", f"${consolidated.get('total_tips', 0):,.2f}"],
            ["Total Surcharges", f"${consolidated.get('total_surcharges', 0):,.2f}"],
            ["", ""],  # Empty row
            ["Successful Transactions", f"{consolidated.get('successful_count', 0):,}"],
            ["Refunded Transactions", f"{consolidated.get('refunded_count', 0):,}"],
            ["Failed Transactions", f"{consolidated.get('failed_count', 0):,}"],
            ["Total Transactions", f"{consolidated.get('total_transactions', 0):,}"],
            ["Average Transaction", f"${consolidated.get('avg_transaction_amount', 0):,.2f}"],
        ]

        summary_table = Table(consolidated_summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 20))

        # === LOCATION COMPARISON TABLE ===
        story.append(Paragraph("Location Comparison", styles["Heading2"]))
        story.append(Spacer(1, 12))

        comparison_data = [
            ["Location", "Successful", "Refunded", "Net", "Tips", "Txns", "Failed"]
        ]

        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})
            loc_summary = loc_report.get('summary', {})
            breakdown = loc_summary.get('breakdown', {})

            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})

            successful_amount = successful.get('amount', 0)
            successful_count = successful.get('count', 0)
            refunded_amount = refunded.get('amount', 0)
            failed_count = failed.get('count', 0)
            net_revenue = loc_summary.get('net_revenue', 0)

            comparison_data.append([
                location_name,
                f"${successful_amount:,.2f}",
                f"${refunded_amount:,.2f}",
                f"${net_revenue:,.2f}",
                f"$0.00",  # Tips (included in successful_amount)
                f"{successful_count:,}",
                f"{failed_count:,}",
            ])

        comparison_table = Table(comparison_data, colWidths=[
            1.6*inch,  # Location
            1.2*inch,  # Successful
            1.0*inch,  # Refunded
            1.0*inch,  # Net
            0.8*inch,  # Tips
            0.7*inch,  # Txns
            0.7*inch,  # Failed
        ])
        comparison_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(comparison_table)
        story.append(Spacer(1, 20))

        # === INDIVIDUAL LOCATION PAGES ===
        for i, location_data in enumerate(locations, 1):
            # Page break before each location
            story.append(PageBreak())

            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Location header
            story.append(Paragraph(f"Location {i}: {location_name}", styles["Title"]))
            story.append(Paragraph("Detailed Payments Report", styles["Heading1"]))
            story.append(Spacer(1, 12))

            # Add date range for this location
            story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))
            story.append(Spacer(1, 20))

            # Export summary data for this location
            loc_summary = loc_report.get('summary', {})
            breakdown = loc_summary.get('breakdown', {})

            successful = breakdown.get('successful', {})
            refunded = breakdown.get('refunded', {})
            failed = breakdown.get('failed', {})

            successful_amount = successful.get('amount', 0)
            successful_count = successful.get('count', 0)
            refunded_amount = refunded.get('amount', 0)
            refunded_count = refunded.get('count', 0)
            failed_count = failed.get('count', 0)
            net_revenue = loc_summary.get('net_revenue', 0)
            total_collected = loc_summary.get('total_collected', 0)

            avg_txn = successful_amount / successful_count if successful_count > 0 else 0

            story.append(Paragraph("Payment Summary", styles["Heading2"]))
            story.append(Spacer(1, 12))

            loc_summary_data = [
                ["Metric", "Value"],
                ["Total Successful Payments", f"${successful_amount:,.2f}"],
                ["Total Refunded", f"${refunded_amount:,.2f}"],
                ["Net Payment Amount", f"${net_revenue:,.2f}"],
                ["Total Collected", f"${total_collected:,.2f}"],
                ["", ""],
                ["Successful Transactions", f"{successful_count:,}"],
                ["Refunded Transactions", f"{refunded_count:,}"],
                ["Failed Transactions", f"{failed_count:,}"],
                ["Average Transaction", f"${avg_txn:,.2f}"],
            ]

            loc_table = Table(loc_summary_data, colWidths=[4*inch, 2*inch])
            loc_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(loc_table)
            story.append(Spacer(1, 20))

            # Payment methods breakdown
            payment_methods = loc_report.get('payment_methods', [])
            if payment_methods:
                story.append(Paragraph("Payment Methods Breakdown", styles["Heading2"]))
                story.append(Spacer(1, 12))

                method_data = [["Method", "Count", "Amount", "Percentage"]]
                for method in payment_methods:
                    method_data.append([
                        method.get('method', 'Unknown'),
                        f"{method.get('count', 0):,}",
                        f"${method.get('amount', 0):,.2f}",
                        f"{method.get('percentage', 0):.1f}%"
                    ])

                method_table = Table(method_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
                method_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("FONTSIZE", (0, 1), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]))

                story.append(method_table)
                story.append(Spacer(1, 20))