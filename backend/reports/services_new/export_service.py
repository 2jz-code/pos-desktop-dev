"""
Unified export service for handling all report export functionality.
Supports CSV, Excel (XLSX), and PDF formats for all report types.
"""
import csv
import io
import logging
from decimal import Decimal
from datetime import datetime
from typing import Dict, Any, List, Optional

# Export functionality imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .base import BaseReportService

logger = logging.getLogger(__name__)


class ExportService(BaseReportService):
    """Service for exporting reports to various formats."""

    @classmethod
    def export_to_csv(cls, report_data: Dict[str, Any], report_type: str) -> bytes:
        """
        Export report data to CSV format.
        
        Args:
            report_data: The report data to export
            report_type: Type of report (summary, sales, products, etc.)
            
        Returns:
            CSV file content as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        try:
            if report_type == "summary":
                cls._export_summary_to_csv(writer, report_data)
            else:
                # For other report types that haven't been migrated yet
                # This is a fallback for any report types not yet handled
                cls._export_generic_to_csv(writer, report_data, report_type)
            
            csv_content = output.getvalue()
            return csv_content.encode('utf-8')
            
        except Exception as e:
            logger.error(f"CSV export failed for {report_type}: {e}")
            raise
        finally:
            output.close()

    @classmethod
    def export_to_xlsx(cls, report_data: Dict[str, Any], report_type: str) -> bytes:
        """
        Export report data to Excel format.
        
        Args:
            report_data: The report data to export
            report_type: Type of report (summary, sales, products, etc.)
            
        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        
        # Set default title based on report type
        ws.title = f"{report_type.capitalize()} Report"
        
        # Define standard styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        try:
            if report_type == "summary":
                cls._export_summary_to_xlsx(ws, report_data, header_font, header_fill, header_alignment)
            else:
                # Fallback for other report types
                cls._export_generic_to_xlsx(ws, report_data, report_type, header_font, header_fill, header_alignment)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Excel export failed for {report_type}: {e}")
            raise

    @classmethod
    def export_to_pdf(
        cls,
        report_data: Dict[str, Any],
        report_type: str,
        title: Optional[str] = None,
        page_size=letter
    ) -> bytes:
        """
        Export report data to PDF format.
        
        Args:
            report_data: The report data to export
            report_type: Type of report (summary, sales, products, etc.)
            title: Optional custom title for the report
            page_size: Page size for the PDF (default: letter)
            
        Returns:
            PDF file content as bytes
        """
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=page_size,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Add custom title style
        styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=styles["Title"],
                alignment=TA_CENTER,
                fontSize=16,
                spaceAfter=30,
            )
        )
        
        # Add report title
        if not title:
            title = f"{report_type.capitalize()} Report"
        story.append(Paragraph(title, styles["CustomTitle"]))
        
        # Add date range if available
        if "date_range" in report_data:
            date_range = report_data["date_range"]
            date_text = f"Period: {date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"
            story.append(Paragraph(date_text, styles["Normal"]))
            story.append(Spacer(1, 0.2 * inch))
        
        try:
            if report_type == "summary":
                cls._export_summary_to_pdf(story, report_data, styles)
            else:
                # Fallback for other report types
                cls._export_generic_to_pdf(story, report_data, styles, report_type)
            
            doc.build(story)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"PDF export failed for {report_type}: {e}")
            raise
        finally:
            output.close()

    # Summary Report Export Methods
    @classmethod
    def _export_summary_to_csv(cls, writer, report_data: Dict[str, Any]):
        """Export summary report to CSV format."""
        # Header
        writer.writerow(["Summary Report"])
        writer.writerow([])
        
        # Date range
        if "date_range" in report_data:
            writer.writerow(["Date Range", report_data["date_range"].get("start", ""), 
                           "to", report_data["date_range"].get("end", "")])
            writer.writerow([])
        
        # Key Metrics
        writer.writerow(["Key Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Sales", f"${report_data.get('total_sales', 0):,.2f}"])
        writer.writerow(["Total Transactions", report_data.get("total_transactions", 0)])
        writer.writerow(["Average Order Value", f"${report_data.get('average_order_value', 0):,.2f}"])
        writer.writerow(["Total Customers", report_data.get("total_customers", 0)])
        writer.writerow([])
        
        # Financial Breakdown
        if "financial_breakdown" in report_data:
            writer.writerow(["Financial Breakdown"])
            fb = report_data["financial_breakdown"]
            writer.writerow(["Subtotal", f"${fb.get('subtotal', 0):,.2f}"])
            writer.writerow(["Discounts", f"${fb.get('discounts', 0):,.2f}"])
            writer.writerow(["Tax", f"${fb.get('tax', 0):,.2f}"])
            writer.writerow(["Tips", f"${fb.get('tips', 0):,.2f}"])
            writer.writerow(["Net Revenue", f"${fb.get('net_revenue', 0):,.2f}"])
            writer.writerow([])
        
        # Top Products
        if "top_products" in report_data and report_data["top_products"]:
            writer.writerow(["Top Products"])
            writer.writerow(["Product", "Quantity Sold", "Revenue"])
            for product in report_data["top_products"][:10]:
                writer.writerow([
                    product.get("name", ""),
                    product.get("quantity", 0),
                    f"${product.get('revenue', 0):,.2f}"
                ])
            writer.writerow([])
        
        # Sales by Category
        if "sales_by_category" in report_data and report_data["sales_by_category"]:
            writer.writerow(["Sales by Category"])
            writer.writerow(["Category", "Revenue", "Percentage"])
            for category in report_data["sales_by_category"]:
                writer.writerow([
                    category.get("name", ""),
                    f"${category.get('revenue', 0):,.2f}",
                    f"{category.get('percentage', 0):.1f}%"
                ])

    @classmethod
    def _export_summary_to_xlsx(cls, ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export summary report to Excel format."""
        row = 1
        
        # Title
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Summary Report"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 2
        
        # Date range
        if "date_range" in report_data:
            ws[f"A{row}"] = "Date Range:"
            ws[f"B{row}"] = f"{report_data['date_range'].get('start', '')} to {report_data['date_range'].get('end', '')}"
            row += 2
        
        # Key Metrics
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        metrics_headers = ["Metric", "Value"]
        for col, header in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        row += 1
        
        metrics = [
            ("Total Sales", f"${report_data.get('total_sales', 0):,.2f}"),
            ("Total Transactions", report_data.get("total_transactions", 0)),
            ("Average Order Value", f"${report_data.get('average_order_value', 0):,.2f}"),
            ("Total Customers", report_data.get("total_customers", 0)),
        ]
        
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1
        
        # Financial Breakdown
        if "financial_breakdown" in report_data:
            ws[f"A{row}"] = "Financial Breakdown"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            fb = report_data["financial_breakdown"]
            breakdown_items = [
                ("Subtotal", f"${fb.get('subtotal', 0):,.2f}"),
                ("Discounts", f"${fb.get('discounts', 0):,.2f}"),
                ("Tax", f"${fb.get('tax', 0):,.2f}"),
                ("Tips", f"${fb.get('tips', 0):,.2f}"),
                ("Net Revenue", f"${fb.get('net_revenue', 0):,.2f}"),
            ]
            
            for item, value in breakdown_items:
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=value)
                row += 1
            row += 1
        
        # Top Products
        if "top_products" in report_data and report_data["top_products"]:
            ws[f"A{row}"] = "Top Products"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            product_headers = ["Product", "Quantity Sold", "Revenue"]
            for col, header in enumerate(product_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1
            
            for product in report_data["top_products"][:10]:
                ws.cell(row=row, column=1, value=product.get("name", ""))
                ws.cell(row=row, column=2, value=product.get("quantity", 0))
                ws.cell(row=row, column=3, value=f"${product.get('revenue', 0):,.2f}")
                row += 1

    @classmethod
    def _export_summary_to_pdf(cls, story, report_data: Dict[str, Any], styles):
        """Export summary report to PDF format."""
        # Key Metrics Section
        story.append(Paragraph("Key Metrics", styles["Heading2"]))
        
        metrics_data = [
            ["Metric", "Value"],
            ["Total Sales", f"${report_data.get('total_sales', 0):,.2f}"],
            ["Total Transactions", str(report_data.get("total_transactions", 0))],
            ["Average Order Value", f"${report_data.get('average_order_value', 0):,.2f}"],
            ["Total Customers", str(report_data.get("total_customers", 0))],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3 * inch, 2 * inch])
        metrics_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ])
        )
        story.append(metrics_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Financial Breakdown
        if "financial_breakdown" in report_data:
            story.append(Paragraph("Financial Breakdown", styles["Heading2"]))
            fb = report_data["financial_breakdown"]
            
            breakdown_data = [
                ["Item", "Amount"],
                ["Subtotal", f"${fb.get('subtotal', 0):,.2f}"],
                ["Discounts", f"${fb.get('discounts', 0):,.2f}"],
                ["Tax", f"${fb.get('tax', 0):,.2f}"],
                ["Tips", f"${fb.get('tips', 0):,.2f}"],
                ["Net Revenue", f"${fb.get('net_revenue', 0):,.2f}"],
            ]
            
            breakdown_table = Table(breakdown_data, colWidths=[3 * inch, 2 * inch])
            breakdown_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ])
            )
            story.append(breakdown_table)
            story.append(Spacer(1, 0.3 * inch))
        
        # Top Products
        if "top_products" in report_data and report_data["top_products"]:
            story.append(Paragraph("Top 10 Products", styles["Heading2"]))
            
            product_data = [["Product", "Quantity", "Revenue"]]
            for product in report_data["top_products"][:10]:
                product_data.append([
                    product.get("name", ""),
                    str(product.get("quantity", 0)),
                    f"${product.get('revenue', 0):,.2f}"
                ])
            
            product_table = Table(product_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
            product_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ])
            )
            story.append(product_table)

    # Generic Export Methods (Fallbacks)
    @classmethod
    def _export_generic_to_csv(cls, writer, report_data: Dict[str, Any], report_type: str):
        """Generic CSV export for any report type."""
        writer.writerow([f"{report_type.capitalize()} Report"])
        writer.writerow([])
        
        # Attempt to export any dict/list data in a reasonable format
        for key, value in report_data.items():
            if isinstance(value, list) and value:
                writer.writerow([key.replace("_", " ").title()])
                if isinstance(value[0], dict):
                    # Write headers from first item's keys
                    headers = list(value[0].keys())
                    writer.writerow(headers)
                    # Write data rows
                    for item in value:
                        row = [item.get(h, "") for h in headers]
                        writer.writerow(row)
                else:
                    # Simple list
                    for item in value:
                        writer.writerow([str(item)])
                writer.writerow([])
            elif isinstance(value, (str, int, float, Decimal)):
                writer.writerow([key.replace("_", " ").title(), str(value)])

    @classmethod
    def _export_generic_to_xlsx(cls, ws, report_data: Dict[str, Any], report_type: str, 
                                header_font, header_fill, header_alignment):
        """Generic Excel export for any report type."""
        row = 1
        
        # Title
        ws[f"A{row}"] = f"{report_type.capitalize()} Report"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 2
        
        # Export data
        for key, value in report_data.items():
            if isinstance(value, list) and value:
                # Section header
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"A{row}"].font = Font(bold=True, size=12)
                row += 1
                
                if isinstance(value[0], dict):
                    # Table data
                    headers = list(value[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header.replace("_", " ").title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    row += 1
                    
                    for item in value:
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=str(item.get(header, "")))
                        row += 1
                row += 1
            elif isinstance(value, (str, int, float, Decimal)):
                ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row, column=2, value=str(value))
                row += 1

    @classmethod
    def _export_generic_to_pdf(cls, story, report_data: Dict[str, Any], styles, report_type: str):
        """Generic PDF export for any report type."""
        for key, value in report_data.items():
            if isinstance(value, list) and value:
                # Add section header
                story.append(Paragraph(key.replace("_", " ").title(), styles["Heading2"]))
                
                if isinstance(value[0], dict):
                    # Create table
                    headers = list(value[0].keys())
                    table_data = [[h.replace("_", " ").title() for h in headers]]
                    
                    for item in value[:50]:  # Limit to 50 rows for PDF
                        row = [str(item.get(h, "")) for h in headers]
                        table_data.append(row)
                    
                    table = Table(table_data)
                    table.setStyle(
                        TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ])
                    )
                    story.append(table)
                    story.append(Spacer(1, 0.2 * inch))
            elif key not in ["date_range"] and isinstance(value, (str, int, float, Decimal)):
                # Add key-value pairs
                story.append(Paragraph(f"<b>{key.replace('_', ' ').title()}:</b> {value}", styles["Normal"]))
                story.append(Spacer(1, 0.1 * inch))