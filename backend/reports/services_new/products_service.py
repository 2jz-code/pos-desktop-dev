"""
Products report service with generation and export functionality.
Refactored from monolithic services.py into modular, scalable components.
"""
import time
import logging
from decimal import Decimal
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

from django.db import transaction
from django.db.models import (
    Sum, Count, Avg, F, Q, Value
)
from django.db.models.functions import (
    TruncDate, TruncWeek, TruncMonth
)
from django.utils import timezone

# Export functionality imports
from openpyxl.styles import Font, PatternFill, Alignment

from orders.models import Order, OrderItem
from products.models import Product, Category
from .base import BaseReportService
from .timezone_utils import TimezoneUtils

logger = logging.getLogger(__name__)


class ProductsReportService(BaseReportService):
    """Service for generating and exporting products reports."""

    @staticmethod
    @transaction.atomic
    def generate_products_report(
        tenant,
        start_date: datetime,
        end_date: datetime,
        location_id: Optional[int] = None,
        category_id: Optional[int] = None,
        limit: Optional[int] = None,  # None means no limit (fetch all products)
        trend_period: str = "auto",  # "daily", "weekly", "monthly", "auto"
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate product performance report"""

        cache_key = ProductsReportService._generate_cache_key(
            "products",
            {
                "start_date": start_date,
                "end_date": end_date,
                "location_id": location_id,
                "category_id": category_id,
                "limit": limit,
                "trend_period": trend_period,
            },
        )

        if use_cache:
            cached_data = ProductsReportService._get_cached_report(cache_key, tenant)
            if cached_data:
                logger.info(f"Products report served from cache: {cache_key[:8]}...")
                return cached_data

        logger.info(f"Generating products report for {start_date} to {end_date}" + (f" at location {location_id}" if location_id else " (all locations)"))
        start_time = time.time()

        # Multi-location report generation
        if location_id is None:
            return ProductsReportService._generate_multi_location_products_report(
                tenant, start_date, end_date, category_id, limit, trend_period, use_cache
            )

        # Get base data
        order_items = ProductsReportService._get_base_order_items_queryset(
            tenant, start_date, end_date, location_id, category_id
        )

        # Calculate core metrics
        products_data = {}

        # Top products by revenue
        products_data["top_products"] = ProductsReportService._get_top_products_by_revenue(
            order_items, limit
        )

        # Best sellers by quantity
        products_data["best_sellers"] = ProductsReportService._get_best_sellers_by_quantity(
            order_items, limit
        )

        # Category performance
        products_data["category_performance"] = ProductsReportService._get_category_performance(
            order_items
        )

        # Product trends
        actual_period = ProductsReportService._determine_trend_period(
            start_date, end_date, trend_period
        )
        products_data["product_trends"] = ProductsReportService._get_product_trends(
            order_items, products_data["top_products"], actual_period
        )

        # Summary stats
        products_data["summary"] = ProductsReportService._calculate_summary_stats(order_items)

        # Add metadata
        products_data["generated_at"] = timezone.now().isoformat()
        products_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Add location metadata
        location_name = "All Locations"
        if location_id is not None:
            from settings.models import StoreLocation
            try:
                location = StoreLocation.objects.get(id=location_id, tenant=tenant)
                location_name = location.name
            except StoreLocation.DoesNotExist:
                location_name = f"Location ID {location_id}"

        products_data["location_info"] = {
            "location_id": location_id,
            "location_name": location_name,
            "is_multi_location": location_id is None
        }

        # Add tenant_id for export filtering
        products_data["tenant_id"] = tenant.id

        products_data["filters"] = {
            "location_id": location_id,
            "category_id": category_id,
            "limit": limit,
            "trend_period": trend_period,
            "actual_period": actual_period,
        }

        # Cache the result
        generation_time = time.time() - start_time
        ProductsReportService._cache_report(
            cache_key, products_data, tenant, report_type="products", ttl_hours=ProductsReportService.CACHE_TTL["products"]
        )

        logger.info(f"Products report generated in {generation_time:.2f}s")
        return products_data

    @staticmethod
    def _generate_multi_location_products_report(
        tenant,
        start_date: datetime,
        end_date: datetime,
        category_id: Optional[int] = None,
        limit: int = 10,
        trend_period: str = "auto",
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate products report for all locations with breakdown per location."""
        from settings.models import StoreLocation

        logger.info(f"Generating multi-location products report for {start_date} to {end_date}")

        # Get all active locations
        active_locations = StoreLocation.objects.filter(
            tenant=tenant,
            is_active=True
        ).order_by('name')

        # Generate report for each location
        location_reports = []
        for location in active_locations:
            location_report = ProductsReportService.generate_products_report(
                tenant=tenant,
                start_date=start_date,
                end_date=end_date,
                location_id=location.id,
                category_id=category_id,
                limit=limit,
                trend_period=trend_period,
                use_cache=use_cache
            )
            location_reports.append({
                "location_id": location.id,
                "location_name": location.name,
                "report_data": location_report
            })

        # Calculate consolidated totals
        consolidated_totals = ProductsReportService._calculate_consolidated_products_totals(location_reports, limit)

        return {
            "is_multi_location": True,
            "location_count": len(location_reports),
            "locations": location_reports,
            "consolidated": consolidated_totals,
            "generated_at": timezone.now().isoformat(),
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "tenant_id": tenant.id,
            "location_info": {
                "location_id": None,
                "location_name": "All Locations",
                "is_multi_location": True
            },
            "filters": {
                "category_id": category_id,
                "limit": limit,
                "trend_period": trend_period,
            }
        }

    @staticmethod
    def _calculate_consolidated_products_totals(location_reports: list, limit: int) -> Dict[str, Any]:
        """Calculate consolidated product totals across all locations."""
        from collections import defaultdict

        # Initialize aggregation structures
        products_revenue_agg = defaultdict(lambda: {'revenue': Decimal('0.00'), 'sold': 0, 'id': None})
        products_quantity_agg = defaultdict(lambda: {'revenue': Decimal('0.00'), 'sold': 0, 'id': None})
        category_performance_agg = defaultdict(lambda: {
            'revenue': Decimal('0.00'),
            'units_sold': 0,
            'unique_products': set()
        })
        trends_by_product_agg = defaultdict(lambda: defaultdict(int))

        # Summary totals
        total_revenue = Decimal('0.00')
        total_units_sold = 0
        all_products = set()

        # Aggregate across locations
        for location_data in location_reports:
            loc_report = location_data.get('report_data', {})

            # Aggregate summary
            summary = loc_report.get('summary', {})
            total_revenue += Decimal(str(summary.get('total_revenue', 0)))
            total_units_sold += summary.get('total_units_sold', 0)

            # Aggregate top products by revenue
            for product in loc_report.get('top_products', []):
                name = product['name']
                products_revenue_agg[name]['revenue'] += Decimal(str(product.get('revenue', 0)))
                products_revenue_agg[name]['sold'] += product.get('sold', 0)
                products_revenue_agg[name]['id'] = product.get('id')
                all_products.add(product.get('id'))

            # Aggregate best sellers by quantity
            for product in loc_report.get('best_sellers', []):
                name = product['name']
                products_quantity_agg[name]['revenue'] += Decimal(str(product.get('revenue', 0)))
                products_quantity_agg[name]['sold'] += product.get('sold', 0)
                products_quantity_agg[name]['id'] = product.get('id')

            # Aggregate category performance
            for category in loc_report.get('category_performance', []):
                cat_name = category['category']
                category_performance_agg[cat_name]['revenue'] += Decimal(str(category.get('revenue', 0)))
                category_performance_agg[cat_name]['units_sold'] += category.get('units_sold', 0)
                # Track unique products per category (using sets)
                # Note: We can't easily track unique products across locations without product IDs per category

            # Aggregate product trends
            for product_name, trends in loc_report.get('product_trends', {}).items():
                for trend in trends:
                    date = trend['date']
                    sold = trend['sold']
                    trends_by_product_agg[product_name][date] += sold

        # Build top_products list (sorted by revenue, limited)
        top_products = [
            {
                'name': name,
                'id': data['id'],
                'total_revenue': float(data['revenue']),
                'revenue': float(data['revenue']),
                'total_sold': data['sold'],
                'sold': data['sold'],
                'avg_price': float(data['revenue'] / data['sold']) if data['sold'] > 0 else 0
            }
            for name, data in sorted(
                products_revenue_agg.items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )[:limit]
        ]

        # Build best_sellers list (sorted by quantity, limited)
        best_sellers = [
            {
                'name': name,
                'id': data['id'],
                'total_sold': data['sold'],
                'sold': data['sold'],
                'total_revenue': float(data['revenue']),
                'revenue': float(data['revenue'])
            }
            for name, data in sorted(
                products_quantity_agg.items(),
                key=lambda x: x[1]['sold'],
                reverse=True
            )[:limit]
        ]

        # Build category_performance list (sorted by revenue)
        category_performance = [
            {
                'category': category,
                'revenue': float(data['revenue']),
                'units_sold': data['units_sold'],
                'unique_products': len(data['unique_products']) if data['unique_products'] else 0
            }
            for category, data in sorted(
                category_performance_agg.items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )
        ]

        # Build product_trends dict
        product_trends = {}
        for product_name, trends in trends_by_product_agg.items():
            product_trends[product_name] = [
                {'date': date, 'sold': sold}
                for date, sold in sorted(trends.items())
            ]

        # Calculate summary
        avg_price_per_unit = float(total_revenue / total_units_sold) if total_units_sold > 0 else 0

        return {
            'top_products': top_products,
            'best_sellers': best_sellers,
            'category_performance': category_performance,
            'product_trends': product_trends,
            'summary': {
                'total_products': len(all_products),
                'unique_products_sold': len(all_products),
                'total_revenue': float(total_revenue),
                'total_units_sold': total_units_sold,
                'avg_price_per_unit': avg_price_per_unit
            }
        }

    @staticmethod
    def _get_base_order_items_queryset(
        tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None, category_id: Optional[int] = None
    ):
        """Get optimized base queryset for order items in date range."""
        filters = {
            "order__tenant": tenant,
            "order__status": Order.OrderStatus.COMPLETED,
            "order__completed_at__range": (start_date, end_date),
            "order__subtotal__gt": 0,  # Exclude orders with $0.00 subtotals
        }

        if location_id is not None:
            filters["order__store_location_id"] = location_id

        queryset = OrderItem.objects.filter(**filters).select_related("product", "product__category", "order__store_location")

        # Filter by category if specified
        if category_id:
            queryset = queryset.filter(product__category_id=category_id)

        return queryset

    @staticmethod
    def _get_top_products_by_revenue(order_items, limit: int) -> list:
        """Get top products by revenue with detailed metrics."""
        top_products = (
            order_items.annotate(revenue=F("quantity") * F("price_at_sale"))
            .values("product__name", "product__id")
            .annotate(
                total_revenue=Sum("revenue"),
                total_sold=Sum("quantity"),
                avg_price=Avg("price_at_sale"),
            )
            .order_by("-total_revenue")[:limit]
        )

        return [
            {
                "name": item["product__name"],
                "id": item["product__id"],
                "total_revenue": float(item["total_revenue"] or 0),
                "revenue": float(item["total_revenue"] or 0),  # Alias for CSV compatibility
                "total_sold": item["total_sold"],
                "sold": item["total_sold"],  # Alias for CSV compatibility
                "avg_price": float(item["avg_price"] or 0),
            }
            for item in top_products
        ]

    @staticmethod
    def _get_best_sellers_by_quantity(order_items, limit: int) -> list:
        """Get best selling products by quantity."""
        best_sellers = (
            order_items.values("product__name", "product__id")
            .annotate(
                total_sold=Sum("quantity"),
                total_revenue=Sum(F("quantity") * F("price_at_sale")),
            )
            .order_by("-total_sold")[:limit]
        )

        return [
            {
                "name": item["product__name"],
                "id": item["product__id"],
                "total_sold": item["total_sold"],
                "sold": item["total_sold"],  # Alias for CSV compatibility
                "total_revenue": float(item["total_revenue"] or 0),
                "revenue": float(item["total_revenue"] or 0),  # Alias for CSV compatibility
            }
            for item in best_sellers
        ]

    @staticmethod
    def _get_category_performance(order_items) -> list:
        """Get performance metrics by product category."""
        category_performance = (
            order_items.values("product__category__name")
            .annotate(
                revenue=Sum(F("quantity") * F("price_at_sale")),
                units_sold=Sum("quantity"),
                unique_products=Count("product__id", distinct=True),
            )
            .order_by("-revenue")
        )

        return [
            {
                "category": item["product__category__name"] or "Uncategorized",
                "revenue": float(item["revenue"] or 0),
                "units_sold": item["units_sold"] or 0,
                "unique_products": item["unique_products"],
            }
            for item in category_performance
        ]

    @staticmethod
    def _determine_trend_period(
        start_date: datetime, end_date: datetime, trend_period: str
    ) -> str:
        """Determine the actual trend period based on date range and user preference."""
        date_diff = (end_date - start_date).days

        if trend_period == "auto":
            if date_diff <= 7:
                return "daily"
            elif date_diff <= 60:
                return "weekly"
            else:
                return "monthly"
        else:
            return trend_period

    @staticmethod
    def _get_product_trends(order_items, top_products: list, actual_period: str) -> dict:
        """Get sales trends for top products over time."""
        if not top_products:
            return {}

        # Get top 5 products for trends to avoid chart clutter
        top_product_ids = [p["id"] for p in top_products[:5]]
        
        trend_items = order_items.filter(product__in=top_product_ids)

        # Apply correct grouping based on period
        if actual_period == "weekly":
            product_trends = (
                trend_items
                .annotate(period=TruncWeek("order__completed_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )
        elif actual_period == "monthly":
            product_trends = (
                trend_items
                .annotate(period=TruncMonth("order__completed_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )
        else:  # daily
            product_trends = (
                trend_items
                .annotate(period=TimezoneUtils.trunc_date_local("order__completed_at"))
                .values("product__name", "period")
                .annotate(sold=Sum("quantity"))
                .order_by("period")
            )

        # Group trends by product
        trends_by_product = {}
        for trend in product_trends:
            product_name = trend["product__name"]
            if product_name not in trends_by_product:
                trends_by_product[product_name] = []

            trends_by_product[product_name].append(
                {"date": trend["period"].strftime("%Y-%m-%d"), "sold": trend["sold"]}
            )

        return trends_by_product

    @staticmethod
    def _calculate_summary_stats(order_items) -> dict:
        """Calculate summary statistics for the products report."""
        total_revenue = float(
            order_items.aggregate(total=Sum(F("quantity") * F("price_at_sale")))["total"] or 0
        )
        total_units_sold = order_items.aggregate(total=Sum("quantity"))["total"] or 0
        total_products = order_items.values("product__id").distinct().count()
        
        # Calculate average price per unit
        avg_price_per_unit = total_revenue / total_units_sold if total_units_sold > 0 else 0
        
        return {
            "total_products": total_products,
            "unique_products_sold": total_products,  # Alias for consistency
            "total_revenue": total_revenue,
            "total_units_sold": total_units_sold,
            "avg_price_per_unit": avg_price_per_unit,
        }

    # Export Methods
    @staticmethod
    def export_products_to_pdf(report_data: Dict[str, Any], story, styles):
        """Export products report to PDF format."""
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            ProductsReportService._export_multi_location_products_to_pdf(story, report_data, styles)
        else:
            ProductsReportService._export_products_to_pdf(story, report_data, styles)

    @staticmethod
    def export_products_to_xlsx(report_data: Dict[str, Any], ws_or_wb, header_font, header_fill, header_alignment):
        """Export products report to Excel format.

        For single-location reports, ws_or_wb is a worksheet.
        For multi-location reports, ws_or_wb is a workbook.
        """
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            ProductsReportService._export_multi_location_products_to_xlsx(
                ws_or_wb, report_data, header_font, header_fill, header_alignment
            )
        else:
            ProductsReportService._export_products_to_xlsx(
                ws_or_wb, report_data, header_font, header_fill, header_alignment
            )

    @staticmethod
    def export_products_to_csv(report_data: Dict[str, Any]) -> bytes:
        """Export products report to CSV format."""
        import io
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            ProductsReportService._export_multi_location_products_to_csv(writer, report_data)
        else:
            ProductsReportService._export_products_to_csv(writer, report_data)
        
        csv_bytes = output.getvalue().encode('utf-8')
        output.close()
        return csv_bytes

    # Multi-Location Export Methods
    @staticmethod
    def _export_multi_location_products_to_xlsx(wb, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export multi-location products report to Excel with separate sheets."""
        from openpyxl.styles import Border, Side

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Create consolidated summary sheet
        summary_ws = wb.create_sheet(title="Consolidated Summary", index=0)

        # Add consolidated summary header
        summary_ws.append(["MULTI-LOCATION PRODUCTS REPORT - CONSOLIDATED SUMMARY"])
        summary_ws['A1'].font = Font(size=14, bold=True)

        date_range = report_data.get('date_range', {})
        summary_ws.append([f"Period: {date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"])
        summary_ws.append([f"Locations: {report_data.get('location_count', 0)} active locations"])
        summary_ws.append([])

        # Consolidated metrics
        consolidated = report_data.get('consolidated', {})
        summary = consolidated.get('summary', {})

        summary_ws.append(["OVERALL METRICS"])
        summary_ws['A5'].font = header_font
        summary_ws.append(["Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"])
        summary_ws.append(["Total Units Sold", f"{summary.get('total_units_sold', 0):,}"])
        summary_ws.append(["Unique Products Sold", f"{summary.get('unique_products_sold', 0):,}"])
        summary_ws.append(["Avg Price Per Unit", f"${summary.get('avg_price_per_unit', 0):,.2f}"])
        summary_ws.append([])

        # Top products consolidated
        summary_ws.append(["TOP PRODUCTS BY REVENUE (CONSOLIDATED)"])
        row = summary_ws.max_row
        summary_ws[f'A{row}'].font = header_font

        # Headers
        headers = ["Rank", "Product", "Total Revenue", "Units Sold", "Avg Price"]
        summary_ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = summary_ws.cell(row=summary_ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for rank, product in enumerate(consolidated.get('top_products', []), 1):
            summary_ws.append([
                rank,
                product.get('name', ''),
                product.get('revenue', 0),
                product.get('sold', 0),
                product.get('avg_price', 0),
            ])
            for col in range(1, 6):
                summary_ws.cell(row=summary_ws.max_row, column=col).border = thin_border

        summary_ws.append([])

        # Location comparison table
        summary_ws.append(["LOCATION COMPARISON"])
        row = summary_ws.max_row
        summary_ws[f'A{row}'].font = header_font

        comparison_headers = ["Location", "Total Revenue", "Units Sold", "Unique Products", "Avg Price/Unit"]
        summary_ws.append(comparison_headers)
        for col, _ in enumerate(comparison_headers, 1):
            cell = summary_ws.cell(row=summary_ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Location comparison rows
        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_summary = location_data.get('report_data', {}).get('summary', {})

            summary_ws.append([
                location_name,
                loc_summary.get('total_revenue', 0),
                loc_summary.get('total_units_sold', 0),
                loc_summary.get('unique_products_sold', 0),
                loc_summary.get('avg_price_per_unit', 0),
            ])
            for col in range(1, 6):
                summary_ws.cell(row=summary_ws.max_row, column=col).border = thin_border

        # Auto-size columns
        for col in summary_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column].width = adjusted_width

        # Create individual location sheets
        for i, location_data in enumerate(locations, 1):
            location_name = location_data.get('location_name', f'Location {i}')
            loc_report = location_data.get('report_data', {})

            # Create sheet for this location
            safe_sheet_name = location_name[:31]  # Excel sheet name limit
            loc_ws = wb.create_sheet(title=safe_sheet_name)

            # Export individual location data using the single-location method
            ProductsReportService._export_products_to_xlsx(
                loc_ws, loc_report, header_font, header_fill, header_alignment
            )

    @staticmethod
    def _export_multi_location_products_to_csv(writer, report_data: Dict[str, Any]):
        """Export multi-location products report to CSV with sectioned format."""

        # Header
        writer.writerow(["MULTI-LOCATION PRODUCTS REPORT - CONSOLIDATED"])
        date_range = report_data.get('date_range', {})
        writer.writerow([f"Period: {date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"])
        writer.writerow([f"Locations: {report_data.get('location_count', 0)} active locations"])
        writer.writerow([])

        # Consolidated summary
        consolidated = report_data.get('consolidated', {})
        summary = consolidated.get('summary', {})

        writer.writerow(["OVERALL METRICS"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"])
        writer.writerow(["Total Units Sold", f"{summary.get('total_units_sold', 0):,}"])
        writer.writerow(["Unique Products Sold", f"{summary.get('unique_products_sold', 0):,}"])
        writer.writerow(["Avg Price Per Unit", f"${summary.get('avg_price_per_unit', 0):,.2f}"])
        writer.writerow([])

        # Top products consolidated
        writer.writerow(["TOP PRODUCTS BY REVENUE (CONSOLIDATED)"])
        writer.writerow(["Rank", "Product", "Total Revenue", "Units Sold", "Avg Price"])
        for rank, product in enumerate(consolidated.get('top_products', []), 1):
            writer.writerow([
                rank,
                product.get('name', ''),
                f"${product.get('revenue', 0):,.2f}",
                f"{product.get('sold', 0):,}",
                f"${product.get('avg_price', 0):,.2f}",
            ])
        writer.writerow([])

        # Location comparison
        writer.writerow(["LOCATION COMPARISON"])
        writer.writerow(["Location", "Total Revenue", "Units Sold", "Unique Products", "Avg Price/Unit"])

        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_summary = location_data.get('report_data', {}).get('summary', {})

            writer.writerow([
                location_name,
                f"${loc_summary.get('total_revenue', 0):,.2f}",
                f"{loc_summary.get('total_units_sold', 0):,}",
                f"{loc_summary.get('unique_products_sold', 0):,}",
                f"${loc_summary.get('avg_price_per_unit', 0):,.2f}",
            ])

        writer.writerow([])
        writer.writerow([])

        # Individual location details
        writer.writerow(["=" * 100])
        writer.writerow(["=== DETAILED BREAKDOWN BY LOCATION ==="])
        writer.writerow(["=" * 100])
        writer.writerow([])

        for i, location_data in enumerate(locations, 1):
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            writer.writerow(["=" * 100])
            writer.writerow([f"LOCATION {i}: {location_name}"])
            writer.writerow(["=" * 100])
            writer.writerow([])

            # Export this location's data using single-location method
            ProductsReportService._export_products_to_csv(writer, loc_report)

            writer.writerow([])
            writer.writerow([])

    @staticmethod
    def _export_multi_location_products_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export multi-location products report to PDF with multi-page format."""
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak

        # Title page / Executive summary
        story.append(Paragraph("Multi-Location Products Report", styles["Title"]))
        story.append(Spacer(1, 12))

        date_range = report_data.get('date_range', {})
        story.append(Paragraph(
            f"Period: {date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}",
            styles["Normal"]
        ))
        story.append(Paragraph(
            f"Locations: {report_data.get('location_count', 0)} active locations",
            styles["Normal"]
        ))
        story.append(Spacer(1, 24))

        # Consolidated metrics
        consolidated = report_data.get('consolidated', {})
        summary = consolidated.get('summary', {})

        story.append(Paragraph("Overall Metrics", styles["Heading1"]))
        story.append(Spacer(1, 12))

        metrics_data = [
            ["Metric", "Value"],
            ["Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"],
            ["Total Units Sold", f"{summary.get('total_units_sold', 0):,}"],
            ["Unique Products Sold", f"{summary.get('unique_products_sold', 0):,}"],
            ["Avg Price Per Unit", f"${summary.get('avg_price_per_unit', 0):,.2f}"],
        ]

        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 24))

        # Top products consolidated
        story.append(Paragraph("Top Products by Revenue (Consolidated)", styles["Heading2"]))
        story.append(Spacer(1, 12))

        top_products_data = [["Rank", "Product", "Revenue", "Units Sold", "Avg Price"]]
        for rank, product in enumerate(consolidated.get('top_products', []), 1):
            top_products_data.append([
                str(rank),
                product.get('name', '')[:30],
                f"${product.get('revenue', 0):,.2f}",
                f"{product.get('sold', 0):,}",
                f"${product.get('avg_price', 0):,.2f}",
            ])

        top_products_table = Table(top_products_data, colWidths=[0.5*inch, 2.5*inch, 1.2*inch, 1*inch, 1*inch])
        top_products_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(top_products_table)
        story.append(Spacer(1, 24))

        # Location comparison
        story.append(Paragraph("Location Comparison", styles["Heading2"]))
        story.append(Spacer(1, 12))

        comparison_data = [["Location", "Revenue", "Units Sold", "Unique Products", "Avg Price/Unit"]]
        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_summary = location_data.get('report_data', {}).get('summary', {})

            comparison_data.append([
                location_name,
                f"${loc_summary.get('total_revenue', 0):,.2f}",
                f"{loc_summary.get('total_units_sold', 0):,}",
                f"{loc_summary.get('unique_products_sold', 0):,}",
                f"${loc_summary.get('avg_price_per_unit', 0):,.2f}",
            ])

        comparison_table = Table(comparison_data, colWidths=[2*inch, 1.3*inch, 1*inch, 1.2*inch, 1.2*inch])
        comparison_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(comparison_table)
        story.append(PageBreak())

        # Individual location pages
        for i, location_data in enumerate(locations):
            location_name = location_data.get('location_name', f'Location {i+1}')
            loc_report = location_data.get('report_data', {})

            story.append(Paragraph(f"Location: {location_name}", styles["Heading1"]))
            story.append(Spacer(1, 12))

            # Export individual location data using the single-location method
            ProductsReportService._export_products_to_pdf(story, loc_report, styles)

            # Add page break between locations (except for the last one)
            if i < len(locations) - 1:
                story.append(PageBreak())

    @staticmethod
    def _get_all_products_for_export(start_date, end_date, category_id=None, tenant_id=None, location_id=None):
        """Get ALL products sold in the time period for comprehensive export"""
        from orders.models import OrderItem, Order
        from django.db.models import Sum, Count, Avg, F

        # Get base queryset - same logic as main report but no limit
        filters = {
            'order__status': Order.OrderStatus.COMPLETED,
            'order__completed_at__range': (start_date, end_date),
            'order__subtotal__gt': 0,
        }

        # Add tenant filter
        if tenant_id:
            filters['order__tenant_id'] = tenant_id

        # Add location filter if specified
        if location_id is not None:
            filters['order__store_location_id'] = location_id

        order_items = OrderItem.objects.filter(**filters).select_related("product", "product__category")

        if category_id:
            order_items = order_items.filter(product__category_id=category_id)

        # Get all products with their performance data
        all_products = (
            order_items.annotate(revenue=F("quantity") * F("price_at_sale"))
            .values("product__name", "product__id")
            .annotate(
                total_revenue=Sum("revenue"),
                total_sold=Sum("quantity"),
                avg_price=Avg("price_at_sale"),
            )
            .order_by("-total_revenue")  # Rank by revenue
        )

        return [
            {
                "name": item["product__name"],
                "id": item["product__id"],
                "total_revenue": float(item["total_revenue"] or 0),
                "total_sold": item["total_sold"],
                "avg_price": float(item["avg_price"] or 0),
            }
            for item in all_products
        ]

    @staticmethod
    def _export_products_to_csv(writer, report_data: Dict[str, Any]):
        """Export comprehensive products report to CSV with enhanced data"""
        # Header
        writer.writerow(["Products Report"])
        writer.writerow(["Generated:", report_data.get("generated_at", "N/A")])
        
        # Extract and format date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "N/A")
        end_str = date_range.get("end", "N/A")
        
        # Clean up date strings (remove timezone info for cleaner display)
        if start_str and start_str != "N/A":
            start_str = start_str.split('T')[0]
        if end_str and end_str != "N/A":
            end_str = end_str.split('T')[0]
            
        writer.writerow(["Date Range:", f"{start_str} to {end_str}"])

        # Add location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        writer.writerow(["Location:", location_name])

        # Add filter information
        filters = report_data.get("filters", {})
        if filters.get("category_id"):
            writer.writerow(["Category Filter:", f"Category ID {filters['category_id']}"])
        if filters.get("trend_period"):
            writer.writerow(["Trend Period:", f"{filters['trend_period']} (Actual: {filters.get('actual_period', 'N/A')})"])
        
        writer.writerow([])  # Empty row

        # === SUMMARY METRICS ===
        summary = report_data.get("summary", {})
        writer.writerow(["=== SUMMARY METRICS ==="])
        writer.writerow(["Total Products Sold (Unique SKUs):", summary.get("total_products", 0)])
        writer.writerow(["Total Units Sold:", summary.get("total_units_sold", 0)])
        writer.writerow(["Total Revenue:", f"${summary.get('total_revenue', 0):,.2f}"])
        
        # Calculate average revenue per product
        total_products = summary.get("total_products", 1)  # Avoid division by zero
        avg_revenue_per_product = summary.get("total_revenue", 0) / total_products
        writer.writerow(["Average Revenue per Product:", f"${avg_revenue_per_product:,.2f}"])
        
        writer.writerow([])  # Empty row

        # === CATEGORY PERFORMANCE ===
        writer.writerow(["=== CATEGORY PERFORMANCE ==="])
        writer.writerow([
            "Category", "Total Revenue", "Units Sold", "Unique Products", 
            "Avg Revenue per Product", "Avg Units per Product", "Category Share %"
        ])

        categories = report_data.get("category_performance", [])
        total_category_revenue = sum(c.get("revenue", 0) for c in categories)
        
        for category in categories:
            revenue = category.get("revenue", 0)
            units_sold = category.get("units_sold", 0)
            unique_products = category.get("unique_products", 1)  # Avoid division by zero
            
            avg_revenue_per_product = revenue / unique_products
            avg_units_per_product = units_sold / unique_products
            category_share = (revenue / total_category_revenue * 100) if total_category_revenue > 0 else 0

            writer.writerow([
                category.get("category", "Uncategorized"),
                f"${revenue:,.2f}",
                units_sold,
                unique_products,
                f"${avg_revenue_per_product:,.2f}",
                f"{avg_units_per_product:.1f}",
                f"{category_share:.1f}%"
            ])

        writer.writerow([])  # Empty row

        # === BUSINESS INSIGHTS ===
        writer.writerow(["=== BUSINESS INSIGHTS ==="])
        
        # Top performer
        top_products = report_data.get("top_products", [])
        if top_products:
            top_performer = top_products[0]
            writer.writerow(["Top Revenue Product:", f"{top_performer.get('name')} (${top_performer.get('revenue', 0):,.2f})"])
            
            # Best seller by quantity
            best_seller = max(top_products, key=lambda x: x.get('sold', 0))
            writer.writerow(["Best Seller by Quantity:", f"{best_seller.get('name')} ({best_seller.get('sold', 0)} units)"])
        
        # Category insights
        if categories:
            top_category = max(categories, key=lambda x: x.get('revenue', 0))
            writer.writerow(["Top Revenue Category:", f"{top_category.get('category')} (${top_category.get('revenue', 0):,.2f})"])
        
        # Performance ratios
        if summary.get("total_units_sold", 0) > 0:
            avg_units_per_sku = summary.get("total_units_sold", 0) / summary.get("total_products", 1)
            writer.writerow(["Average Units Sold per SKU:", f"{avg_units_per_sku:.1f}"])
        
        writer.writerow([])  # Empty row

        # === ALL PRODUCTS PERFORMANCE (RANKED) ===  
        writer.writerow(["=== ALL PRODUCTS PERFORMANCE (RANKED BY REVENUE) ==="])
        writer.writerow([
            "Rank", "Product Name", "Product ID", "Units Sold", "Total Revenue", "Average Sale Price"
        ])

        # Extract date range and filters to get ALL products
        try:
            from datetime import datetime
            
            # Parse dates from report_data
            date_range = report_data.get("date_range", {})
            start_str = date_range.get("start")
            end_str = date_range.get("end")
            
            if start_str and end_str:
                # Convert back to datetime objects
                start_date = datetime.fromisoformat(start_str.replace("Z", "+00:00"))
                end_date = datetime.fromisoformat(end_str.replace("Z", "+00:00"))
                
                # Get category filter if present
                filters = report_data.get("filters", {})
                category_id = filters.get("category_id")

                # Extract tenant and location filters
                tenant_id = report_data.get('tenant_id')
                location_id = report_data.get('location_info', {}).get('location_id')

                # Get ALL products for ranking
                all_products = ProductsReportService._get_all_products_for_export(
                    start_date, end_date, category_id, tenant_id, location_id
                )
                
                # Write all products ranked by revenue
                for i, product in enumerate(all_products, 1):
                    revenue = product.get("revenue", 0)
                    units_sold = product.get("sold", 0)
                    avg_price = product.get("avg_price", 0)

                    writer.writerow([
                        i,  # Rank
                        product.get("name", "N/A"),
                        product.get("id", "N/A"),
                        units_sold,
                        f"${revenue:,.2f}",
                        f"${avg_price:.2f}"
                    ])
            else:
                writer.writerow(["Error: Could not parse date range for complete product listing"])
                
        except Exception as e:
            writer.writerow([f"Error generating complete product ranking: {str(e)}"])
            # Fallback to limited data
            top_products = report_data.get("top_products", [])
            for i, product in enumerate(top_products, 1):
                revenue = product.get("revenue", 0)
                units_sold = product.get("sold", 0)
                avg_price = product.get("avg_price", 0)

                writer.writerow([
                    i,  # Rank
                    product.get("name", "N/A"),
                    product.get("id", "N/A"),
                    units_sold,
                    f"${revenue:,.2f}",
                    f"${avg_price:.2f}"
                ])

        writer.writerow(["Report Generation Complete"])

    @staticmethod
    def _export_products_to_xlsx(ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export comprehensive products report to Excel matching CSV format (Excel-safe headers)"""
        from django.utils import timezone
        from datetime import datetime
        
        # Define number formats (using working patterns)
        currency_format = '"$"#,##0.00'
        number_format = '#,##0'
        decimal_format = '0.0'
        percentage_format = '0.0%'
        
        row = 1
        
        # --- HEADER INFORMATION ---
        ws.cell(row=row, column=1, value="--- PRODUCTS REPORT ---")
        row += 1
        
        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "N/A")
        end_str = date_range.get("end", "N/A")
        
        if start_str != "N/A" and end_str != "N/A":
            try:
                start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
                start_str = start_date.strftime("%B %d, %Y")
                end_str = end_date.strftime("%B %d, %Y")
            except:
                pass
        
        ws.cell(row=row, column=1, value="Date Range:")
        ws.cell(row=row, column=2, value=f"{start_str} to {end_str}")
        row += 1

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        ws.cell(row=row, column=1, value="Location:")
        ws.cell(row=row, column=2, value=location_name)
        row += 1

        # Filter information
        filters = report_data.get("filters", {})
        if filters.get("category_id"):
            ws.cell(row=row, column=1, value="Category Filter:")
            ws.cell(row=row, column=2, value=f"Category ID {filters['category_id']}")
            row += 1
        if filters.get("trend_period"):
            ws.cell(row=row, column=1, value="Trend Period:")
            ws.cell(row=row, column=2, value=f"{filters['trend_period']} (Actual: {filters.get('actual_period', 'N/A')})")
            row += 1
        
        row += 1  # Empty row
        
        # --- SUMMARY METRICS ---
        ws.cell(row=row, column=1, value="--- SUMMARY METRICS ---")
        row += 1
        
        summary = report_data.get("summary", {})
        
        # Calculate average revenue per product to match CSV
        total_products = summary.get("total_products", 1)
        avg_revenue_per_product = summary.get("total_revenue", 0) / total_products
        
        metrics = [
            ("Total Products Sold (Unique SKUs)", int(summary.get('total_products', 0)), "number"),
            ("Total Units Sold", int(summary.get('total_units_sold', 0)), "number"),
            ("Total Revenue", float(summary.get('total_revenue', 0)), "currency"),
            ("Average Revenue per Product", float(avg_revenue_per_product), "currency"),
        ]
        
        for metric_name, metric_value, format_type in metrics:
            ws.cell(row=row, column=1, value=metric_name)
            cell = ws.cell(row=row, column=2, value=metric_value)
            if format_type == "currency":
                cell.number_format = currency_format
            elif format_type == "number":
                cell.number_format = number_format
            row += 1
        
        row += 1  # Empty row
        
        # --- CATEGORY PERFORMANCE ---
        category_performance = report_data.get("category_performance", [])
        if category_performance:
            ws.cell(row=row, column=1, value="--- CATEGORY PERFORMANCE ---")
            row += 1
            
            # Headers
            headers = ["Category", "Revenue", "Units Sold", "Products", "Avg Revenue/Product", "Avg Units/Product", "Share of Total"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            
            # Calculate total for percentage
            total_category_revenue = sum(cat.get("revenue", 0) for cat in category_performance)
            
            # Data rows
            for category in category_performance:
                revenue = category.get("revenue", 0)
                units_sold = category.get("units_sold", 0)
                unique_products = category.get("unique_products", 1)  # Avoid division by zero
                
                avg_revenue_per_product = revenue / unique_products
                avg_units_per_product = units_sold / unique_products
                category_share = (revenue / total_category_revenue * 100) if total_category_revenue > 0 else 0

                # Write data with proper formatting
                ws.cell(row=row, column=1, value=category.get("category", "Uncategorized"))
                
                revenue_cell = ws.cell(row=row, column=2, value=float(revenue))
                revenue_cell.number_format = currency_format
                
                ws.cell(row=row, column=3, value=int(units_sold))
                ws.cell(row=row, column=4, value=int(unique_products))
                
                avg_rev_cell = ws.cell(row=row, column=5, value=float(avg_revenue_per_product))
                avg_rev_cell.number_format = currency_format
                
                avg_units_cell = ws.cell(row=row, column=6, value=float(avg_units_per_product))
                avg_units_cell.number_format = decimal_format
                
                share_cell = ws.cell(row=row, column=7, value=float(category_share)/100)
                share_cell.number_format = percentage_format
                
                row += 1
            
            row += 1  # Empty row
        
        # --- BUSINESS INSIGHTS ---
        ws.cell(row=row, column=1, value="--- BUSINESS INSIGHTS ---")
        row += 1
        
        # Top performer
        top_products = report_data.get("top_products", [])
        if top_products:
            top_performer = top_products[0]
            ws.cell(row=row, column=1, value="Top Revenue Product:")
            ws.cell(row=row, column=2, value=f"{top_performer.get('name', 'N/A')} (${top_performer.get('revenue', 0):,.2f})")
            row += 1
            
        # Best seller by quantity (same logic as CSV)
        if top_products:
            best_seller = max(top_products, key=lambda x: x.get('sold', 0))
            ws.cell(row=row, column=1, value="Best Seller by Quantity:")
            ws.cell(row=row, column=2, value=f"{best_seller.get('name', 'N/A')} ({best_seller.get('sold', 0):,} units)")
            row += 1
        
        # Category insights
        if category_performance:
            top_category = max(category_performance, key=lambda x: x.get("revenue", 0))
            ws.cell(row=row, column=1, value="Top Revenue Category:")
            ws.cell(row=row, column=2, value=f"{top_category.get('category', 'N/A')} (${top_category.get('revenue', 0):,.2f})")
            row += 1
        
        row += 1  # Empty row
        
        # --- ALL PRODUCTS SOLD (Ranked by Revenue) ---
        ws.cell(row=row, column=1, value="--- ALL PRODUCTS SOLD (Ranked by Revenue) ---")
        row += 1
        
        # Get ALL products for comprehensive ranking
        if date_range.get("start") and date_range.get("end"):
            try:
                start_date_obj = datetime.fromisoformat(date_range["start"].replace('Z', '+00:00'))
                end_date_obj = datetime.fromisoformat(date_range["end"].replace('Z', '+00:00'))
                category_id = filters.get("category_id") if filters.get("category_id") != "all" else None

                # Extract tenant and location filters
                tenant_id = report_data.get('tenant_id')
                location_id = report_data.get('location_info', {}).get('location_id')

                all_products = ProductsReportService._get_all_products_for_export(
                    start_date_obj, end_date_obj, category_id, tenant_id, location_id
                )
                
                if all_products:
                    # Headers for ranking table
                    ranking_headers = ["Rank", "Product Name", "Product ID", "Units Sold", "Revenue", "Avg Price"]
                    for col, header in enumerate(ranking_headers, 1):
                        ws.cell(row=row, column=col, value=header)
                    row += 1
                    
                    # Data rows for ALL products
                    for i, product in enumerate(all_products, 1):
                        ws.cell(row=row, column=1, value=int(i))
                        ws.cell(row=row, column=2, value=str(product.get("name", "N/A")))
                        ws.cell(row=row, column=3, value=str(product.get("id", "N/A")))
                        ws.cell(row=row, column=4, value=int(product.get("total_sold", 0)))
                        
                        revenue_cell = ws.cell(row=row, column=5, value=float(product.get("total_revenue", 0)))
                        revenue_cell.number_format = currency_format
                        
                        price_cell = ws.cell(row=row, column=6, value=float(product.get("avg_price", 0)))
                        price_cell.number_format = currency_format
                        
                        row += 1
                        
            except Exception as e:
                # Fallback to limited data
                ws.cell(row=row, column=1, value="Error loading comprehensive data. Showing limited results:")
                row += 1
                
                if top_products:
                    # Headers
                    ranking_headers = ["Rank", "Product Name", "Units Sold", "Revenue", "Avg Price"]
                    for col, header in enumerate(ranking_headers, 1):
                        ws.cell(row=row, column=col, value=header)
                    row += 1
                    
                    # Fallback data
                    for i, product in enumerate(top_products, 1):
                        ws.cell(row=row, column=1, value=int(i))
                        ws.cell(row=row, column=2, value=str(product.get("name", "N/A")))
                        ws.cell(row=row, column=3, value=int(product.get("total_sold", 0)))
                        
                        revenue_cell = ws.cell(row=row, column=4, value=float(product.get('total_revenue', 0)))
                        revenue_cell.number_format = currency_format
                        
                        price_cell = ws.cell(row=row, column=5, value=float(product.get('avg_price', 0)))
                        price_cell.number_format = currency_format
                        
                        row += 1
        
        # Final row
        ws.cell(row=row + 1, column=1, value="Report Generation Complete")

    @staticmethod
    def _export_products_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export comprehensive products report to PDF matching the CSV and XLSX formats"""
        from django.utils import timezone
        from datetime import datetime
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Spacer, PageBreak, Paragraph, Table, TableStyle
        
        # Use landscape orientation for better table formatting
        story.append(PageBreak())
        
        # === HEADER INFORMATION ===
        story.append(Paragraph("PRODUCTS REPORT", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "N/A")
        end_str = date_range.get("end", "N/A")
        
        if start_str != "N/A" and end_str != "N/A":
            try:
                start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
                start_str = start_date.strftime("%B %d, %Y")
                end_str = end_date.strftime("%B %d, %Y")
            except:
                pass
        
        story.append(Paragraph(f"<b>Date Range:</b> {start_str} to {end_str}", styles['Normal']))

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        story.append(Paragraph(f"<b>Location:</b> {location_name}", styles['Normal']))

        # Filter information
        filters = report_data.get("filters", {})
        if filters.get("category_id"):
            story.append(Paragraph(f"<b>Category Filter:</b> Category ID {filters['category_id']}", styles['Normal']))
        if filters.get("trend_period"):
            story.append(Paragraph(f"<b>Trend Period:</b> {filters['trend_period']} (Actual: {filters.get('actual_period', 'N/A')})", styles['Normal']))
        
        story.append(Spacer(1, 12))
        
        # === SUMMARY METRICS ===
        story.append(Paragraph("SUMMARY METRICS", styles['Heading1']))
        story.append(Spacer(1, 6))
        
        summary = report_data.get("summary", {})
        
        # Calculate average revenue per product to match CSV
        total_products = summary.get("total_products", 1)
        avg_revenue_per_product = summary.get("total_revenue", 0) / total_products
        
        metrics_data = [
            ["Metric", "Value"],
            ["Total Products Sold (Unique SKUs)", f"{summary.get('total_products', 0):,}"],
            ["Total Units Sold", f"{summary.get('total_units_sold', 0):,}"],
            ["Total Revenue", f"${summary.get('total_revenue', 0):,.2f}"],
            ["Average Revenue per Product", f"${avg_revenue_per_product:,.2f}"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Make metric names bold
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 12))
        
        # === CATEGORY PERFORMANCE ===
        category_performance = report_data.get("category_performance", [])
        if category_performance:
            story.append(Paragraph("CATEGORY PERFORMANCE", styles['Heading1']))
            story.append(Spacer(1, 6))
            
            category_data = [["Category", "Revenue", "Units Sold", "Products", "Avg Rev/Prod", "Avg Units/Prod", "Share %"]]
            
            # Calculate total for percentage
            total_category_revenue = sum(cat.get("revenue", 0) for cat in category_performance)
            
            for category in category_performance:
                revenue = category.get("revenue", 0)
                units_sold = category.get("units_sold", 0)
                unique_products = category.get("unique_products", 1)  # Avoid division by zero
                
                avg_revenue_per_product = revenue / unique_products
                avg_units_per_product = units_sold / unique_products
                category_share = (revenue / total_category_revenue * 100) if total_category_revenue > 0 else 0

                category_data.append([
                    category.get("category", "Uncategorized"),
                    f"${revenue:,.2f}",
                    f"{units_sold:,}",
                    f"{unique_products:,}",
                    f"${avg_revenue_per_product:,.2f}",
                    f"{avg_units_per_product:.1f}",
                    f"{category_share:.1f}%"
                ])
            
            category_table = Table(category_data, colWidths=[1.2*inch, 1*inch, 0.8*inch, 0.6*inch, 0.9*inch, 0.8*inch, 0.7*inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align numeric columns
            ]))
            
            story.append(category_table)
            story.append(Spacer(1, 12))
        
        # === BUSINESS INSIGHTS ===
        story.append(Paragraph("BUSINESS INSIGHTS", styles['Heading1']))
        story.append(Spacer(1, 6))
        
        insights = []
        
        # Top performer
        top_products = report_data.get("top_products", [])
        if top_products:
            top_performer = top_products[0]
            insights.append(f"<b>Top Revenue Product:</b> {top_performer.get('name', 'N/A')} (${top_performer.get('revenue', 0):,.2f})")
            
        # Best seller by quantity (same logic as CSV)
        if top_products:
            best_seller = max(top_products, key=lambda x: x.get('sold', 0))
            insights.append(f"<b>Best Seller by Quantity:</b> {best_seller.get('name', 'N/A')} ({best_seller.get('sold', 0):,} units)")
        
        # Category insights
        if category_performance:
            top_category = max(category_performance, key=lambda x: x.get("revenue", 0))
            insights.append(f"<b>Top Revenue Category:</b> {top_category.get('category', 'N/A')} (${top_category.get('revenue', 0):,.2f})")
        
        for insight in insights:
            story.append(Paragraph(insight, styles['Normal']))
            story.append(Spacer(1, 4))
        
        story.append(Spacer(1, 12))
        
        # === DETAILED PERFORMANCE ANALYSIS ===
        story.append(Paragraph("DETAILED PERFORMANCE ANALYSIS", styles['Heading1']))
        story.append(Spacer(1, 6))
        
        # Top performers breakdown
        if top_products and len(top_products) >= 3:
            story.append(Paragraph("Top 3 Revenue Generators:", styles['Heading2']))
            
            top_3_data = [["Rank", "Product Name", "Revenue", "Units Sold", "Avg Price", "Revenue %"]]
            total_revenue = summary.get('total_revenue', 1)  # Avoid division by zero
            
            for i, product in enumerate(top_products[:3], 1):
                revenue = product.get('revenue', 0)  # Use 'revenue' key like CSV
                revenue_percent = (revenue / total_revenue * 100) if total_revenue > 0 else 0
                product_name = product.get("name", "N/A")
                if len(product_name) > 25:
                    product_name = product_name[:22] + "..."
                    
                top_3_data.append([
                    str(i),
                    product_name,
                    f"${revenue:,.2f}",
                    f"{product.get('sold', 0):,}",  # Use 'sold' key like CSV
                    f"${product.get('avg_price', 0):.2f}",
                    f"{revenue_percent:.1f}%"
                ])
            
            top_3_table = Table(top_3_data, colWidths=[0.4*inch, 1.8*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch])
            top_3_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Center rank
                ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Right align numbers
            ]))
            
            story.append(top_3_table)
            story.append(Spacer(1, 8))
        
        # Category summary if available
        if category_performance and len(category_performance) > 1:
            story.append(Paragraph("Category Distribution:", styles['Heading2']))
            
            # Calculate some category stats
            total_categories = len(category_performance)
            total_category_revenue = sum(cat.get("revenue", 0) for cat in category_performance)
            top_category = max(category_performance, key=lambda x: x.get("revenue", 0))
            
            category_stats = [
                f"• Total Categories: {total_categories}",
                f"• Leading Category: {top_category.get('category', 'N/A')} (${top_category.get('revenue', 0):,.2f})",
                f"• Category Revenue Range: ${min(cat.get('revenue', 0) for cat in category_performance):,.2f} - ${max(cat.get('revenue', 0) for cat in category_performance):,.2f}",
            ]
            
            for stat in category_stats:
                story.append(Paragraph(stat, styles['Normal']))
                story.append(Spacer(1, 2))
            
            story.append(Spacer(1, 8))
        
        
        # === COMPREHENSIVE PRODUCT RANKING ===
        story.append(Paragraph("ALL PRODUCTS SOLD (Ranked by Revenue)", styles['Heading1']))
        story.append(Spacer(1, 6))
        
        # Get ALL products for comprehensive ranking
        filters = report_data.get("filters", {})
        date_range = report_data.get("date_range", {})
        
        if date_range.get("start") and date_range.get("end"):
            try:
                start_date_obj = datetime.fromisoformat(date_range["start"].replace('Z', '+00:00'))
                end_date_obj = datetime.fromisoformat(date_range["end"].replace('Z', '+00:00'))
                category_id = filters.get("category_id") if filters.get("category_id") != "all" else None

                # Extract tenant and location filters
                tenant_id = report_data.get('tenant_id')
                location_id = report_data.get('location_info', {}).get('location_id')

                all_products = ProductsReportService._get_all_products_for_export(
                    start_date_obj, end_date_obj, category_id, tenant_id, location_id
                )
                
                if all_products:
                    # Create table with pagination for long lists
                    ranking_data = [["Rank", "Product Name", "Product ID", "Units Sold", "Revenue", "Avg Price"]]
                    
                    # Show top 100 products for PDF (more detailed than before)
                    displayed_products = all_products[:100]
                    total_products_count = len(all_products)
                    
                    for i, product in enumerate(displayed_products, 1):
                        revenue = product.get("total_revenue", 0)
                        units_sold = product.get("total_sold", 0)
                        avg_price = product.get("avg_price", 0)
                        product_name = product.get("name", "N/A")
                        
                        # Truncate long product names for PDF display
                        if len(product_name) > 30:
                            product_name = product_name[:27] + "..."

                        ranking_data.append([
                            str(i),
                            product_name,
                            str(product.get("id", "N/A")),
                            f"{units_sold:,}",
                            f"${revenue:,.2f}",
                            f"${avg_price:.2f}"
                        ])
                    
                    ranking_table = Table(ranking_data, colWidths=[0.5*inch, 2.2*inch, 0.8*inch, 0.8*inch, 1*inch, 0.7*inch])
                    ranking_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Center rank column
                        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Right align numeric columns
                    ]))
                    
                    story.append(ranking_table)
                    
                    # Add note if there are more products
                    if total_products_count > 100:
                        story.append(Spacer(1, 6))
                        story.append(Paragraph(f"<i>Showing top 100 of {total_products_count} total products. For complete list, use CSV or Excel export.</i>", styles['Normal']))
                    
                    story.append(Spacer(1, 12))
                    
            except Exception as e:
                # Fallback to limited data
                story.append(Paragraph("Error loading comprehensive data. Showing limited results:", styles['Normal']))
                story.append(Spacer(1, 6))
                
                if top_products:
                    fallback_data = [["Rank", "Product Name", "Units Sold", "Revenue", "Avg Price"]]
                    
                    for i, product in enumerate(top_products, 1):
                        product_name = product.get("name", "N/A")
                        if len(product_name) > 35:
                            product_name = product_name[:32] + "..."
                            
                        fallback_data.append([
                            str(i),
                            product_name,
                            f"{product.get('total_sold', 0):,}",
                            f"${product.get('total_revenue', 0):,.2f}",
                            f"${product.get('avg_price', 0):.2f}"
                        ])
                    
                    fallback_table = Table(fallback_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 0.8*inch])
                    fallback_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                    ]))
                    
                    story.append(fallback_table)
        
        # Final note
        story.append(Spacer(1, 12))
        story.append(Paragraph("<i>Report Generation Complete</i>", styles['Normal']))