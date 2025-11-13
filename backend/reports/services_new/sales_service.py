"""
Sales report service with generation and export functionality.
Refactored into modular, scalable components.
"""
import time
import logging
import csv
import io
from calendar import monthrange
from decimal import Decimal
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

from django.db import transaction
from django.db.models import (
    Sum, Count, Avg, F, Q, Value, ExpressionWrapper,
    DecimalField
)
from django.db.models.functions import (
    TruncDate, TruncHour, TruncWeek, TruncMonth, Coalesce, Extract
)
from django.utils import timezone

# Export functionality imports
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors

from orders.models import Order, OrderItem
from payments.models import Payment, PaymentTransaction
from .base import BaseReportService
from .timezone_utils import TimezoneUtils

logger = logging.getLogger(__name__)


class SalesReportService(BaseReportService):
    """Service for generating and exporting sales reports."""

    @staticmethod
    @transaction.atomic
    def generate_sales_report(
        tenant,
        start_date: datetime,
        end_date: datetime,
        location_id: Optional[int] = None,
        group_by: str = "day",
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """Generate detailed sales report"""

        cache_key = SalesReportService._generate_cache_key(
            "sales",
            {"start_date": start_date, "end_date": end_date, "location_id": location_id, "group_by": group_by},
        )

        if use_cache:
            cached_data = SalesReportService._get_cached_report(cache_key, tenant)
            if cached_data:
                return cached_data

        logger.info(f"Generating sales report for {start_date} to {end_date}" + (f" at location {location_id}" if location_id else " (all locations)"))
        start_time = time.time()

        # Multi-location report generation
        if location_id is None:
            return SalesReportService._generate_multi_location_sales_report(
                tenant, start_date, end_date, group_by, use_cache
            )

        # Get base data
        orders_queryset = SalesReportService._get_base_orders_queryset(tenant, start_date, end_date, location_id)

        # Calculate core metrics
        sales_data = SalesReportService._calculate_core_sales_metrics(orders_queryset)

        # Add detailed breakdowns
        sales_data.update(SalesReportService._calculate_sales_by_period(orders_queryset, group_by))
        sales_data.update(SalesReportService._calculate_category_sales(orders_queryset))
        sales_data.update(SalesReportService._calculate_peak_hours(orders_queryset))
        sales_data.update(SalesReportService._calculate_payment_reconciliation(tenant, start_date, end_date, location_id))

        # Add metadata
        sales_data["generated_at"] = timezone.now().isoformat()
        sales_data["date_range"] = {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        }

        # Add location metadata
        location_name = "All Locations"
        if location_id is not None:
            from settings.models import StoreLocation
            try:
                location = StoreLocation.objects.get(id=location_id, tenant=tenant)
                location_name = location.name
            except StoreLocation.DoesNotExist:
                location_name = f"Location ID {location_id}"

        sales_data["location_info"] = {
            "location_id": location_id,
            "location_name": location_name,
            "is_multi_location": location_id is None
        }

        # Add tenant_id for export filtering
        sales_data["tenant_id"] = tenant.id

        # Cache the result
        generation_time = time.time() - start_time
        SalesReportService._cache_report(
            cache_key, sales_data, tenant, report_type="sales", ttl_hours=SalesReportService.CACHE_TTL["sales"]
        )

        logger.info(f"Sales report generated in {generation_time:.2f}s")
        return sales_data

    @staticmethod
    def _generate_multi_location_sales_report(
        tenant,
        start_date: datetime,
        end_date: datetime,
        group_by: str = "day",
        use_cache: bool = True,
    ) -> Dict[str, Any]:
        """
        Generate sales report for all locations with breakdown per location.
        Returns consolidated totals + individual location reports.
        """
        from settings.models import StoreLocation

        logger.info(f"Generating multi-location sales report for {start_date} to {end_date}")
        start_time = time.time()

        # Get all active locations for this tenant
        active_locations = StoreLocation.objects.filter(
            tenant=tenant,
            is_active=True
        ).order_by('name')

        if not active_locations.exists():
            logger.warning(f"No active locations found for tenant {tenant.id}")
            # Return empty multi-location report
            return {
                "is_multi_location": True,
                "location_count": 0,
                "locations": [],
                "consolidated": SalesReportService._get_empty_report_structure(),
                "generated_at": timezone.now().isoformat(),
                "date_range": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat(),
                },
                "tenant_id": tenant.id,
            }

        # Generate report for each location
        location_reports = []
        for location in active_locations:
            logger.info(f"Generating report for location: {location.name} (ID: {location.id})")

            # Generate individual location report (this will use cache if available)
            location_report = SalesReportService.generate_sales_report(
                tenant=tenant,
                start_date=start_date,
                end_date=end_date,
                location_id=location.id,
                group_by=group_by,
                use_cache=use_cache
            )

            location_reports.append({
                "location_id": location.id,
                "location_name": location.name,
                "report_data": location_report
            })

        # Calculate consolidated totals across all locations
        consolidated_totals = SalesReportService._calculate_consolidated_totals(location_reports)

        # Build multi-location response
        multi_location_data = {
            "is_multi_location": True,
            "location_count": len(location_reports),
            "locations": location_reports,
            "consolidated": consolidated_totals,
            "generated_at": timezone.now().isoformat(),
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "tenant_id": tenant.id,
            "location_info": {
                "location_id": None,
                "location_name": "All Locations",
                "is_multi_location": True
            }
        }

        generation_time = time.time() - start_time
        logger.info(f"Multi-location sales report generated in {generation_time:.2f}s for {len(location_reports)} locations")

        return multi_location_data

    @staticmethod
    def _calculate_consolidated_totals(location_reports: list) -> Dict[str, Any]:
        """Calculate consolidated totals from multiple location reports."""
        from collections import defaultdict

        if not location_reports:
            return SalesReportService._get_empty_report_structure()

        consolidated = {
            "total_revenue": Decimal("0.00"),
            "net_revenue": Decimal("0.00"),
            "total_orders": 0,
            "total_items": 0,
            "total_subtotal": Decimal("0.00"),
            "total_tax": Decimal("0.00"),
            "total_tips": Decimal("0.00"),
            "total_surcharges": Decimal("0.00"),
            "total_discounts": Decimal("0.00"),
            "total_refunds": Decimal("0.00"),
        }

        # Initialize aggregation structures for charts
        sales_by_period_agg = defaultdict(lambda: {
            'revenue': Decimal('0.00'),
            'orders': 0,
            'items': 0,
            'transaction_details': {
                'transactions': [],
                'payment_totals': {
                    'total_tips': Decimal('0.00'),
                    'total_surcharges': Decimal('0.00'),
                    'total_collected': Decimal('0.00'),
                },
                'method_breakdown': defaultdict(lambda: {
                    'count': 0,
                    'total_amount': Decimal('0.00'),
                    'total_tips': Decimal('0.00'),
                    'total_surcharges': Decimal('0.00'),
                })
            }
        })

        sales_by_category_agg = defaultdict(lambda: {'revenue': Decimal('0.00'), 'quantity': 0})
        top_hours_agg = defaultdict(lambda: {'revenue': Decimal('0.00'), 'orders': 0})

        # Aggregate metrics across all locations
        for loc_data in location_reports:
            report = loc_data["report_data"]

            consolidated["total_revenue"] += Decimal(str(report.get("total_revenue", 0)))
            consolidated["net_revenue"] += Decimal(str(report.get("net_revenue", 0)))
            consolidated["total_orders"] += report.get("total_orders", 0)
            consolidated["total_items"] += report.get("total_items", 0)
            consolidated["total_subtotal"] += Decimal(str(report.get("total_subtotal", 0)))
            consolidated["total_tax"] += Decimal(str(report.get("total_tax", 0)))
            consolidated["total_tips"] += Decimal(str(report.get("total_tips", 0)))
            consolidated["total_surcharges"] += Decimal(str(report.get("total_surcharges", 0)))
            consolidated["total_discounts"] += Decimal(str(report.get("total_discounts", 0)))
            consolidated["total_refunds"] += Decimal(str(report.get("total_refunds", 0)))

            # Aggregate sales_by_period
            for period in report.get("sales_by_period", []):
                date = period['date']
                sales_by_period_agg[date]['revenue'] += Decimal(str(period.get('revenue', 0)))
                sales_by_period_agg[date]['orders'] += period.get('orders', 0)
                sales_by_period_agg[date]['items'] += period.get('items', 0)

                # Aggregate transaction details
                txn_details = period.get('transaction_details', {})
                if txn_details:
                    # Add transactions
                    sales_by_period_agg[date]['transaction_details']['transactions'].extend(
                        txn_details.get('transactions', [])
                    )

                    # Aggregate payment totals
                    payment_totals = txn_details.get('payment_totals', {})
                    sales_by_period_agg[date]['transaction_details']['payment_totals']['total_tips'] += \
                        Decimal(str(payment_totals.get('total_tips', 0)))
                    sales_by_period_agg[date]['transaction_details']['payment_totals']['total_surcharges'] += \
                        Decimal(str(payment_totals.get('total_surcharges', 0)))
                    sales_by_period_agg[date]['transaction_details']['payment_totals']['total_collected'] += \
                        Decimal(str(payment_totals.get('total_collected', 0)))

                    # Aggregate method breakdown
                    for method, breakdown in txn_details.get('method_breakdown', {}).items():
                        sales_by_period_agg[date]['transaction_details']['method_breakdown'][method]['count'] += \
                            breakdown.get('count', 0)
                        sales_by_period_agg[date]['transaction_details']['method_breakdown'][method]['total_amount'] += \
                            Decimal(str(breakdown.get('total_amount', 0)))
                        sales_by_period_agg[date]['transaction_details']['method_breakdown'][method]['total_tips'] += \
                            Decimal(str(breakdown.get('total_tips', 0)))
                        sales_by_period_agg[date]['transaction_details']['method_breakdown'][method]['total_surcharges'] += \
                            Decimal(str(breakdown.get('total_surcharges', 0)))

            # Aggregate sales_by_category
            for category in report.get("sales_by_category", []):
                cat_name = category['category']
                sales_by_category_agg[cat_name]['revenue'] += Decimal(str(category.get('revenue', 0)))
                sales_by_category_agg[cat_name]['quantity'] += category.get('quantity', 0)

            # Aggregate top_hours
            for hour in report.get("top_hours", []):
                hour_key = hour['hour']
                top_hours_agg[hour_key]['revenue'] += Decimal(str(hour.get('revenue', 0)))
                top_hours_agg[hour_key]['orders'] += hour.get('orders', 0)

        # Calculate average order value
        if consolidated["total_orders"] > 0:
            consolidated["avg_order_value"] = float(consolidated["total_revenue"]) / consolidated["total_orders"]
        else:
            consolidated["avg_order_value"] = 0.0

        # Build sales_by_period array
        sales_by_period = []
        for date in sorted(sales_by_period_agg.keys()):
            data = sales_by_period_agg[date]

            # Convert method_breakdown from defaultdict to regular dict with floats
            method_breakdown = {}
            for method, breakdown in data['transaction_details']['method_breakdown'].items():
                method_breakdown[method] = {
                    'count': breakdown['count'],
                    'total_amount': float(breakdown['total_amount']),
                    'total_tips': float(breakdown['total_tips']),
                    'total_surcharges': float(breakdown['total_surcharges']),
                }

            sales_by_period.append({
                'date': date,
                'revenue': float(data['revenue']),
                'orders': data['orders'],
                'items': data['items'],
                'transaction_details': {
                    'transactions': data['transaction_details']['transactions'],
                    'payment_totals': {
                        'total_tips': float(data['transaction_details']['payment_totals']['total_tips']),
                        'total_surcharges': float(data['transaction_details']['payment_totals']['total_surcharges']),
                        'total_collected': float(data['transaction_details']['payment_totals']['total_collected']),
                    },
                    'method_breakdown': method_breakdown
                }
            })

        # Build sales_by_category array (sorted by revenue)
        sales_by_category = [
            {
                'category': category,
                'revenue': float(data['revenue']),
                'quantity': data['quantity']
            }
            for category, data in sorted(
                sales_by_category_agg.items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )
        ]

        # Build top_hours array (sorted by revenue, top 10)
        top_hours = [
            {
                'hour': hour,
                'revenue': float(data['revenue']),
                'orders': data['orders']
            }
            for hour, data in sorted(
                top_hours_agg.items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )[:10]
        ]

        # Convert Decimals to floats for JSON serialization
        for key in ["total_revenue", "net_revenue", "total_subtotal", "total_tax",
                    "total_tips", "total_surcharges", "total_discounts", "total_refunds"]:
            consolidated[key] = float(consolidated[key])

        # Add the aggregated arrays
        consolidated["sales_by_period"] = sales_by_period
        consolidated["sales_by_category"] = sales_by_category
        consolidated["top_hours"] = top_hours

        # Add revenue_breakdown structure
        consolidated["revenue_breakdown"] = {
            "total_revenue": consolidated["total_revenue"],
            "components": {
                "subtotal": consolidated["total_subtotal"],
                "tips": consolidated["total_tips"],
                "surcharges": consolidated["total_surcharges"],
                "tax": consolidated["total_tax"],
                "discounts_applied": consolidated["total_discounts"],
                "refunds": consolidated["total_refunds"],
                "net_revenue": consolidated["net_revenue"],
            }
        }

        return consolidated

    @staticmethod
    def _get_empty_report_structure() -> Dict[str, Any]:
        """Return an empty report structure with zero values."""
        return {
            "total_revenue": 0.0,
            "net_revenue": 0.0,
            "total_orders": 0,
            "total_items": 0,
            "avg_order_value": 0.0,
            "total_subtotal": 0.0,
            "total_tax": 0.0,
            "total_tips": 0.0,
            "total_surcharges": 0.0,
            "total_discounts": 0.0,
            "total_refunds": 0.0,
        }

    @staticmethod
    def _get_base_orders_queryset(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None):
        """Get optimized base queryset for completed orders in date range."""
        filters = {
            "tenant": tenant,
            "status": Order.OrderStatus.COMPLETED,
            "completed_at__range": (start_date, end_date),  # Use completed_at for accurate revenue attribution
            "subtotal__gt": 0,  # Exclude orders with $0.00 subtotals
        }

        # Add location filter if specified
        if location_id is not None:
            filters["store_location_id"] = location_id

        return (
            Order.objects.filter(**filters)
            .select_related("cashier", "customer", "payment_details", "store_location")
            .prefetch_related("items__product")
        )
    
    @staticmethod
    def _calculate_core_sales_metrics(orders_queryset) -> Dict[str, Any]:
        """Calculate core sales metrics from orders queryset."""

        # Calculate total refunds separately
        total_refunds = (
            PaymentTransaction.objects.select_related("payment", "payment__order")
            .filter(
                payment__order__in=orders_queryset,
                status=PaymentTransaction.TransactionStatus.REFUNDED,
            )
            .aggregate(
                total_refunds=Coalesce(Sum("refunded_amount"), Value(Decimal("0.00")))
            )["total_refunds"]
        )

        # Core order metrics (WITHOUT items to avoid JOIN duplication)
        order_data = orders_queryset.aggregate(
            total_subtotal=Coalesce(Sum("subtotal"), Value(Decimal("0.00"))),
            total_orders=Count("id"),
            total_tax=Coalesce(Sum("tax_total"), Value(Decimal("0.00"))),
            total_discounts=Coalesce(
                Sum("total_discounts_amount"), Value(Decimal("0.00"))
            ),
        )

        # Calculate payment totals separately - this is our PRIMARY revenue metric
        payment_totals = Payment.objects.filter(order__in=orders_queryset).aggregate(
            total_revenue=Coalesce(Sum("total_collected"), Value(Decimal("0.00"))),  # PRIMARY metric
            total_surcharges=Coalesce(Sum("total_surcharges"), Value(Decimal("0.00"))),
            total_tips=Coalesce(Sum("total_tips"), Value(Decimal("0.00"))),
            avg_order_value=Coalesce(Avg("total_collected"), Value(Decimal("0.00"))),  # Based on collected
        )

        # Merge order and payment data
        sales_data = {**order_data, **payment_totals}

        # Convert Decimal to float for JSON serialization

        # Convert Decimal to float
        sales_data = {
            k: float(v) if isinstance(v, Decimal) else v for k, v in sales_data.items()
        }

        # Add refunds to the sales data
        sales_data["total_refunds"] = float(total_refunds)
        
        # Calculate net revenue properly: what the business actually keeps
        # Net revenue = subtotal + tips - discounts - refunds (excludes tax/surcharges)
        sales_data["net_revenue"] = (
            sales_data["total_subtotal"] + sales_data["total_tips"] - sales_data["total_discounts"] - sales_data["total_refunds"]
        )
        
        # Add detailed revenue breakdown for frontend
        sales_data["revenue_breakdown"] = {
            "total_revenue": sales_data["total_revenue"],  # What was actually collected
            "components": {
                "subtotal": sales_data["total_subtotal"],
                "tips": sales_data["total_tips"],
                "surcharges": sales_data["total_surcharges"],
                "tax": sales_data["total_tax"],
                "discounts_applied": -sales_data["total_discounts"],  # Negative for display
                "refunds": -sales_data["total_refunds"],  # Negative for display
                "net_revenue": sales_data["net_revenue"]
            }
        }

        # Calculate total_items separately to avoid ORDER duplication from JOINs
        sales_data["total_items"] = orders_queryset.aggregate(
            total_items=Coalesce(Sum("items__quantity"), Value(0))
        )["total_items"]
        
        return sales_data
        
    @staticmethod
    def _calculate_sales_by_period(orders_queryset, group_by: str) -> Dict[str, Any]:
        """Calculate sales breakdown by time period."""
        
        # Determine truncation function - use completed_at for accurate revenue attribution
        if group_by == "week":
            trunc_period = TruncWeek("completed_at")
        elif group_by == "month":
            trunc_period = TruncMonth("completed_at")
        else:
            trunc_period = TruncDate("completed_at")

        # Step 1: Aggregate revenue and orders. This is correct as it doesn't involve joins that cause duplication.
        sales_agg = (
            orders_queryset.annotate(period=trunc_period)
            .values("period")
            .annotate(
                revenue=Sum("grand_total"),
                orders=Count("id"),
            )
            .order_by("period")
        )

        # Step 2: Aggregate item quantities separately to avoid inflating the sales/order counts.
        # This query introduces a JOIN on OrderItem, so it's done independently.
        items_agg = (
            orders_queryset.annotate(period=trunc_period)
            .values("period")
            .annotate(items=Sum("items__quantity"))
            .order_by("period")
        )

        # Step 3: Merge the two aggregations in memory for the final result.
        items_dict = {item["period"]: item["items"] for item in items_agg}

        # Get detailed transaction data for each period
        transaction_details_by_period = SalesReportService._get_transaction_details_by_period(
            orders_queryset, sales_agg, group_by
        )

        sales_by_period = [
            {
                "date": item["period"].strftime("%Y-%m-%d"),
                "revenue": float(item["revenue"] or 0),
                "orders": item["orders"],
                "items": items_dict.get(item["period"], 0) or 0,
                "transaction_details": transaction_details_by_period.get(
                    item["period"],
                    {
                        "transactions": [],
                        "payment_totals": {
                            "total_tips": 0,
                            "total_surcharges": 0,
                            "total_collected": 0,
                        },
                        "method_breakdown": {},
                    },
                ),
            }
            for item in sales_agg
        ]
        
        return {"sales_by_period": sales_by_period}
    
    @staticmethod
    def _get_transaction_details_by_period(orders_queryset, sales_agg, group_by: str) -> Dict[str, Any]:
        """Get detailed transaction information for each period."""
        transaction_details_by_period = {}

        for period_item in sales_agg:
            period_start = period_item["period"]

            # Create timezone-aware datetime objects for the period range
            local_tz = TimezoneUtils.get_local_timezone()
            start_dt = timezone.make_aware(
                datetime.combine(period_start, datetime.min.time()), local_tz
            )

            if group_by == "week":
                end_dt = start_dt + timedelta(days=7)
            elif group_by == "month":
                days_in_month = monthrange(start_dt.year, start_dt.month)[1]
                end_dt = start_dt + timedelta(days=days_in_month)
            else:  # day
                end_dt = start_dt + timedelta(days=1)

            # Get transactions for this period using the precise datetime range
            # Note: Filter by order's completed_at for accurate period attribution
            period_transactions = (
                PaymentTransaction.objects.select_related("payment", "payment__order")
                .filter(
                    payment__order__in=orders_queryset,
                    payment__order__completed_at__gte=start_dt,
                    payment__order__completed_at__lt=end_dt,
                    status=PaymentTransaction.TransactionStatus.SUCCESSFUL,
                )
                .order_by("-payment__order__completed_at")
            )

            # Get payment totals for this period
            period_payments = Payment.objects.filter(
                order__in=orders_queryset,
                order__completed_at__gte=start_dt,
                order__completed_at__lt=end_dt,
            ).aggregate(
                total_tips=Coalesce(Sum("total_tips"), Value(Decimal("0.00"))),
                total_surcharges=Coalesce(
                    Sum("total_surcharges"), Value(Decimal("0.00"))
                ),
                total_collected=Coalesce(
                    Sum("total_collected"), Value(Decimal("0.00"))
                ),
            )

            # Group transactions by payment method
            method_breakdown = SalesReportService._group_transactions_by_method(period_transactions)

            transaction_details_by_period[period_item["period"]] = {
                "transactions": [
                    {
                        "order_number": trans.payment.order.order_number,
                        "completed_at": (trans.payment.order.completed_at or trans.payment.order.updated_at).isoformat(),
                        "amount": float(trans.amount or 0),
                        "tip": float(trans.tip or 0),
                        "surcharge": float(trans.surcharge or 0),
                        "method": trans.method,
                        "transaction_id": trans.transaction_id,
                        "card_brand": trans.card_brand,
                        "card_last4": trans.card_last4,
                    }
                    for trans in period_transactions
                ],
                "payment_totals": {
                    "total_tips": float(period_payments["total_tips"]),
                    "total_surcharges": float(period_payments["total_surcharges"]),
                    "total_collected": float(period_payments["total_collected"]),
                },
                "method_breakdown": method_breakdown,
            }
        
        return transaction_details_by_period
    
    @staticmethod
    def _group_transactions_by_method(period_transactions) -> Dict[str, Any]:
        """Group transactions by payment method for breakdown analysis."""
        method_breakdown = {}
        for transaction in period_transactions:
            method = transaction.method
            if method not in method_breakdown:
                method_breakdown[method] = {
                    "count": 0,
                    "total_amount": 0,
                    "total_tips": 0,
                    "total_surcharges": 0,
                }
            method_breakdown[method]["count"] += 1
            method_breakdown[method]["total_amount"] += float(
                transaction.amount or 0
            )
            method_breakdown[method]["total_tips"] += float(transaction.tip or 0)
            method_breakdown[method]["total_surcharges"] += float(
                transaction.surcharge or 0
            )
        return method_breakdown
    
    @staticmethod
    def _calculate_category_sales(orders_queryset) -> Dict[str, Any]:
        """Calculate sales by product category."""
        category_sales = (
            OrderItem.objects.filter(order__in=orders_queryset)
            .values("product__category__name")
            .annotate(
                revenue=Sum(F("quantity") * F("price_at_sale")),
                quantity=Sum("quantity"),
            )
            .order_by("-revenue")
        )

        sales_by_category = [
            {
                "category": item["product__category__name"] or "Uncategorized",
                "revenue": float(item["revenue"] or 0),
                "quantity": item["quantity"] or 0,
            }
            for item in category_sales
        ]
        
        return {"sales_by_category": sales_by_category}
    
    @staticmethod
    def _calculate_peak_hours(orders_queryset) -> Dict[str, Any]:
        """Calculate top performing hours by revenue."""
        hourly_sales = (
            orders_queryset.annotate(hour=Extract("completed_at", "hour"))
            .values("hour")
            .annotate(revenue=Sum("grand_total"), orders=Count("id"))
            .order_by("-revenue")[:10]
        )

        top_hours = [
            {
                "hour": f"{item['hour']:02d}:00",
                "revenue": float(item["revenue"] or 0),
                "orders": item["orders"],
            }
            for item in hourly_sales
        ]
        
        return {"top_hours": top_hours}
    
    @staticmethod
    def _calculate_payment_reconciliation(tenant, start_date: datetime, end_date: datetime, location_id: Optional[int] = None) -> Dict[str, Any]:
        """Calculate payment success metrics focusing on transaction success rates."""

        # Get all orders in the date range
        filters = {
            "tenant": tenant,
            "completed_at__range": (start_date, end_date),
            "subtotal__gt": 0  # Exclude $0 orders
        }

        if location_id is not None:
            filters["store_location_id"] = location_id

        all_orders = Order.objects.filter(**filters)

        # Get payments for these orders
        all_payments = Payment.objects.filter(
            order__in=all_orders
        )
        
        # Calculate order breakdown by status
        order_breakdown = all_orders.aggregate(
            completed_count=Count("id", filter=Q(status=Order.OrderStatus.COMPLETED)),
            voided_count=Count("id", filter=Q(status=Order.OrderStatus.VOID)),
            canceled_count=Count("id", filter=Q(status=Order.OrderStatus.CANCELLED)),
            total_orders=Count("id")
        )
        
        # Calculate payment success rates
        payment_breakdown = all_payments.aggregate(
            paid_count=Count("id", filter=Q(status=Payment.PaymentStatus.PAID)),
            unpaid_count=Count("id", filter=Q(status=Payment.PaymentStatus.UNPAID)),
            partially_paid_count=Count("id", filter=Q(status=Payment.PaymentStatus.PARTIALLY_PAID)),
            total_payments=Count("id")
        )
        
        # Payment success rate = successful payments / total payment attempts
        payment_success_rate = round(
            (payment_breakdown["paid_count"] / payment_breakdown["total_payments"] * 100), 2
        ) if payment_breakdown["total_payments"] > 0 else 100.0
        
        # Order completion rate = completed orders / total orders
        order_completion_rate = round(
            (order_breakdown["completed_count"] / order_breakdown["total_orders"] * 100), 2
        ) if order_breakdown["total_orders"] > 0 else 100.0
        
        # Get transaction-level breakdown for detailed analysis
        transaction_breakdown = PaymentTransaction.objects.filter(
            payment__order__in=all_orders
        ).aggregate(
            successful_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.SUCCESSFUL)),
            refunded_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.REFUNDED)),
            failed_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.FAILED)),
            canceled_count=Count("id", filter=Q(status=PaymentTransaction.TransactionStatus.CANCELED)),
            total_transactions=Count("id")
        )
        
        # Transaction success rate
        transaction_success_rate = round(
            (transaction_breakdown["successful_count"] / transaction_breakdown["total_transactions"] * 100), 2
        ) if transaction_breakdown["total_transactions"] > 0 else 100.0
        
        return {
            "payment_performance": {
                "order_completion_rate": order_completion_rate,
                "payment_success_rate": payment_success_rate,
                "transaction_success_rate": transaction_success_rate,
                "order_breakdown": {
                    "completed": order_breakdown["completed_count"],
                    "voided": order_breakdown["voided_count"],
                    "canceled": order_breakdown["canceled_count"],
                    "total": order_breakdown["total_orders"]
                },
                "payment_breakdown": {
                    "paid": payment_breakdown["paid_count"],
                    "unpaid": payment_breakdown["unpaid_count"],
                    "partially_paid": payment_breakdown["partially_paid_count"],
                    "total": payment_breakdown["total_payments"]
                },
                "transaction_breakdown": {
                    "successful": transaction_breakdown["successful_count"],
                    "failed": transaction_breakdown["failed_count"],
                    "refunded": transaction_breakdown["refunded_count"],
                    "canceled": transaction_breakdown["canceled_count"],
                    "total": transaction_breakdown["total_transactions"]
                }
            }
        }

    @staticmethod
    def export_sales_to_csv(report_data: Dict[str, Any]) -> bytes:
        """Export sales report to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            SalesReportService._export_multi_location_sales_to_csv(writer, report_data)
        else:
            SalesReportService._export_sales_to_csv(writer, report_data)

        csv_bytes = output.getvalue().encode('utf-8')
        output.close()
        return csv_bytes

    @staticmethod
    def _export_sales_to_csv(writer, report_data: Dict[str, Any]):
        """Export comprehensive sales report to CSV matching the old format"""
        from django.utils import timezone
        from datetime import datetime
        
        # Header
        writer.writerow(["Sales Report"])
        writer.writerow(["Generated:", report_data.get("generated_at", timezone.now().isoformat())])
        
        # Extract date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")
        
        # Format dates for display (remove time and timezone for cleaner look)
        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            writer.writerow([f"Date Range: {start_date} to {end_date}"])
        except:
            writer.writerow([f"Date Range: {start_str} to {end_str}"])

        # Add location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        writer.writerow([f"Location: {location_name}"])

        writer.writerow([])
        
        # === FINANCIAL SUMMARY ===
        writer.writerow(["=== FINANCIAL SUMMARY ==="])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Revenue (Collected)", f"${report_data.get('total_revenue', 0):,.2f}"])
        writer.writerow(["Net Revenue", f"${report_data.get('net_revenue', 0):,.2f}"])
        writer.writerow(["Total Orders", f"{report_data.get('total_orders', 0):,}"])
        writer.writerow(["Average Order Value", f"${report_data.get('avg_order_value', 0):,.2f}"])
        writer.writerow(["Total Items Sold", f"{report_data.get('total_items', 0):,}"])
        writer.writerow([])
        
        # === REVENUE BREAKDOWN ===
        writer.writerow(["=== REVENUE BREAKDOWN ==="])
        writer.writerow(["Component", "Amount"])
        writer.writerow(["Subtotal", f"${report_data.get('total_subtotal', 0):,.2f}"])
        writer.writerow(["Tax Collected", f"${report_data.get('total_tax', 0):,.2f}"])
        writer.writerow(["Tips", f"${report_data.get('total_tips', 0):,.2f}"])
        writer.writerow(["Surcharges", f"${report_data.get('total_surcharges', 0):,.2f}"])
        writer.writerow(["Discounts Applied", f"-${report_data.get('total_discounts', 0):,.2f}"])
        writer.writerow([])
        
        # === ORDER DETAILS ===
        writer.writerow(["=== ORDER DETAILS ==="])
        headers = [
            "Order Number", "Date", "Time", "Order Type", "Status", "Payment Status",
            "Payment Method", "Customer Name", "Customer Email", "Customer Phone", 
            "Cashier", "Subtotal", "Tax", "Tips", "Surcharges", 
            "Total Collected", "Discounts", "Items Count", "Total Quantity"
        ]
        writer.writerow(headers)
        
        # Get individual order details - need to query from the date range
        try:
            start_date_obj = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date_obj = datetime.fromisoformat(end_str.replace('Z', '+00:00'))

            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            # Query orders for detailed export
            from orders.models import Order
            filters = {
                'status': Order.OrderStatus.COMPLETED,
                'completed_at__range': (start_date_obj, end_date_obj),
                'subtotal__gt': 0
            }

            # Add tenant filter
            if tenant_id:
                filters['tenant_id'] = tenant_id

            # Add location filter if specified
            if location_id is not None:
                filters['store_location_id'] = location_id

            orders = Order.objects.filter(**filters).select_related(
                'cashier', 'customer', 'payment_details'
            ).prefetch_related(
                'items__product'
            ).order_by('order_number')
            
            for order in orders:
                # Calculate order metrics
                items_count = order.items.count()
                total_quantity = sum(item.quantity for item in order.items.all())
                
                # Get payment method
                payment_method = "N/A"
                if hasattr(order, 'payment_details') and order.payment_details:
                    transactions = order.payment_details.transactions.filter(
                        status="SUCCESSFUL"
                    ).first()
                    if transactions:
                        method_map = {
                            "CARD_TERMINAL": "Card (Terminal)",
                            "CASH": "Cash",
                            "GIFT_CARD": "Gift Card"
                        }
                        payment_method = method_map.get(transactions.method, transactions.method)
                
                # Customer info using the Order model's built-in method
                customer_name = order.customer_display_name
                customer_email = order.customer_email or ""
                customer_phone = order.customer_phone or ""
                
                # Cashier
                cashier_name = ""
                if order.cashier:
                    cashier_name = f"{order.cashier.first_name} {order.cashier.last_name}".strip()
                
                # Payment details - use safe access
                total_tips = 0
                total_surcharges = 0
                total_collected = 0
                if hasattr(order, 'payment_details') and order.payment_details:
                    total_tips = order.payment_details.total_tips or 0
                    total_surcharges = order.payment_details.total_surcharges or 0
                    total_collected = order.payment_details.total_collected or 0
                
                # Format the order row
                order_row = [
                    order.order_number,
                    order.completed_at.strftime("%Y-%m-%d") if order.completed_at else "N/A",
                    order.completed_at.strftime("%H:%M:%S") if order.completed_at else "N/A", 
                    order.order_type,
                    order.status,
                    "PAID" if order.payment_status == "PAID" else order.payment_status,
                    payment_method,
                    customer_name,
                    customer_email,
                    customer_phone,
                    cashier_name,
                    f"${order.subtotal:.2f}",
                    f"${order.tax_total:.2f}",
                    f"${total_tips:.2f}",
                    f"${total_surcharges:.2f}",
                    f"${total_collected:.2f}",
                    f"${order.total_discounts_amount:.2f}",
                    items_count,
                    total_quantity
                ]
                writer.writerow(order_row)
                
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to export detailed order data: {e}")
            writer.writerow(["Error: Could not load detailed order data"])

    @staticmethod
    def _export_multi_location_sales_to_csv(writer, report_data: Dict[str, Any]):
        """Export multi-location sales report to CSV with consolidated and per-location sections."""
        from django.utils import timezone
        from datetime import datetime

        # Header
        writer.writerow(["Sales Report - All Locations"])
        writer.writerow(["Generated:", report_data.get("generated_at", timezone.now().isoformat())])

        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            writer.writerow([f"Date Range: {start_date} to {end_date}"])
        except:
            writer.writerow([f"Date Range: {start_str} to {end_str}"])

        # Location count
        location_count = report_data.get('location_count', 0)
        writer.writerow([f"Locations Included: {location_count}"])
        writer.writerow([])

        # === CONSOLIDATED FINANCIAL SUMMARY ===
        consolidated = report_data.get('consolidated', {})

        writer.writerow(["=== CONSOLIDATED FINANCIAL SUMMARY (ALL LOCATIONS) ==="])
        writer.writerow(["Metric", "Total Value"])
        writer.writerow(["Total Revenue (All Locations)", f"${consolidated.get('total_revenue', 0):,.2f}"])
        writer.writerow(["Net Revenue (All Locations)", f"${consolidated.get('net_revenue', 0):,.2f}"])
        writer.writerow(["Total Orders (All Locations)", f"{consolidated.get('total_orders', 0):,}"])
        writer.writerow(["Average Order Value (All Locations)", f"${consolidated.get('avg_order_value', 0):,.2f}"])
        writer.writerow(["Total Items Sold (All Locations)", f"{consolidated.get('total_items', 0):,}"])
        writer.writerow([])

        # === CONSOLIDATED REVENUE BREAKDOWN ===
        writer.writerow(["=== CONSOLIDATED REVENUE BREAKDOWN ==="])
        writer.writerow(["Component", "Amount"])
        writer.writerow(["Total Subtotal", f"${consolidated.get('total_subtotal', 0):,.2f}"])
        writer.writerow(["Total Tax Collected", f"${consolidated.get('total_tax', 0):,.2f}"])
        writer.writerow(["Total Tips", f"${consolidated.get('total_tips', 0):,.2f}"])
        writer.writerow(["Total Surcharges", f"${consolidated.get('total_surcharges', 0):,.2f}"])
        writer.writerow(["Total Discounts Applied", f"-${abs(consolidated.get('total_discounts', 0)):,.2f}"])
        writer.writerow(["Total Refunds", f"${consolidated.get('total_refunds', 0):,.2f}"])
        writer.writerow([])

        # === LOCATION COMPARISON TABLE ===
        writer.writerow(["=== LOCATION COMPARISON ==="])
        writer.writerow(["Location", "Revenue", "Orders", "Avg Order", "Items", "Subtotal", "Tax", "Tips"])

        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            writer.writerow([
                location_name,
                f"${loc_report.get('total_revenue', 0):,.2f}",
                f"{loc_report.get('total_orders', 0):,}",
                f"${loc_report.get('avg_order_value', 0):,.2f}",
                f"{loc_report.get('total_items', 0):,}",
                f"${loc_report.get('total_subtotal', 0):,.2f}",
                f"${loc_report.get('total_tax', 0):,.2f}",
                f"${loc_report.get('total_tips', 0):,.2f}",
            ])

        writer.writerow([])
        writer.writerow([])

        # === INDIVIDUAL LOCATION DETAILS ===
        writer.writerow(["=" * 100])
        writer.writerow(["=== DETAILED BREAKDOWN BY LOCATION ==="])
        writer.writerow(["=" * 100])
        writer.writerow([])

        for i, location_data in enumerate(locations, 1):
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Location separator
            writer.writerow(["=" * 100])
            writer.writerow([f"LOCATION {i}: {location_name}"])
            writer.writerow(["=" * 100])
            writer.writerow([])

            # Export this location's data using the existing single-location logic
            # We'll write the same sections as the single-location export but without the header
            SalesReportService._export_single_location_section_to_csv(writer, loc_report, include_header=False)

            writer.writerow([])
            writer.writerow([])

    @staticmethod
    def _export_single_location_section_to_csv(writer, report_data: Dict[str, Any], include_header=True):
        """Helper method to export a single location's data section for multi-location CSV reports."""
        from django.utils import timezone
        from datetime import datetime

        # If including header, add it
        if include_header:
            writer.writerow(["Sales Report"])
            writer.writerow(["Generated:", report_data.get("generated_at", timezone.now().isoformat())])

            date_range = report_data.get("date_range", {})
            start_str = date_range.get("start", "")
            end_str = date_range.get("end", "")

            try:
                start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
                end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
                writer.writerow([f"Date Range: {start_date} to {end_date}"])
            except:
                writer.writerow([f"Date Range: {start_str} to {end_str}"])

            location_info = report_data.get("location_info", {})
            location_name = location_info.get("location_name", "All Locations")
            writer.writerow([f"Location: {location_name}"])
            writer.writerow([])

        # === FINANCIAL SUMMARY ===
        writer.writerow(["FINANCIAL SUMMARY"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Revenue (Collected)", f"${report_data.get('total_revenue', 0):,.2f}"])
        writer.writerow(["Net Revenue", f"${report_data.get('net_revenue', 0):,.2f}"])
        writer.writerow(["Total Orders", f"{report_data.get('total_orders', 0):,}"])
        writer.writerow(["Average Order Value", f"${report_data.get('avg_order_value', 0):,.2f}"])
        writer.writerow(["Total Items Sold", f"{report_data.get('total_items', 0):,}"])
        writer.writerow([])

        # === REVENUE BREAKDOWN ===
        writer.writerow(["REVENUE BREAKDOWN"])
        writer.writerow(["Component", "Amount"])
        writer.writerow(["Subtotal", f"${report_data.get('total_subtotal', 0):,.2f}"])
        writer.writerow(["Tax Collected", f"${report_data.get('total_tax', 0):,.2f}"])
        writer.writerow(["Tips", f"${report_data.get('total_tips', 0):,.2f}"])
        writer.writerow(["Surcharges", f"${report_data.get('total_surcharges', 0):,.2f}"])
        writer.writerow(["Discounts Applied", f"-${abs(report_data.get('total_discounts', 0)):,.2f}"])
        writer.writerow([])

        # === SALES BY CATEGORY ===
        if 'sales_by_category' in report_data and report_data['sales_by_category']:
            writer.writerow(["SALES BY CATEGORY (Top 10)"])
            writer.writerow(["Category", "Revenue", "Quantity"])
            categories_to_write = report_data['sales_by_category'][:10]
            for cat in categories_to_write:
                writer.writerow([
                    cat['category'],
                    f"${cat['revenue']:,.2f}",
                    f"{cat['quantity']:,}"
                ])
            writer.writerow([])

    @staticmethod
    def export_sales_to_xlsx(report_data: Dict[str, Any], ws, header_font, header_fill, header_alignment):
        """Export sales report to Excel format."""
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            # For multi-location reports, ws is actually the workbook
            SalesReportService._export_multi_location_sales_to_xlsx(ws, report_data, header_font, header_fill, header_alignment)
        else:
            # Single location report - ws is a worksheet
            SalesReportService._export_sales_to_xlsx(ws, report_data, header_font, header_fill, header_alignment)

    @staticmethod
    def _export_sales_to_xlsx(
        ws, report_data: Dict[str, Any], header_font, header_fill, header_alignment
    ):
        """Export comprehensive sales report to Excel matching the CSV format"""
        from django.utils import timezone
        from datetime import datetime
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Define additional styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        section_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_header_font = Font(bold=True, color="FFFFFF", size=12)
        
        subsection_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subsection_font = Font(bold=True, size=11)
        
        currency_format = '"$"#,##0.00'
        number_format = '#,##0'
        
        row = 1
        
        # Title
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "Sales Report"
        ws[f"A{row}"].font = Font(bold=True, size=16)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Generated date
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"Generated: {generated_at}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1
        
        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")
        
        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"Date Range: {start_date} to {end_date}"
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
        except:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"Date Range: {start_str} to {end_str}"
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"Location: {location_name}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].font = Font(bold=True)
        row += 2

        # === FINANCIAL SUMMARY ===
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "FINANCIAL SUMMARY"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Headers for financial summary
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].font = subsection_font
        ws[f"B{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        ws[f"B{row}"].fill = subsection_fill
        row += 1
        
        # Financial metrics
        financial_metrics = [
            ("Total Revenue (Collected)", report_data.get('total_revenue', 0), currency_format),
            ("Net Revenue", report_data.get('net_revenue', 0), currency_format),
            ("Total Orders", report_data.get("total_orders", 0), number_format),
            ("Average Order Value", report_data.get('avg_order_value', 0), currency_format),
            ("Total Items Sold", report_data.get('total_items', 0), number_format),
        ]
        
        for metric, value, fmt in financial_metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = fmt
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1
        
        row += 1  # Empty row
        
        # === REVENUE BREAKDOWN ===
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "REVENUE BREAKDOWN"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Headers for revenue breakdown
        ws[f"A{row}"] = "Component"
        ws[f"B{row}"] = "Amount"
        ws[f"A{row}"].font = subsection_font
        ws[f"B{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        ws[f"B{row}"].fill = subsection_fill
        row += 1
        
        # Revenue components
        revenue_components = [
            ("Subtotal", report_data.get('total_subtotal', 0)),
            ("Tax Collected", report_data.get('total_tax', 0)),
            ("Tips", report_data.get('total_tips', 0)),
            ("Surcharges", report_data.get('total_surcharges', 0)),
            ("Discounts Applied", -abs(report_data.get('total_discounts', 0))),  # Show as negative
        ]
        
        for component, value in revenue_components:
            ws[f"A{row}"] = component
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = currency_format
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            # Color code negative values (discounts)
            if value < 0:
                ws[f"B{row}"].font = Font(color="FF0000")  # Red for negative
            row += 1
        
        row += 1  # Empty row
        
        # === SALES BY CATEGORY ===
        cat_row = row  # Initialize cat_row with current row value
        if 'sales_by_category' in report_data and report_data['sales_by_category']:
            ws.merge_cells(f"D{5}:F{5}")
            ws[f"D{5}"] = "SALES BY CATEGORY"
            ws[f"D{5}"].font = section_header_font
            ws[f"D{5}"].fill = section_header_fill
            ws[f"D{5}"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Category headers
            ws[f"D{6}"] = "Category"
            ws[f"E{6}"] = "Revenue"
            ws[f"F{6}"] = "Quantity"
            ws[f"D{6}"].font = subsection_font
            ws[f"E{6}"].font = subsection_font
            ws[f"F{6}"].font = subsection_font
            ws[f"D{6}"].fill = subsection_fill
            ws[f"E{6}"].fill = subsection_fill
            ws[f"F{6}"].fill = subsection_fill
            
            cat_row = 7
            categories_to_write = report_data['sales_by_category'][:10]  # Top 10 categories
            for cat in categories_to_write:
                ws[f"D{cat_row}"] = cat['category']
                ws[f"E{cat_row}"] = cat['revenue']
                ws[f"E{cat_row}"].number_format = currency_format
                ws[f"F{cat_row}"] = cat['quantity']
                ws[f"F{cat_row}"].number_format = number_format
                ws[f"D{cat_row}"].border = thin_border
                ws[f"E{cat_row}"].border = thin_border
                ws[f"F{cat_row}"].border = thin_border
                cat_row += 1
        
        # Move to next section
        row = max(row, cat_row) + 2
        
        # === ORDER DETAILS ===
        ws.merge_cells(f"A{row}:S{row}")
        ws[f"A{row}"] = "ORDER DETAILS"
        ws[f"A{row}"].font = section_header_font
        ws[f"A{row}"].fill = section_header_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Order details headers
        order_headers = [
            "Order Number", "Date", "Time", "Order Type", "Status", "Payment Status",
            "Payment Method", "Customer Name", "Customer Email", "Customer Phone", 
            "Cashier", "Subtotal", "Tax", "Tips", "Surcharges", 
            "Total Collected", "Discounts", "Items Count", "Total Quantity"
        ]
        
        for col, header in enumerate(order_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        
        try:
            start_date_obj = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date_obj = datetime.fromisoformat(end_str.replace('Z', '+00:00'))

            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            from orders.models import Order

            filters = {
                'status': Order.OrderStatus.COMPLETED,
                'completed_at__range': (start_date_obj, end_date_obj),
                'subtotal__gt': 0
            }

            # Add tenant filter
            if tenant_id:
                filters['tenant_id'] = tenant_id

            # Add location filter if specified
            if location_id is not None:
                filters['store_location_id'] = location_id

            orders = Order.objects.filter(**filters).select_related(
                'cashier', 'customer', 'payment_details'
            ).prefetch_related(
                'items__product'
            ).order_by('order_number')
            
            for order in orders:
                try:
                    items_count = order.items.count()
                    total_quantity = sum(item.quantity for item in order.items.all())
                    
                    # Get payment method
                    payment_method = "N/A"
                    if hasattr(order, 'payment_details') and order.payment_details:
                        transactions = order.payment_details.transactions.filter(
                            status="SUCCESSFUL"
                        ).first()
                        if transactions:
                            method_map = {
                                "CARD_TERMINAL": "Card (Terminal)",
                                "CASH": "Cash",
                                "GIFT_CARD": "Gift Card"
                            }
                            payment_method = method_map.get(transactions.method, transactions.method)
                    
                    # Customer info
                    customer_name = order.customer_display_name
                    customer_email = order.customer_email or ""
                    customer_phone = order.customer_phone or ""
                    
                    # Cashier
                    cashier_name = ""
                    if order.cashier:
                        cashier_name = f"{order.cashier.first_name} {order.cashier.last_name}".strip()
                    
                    # Payment details
                    total_tips = 0
                    total_surcharges = 0
                    total_collected = 0
                    if hasattr(order, 'payment_details') and order.payment_details:
                        total_tips = order.payment_details.total_tips or 0
                        total_surcharges = order.payment_details.total_surcharges or 0
                        total_collected = order.payment_details.total_collected or 0
                    
                    # Write order row
                    order_data = [
                        order.order_number,
                        order.completed_at.strftime("%Y-%m-%d") if order.completed_at else "N/A",
                        order.completed_at.strftime("%H:%M:%S") if order.completed_at else "N/A", 
                        order.order_type,
                        order.status,
                        "PAID" if order.payment_status == "PAID" else order.payment_status,
                        payment_method,
                        customer_name,
                        customer_email,
                        customer_phone,
                        cashier_name,
                        float(order.subtotal),
                        float(order.tax_total),
                        float(total_tips),
                        float(total_surcharges),
                        float(total_collected),
                        float(order.total_discounts_amount),
                        items_count,
                        total_quantity
                    ]
                    
                    for col, value in enumerate(order_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = thin_border
                        
                        # Apply number formats to currency columns
                        if col in [12, 13, 14, 15, 16, 17]:  # Currency columns (Subtotal, Tax, Tips, Surcharges, Total Collected, Discounts)
                            cell.number_format = currency_format
                        elif col in [18, 19]:  # Number columns (Items Count, Total Quantity)
                            cell.number_format = number_format
                            
                    row += 1
                        
                except Exception as order_error:
                    continue
                    
        except Exception as e:
            ws[f"A{row}"] = "Error: Could not load detailed order data"
            row += 1
        
        # Auto-adjust column widths
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            pass

    @staticmethod
    def _export_multi_location_sales_to_xlsx(wb, report_data: Dict[str, Any], header_font, header_fill, header_alignment):
        """Export multi-location sales report to Excel with multiple sheets."""
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.utils import timezone

        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        section_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_header_font = Font(bold=True, color="FFFFFF", size=12)
        subsection_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        subsection_font = Font(bold=True, size=11)
        currency_format = '"$"#,##0.00'
        number_format = '#,##0'

        # Sheet 1: Summary - All Locations
        summary_ws = wb.active
        summary_ws.title = "Summary - All Locations"

        row = 1

        # Title
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = "Sales Report - All Locations Summary"
        summary_ws[f"A{row}"].font = Font(bold=True, size=16)
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Generated date
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = f"Generated: {generated_at}"
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Date range
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        from datetime import datetime
        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            summary_ws.merge_cells(f"A{row}:F{row}")
            summary_ws[f"A{row}"] = f"Date Range: {start_date} to {end_date}"
            summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        except:
            summary_ws.merge_cells(f"A{row}:F{row}")
            summary_ws[f"A{row}"] = f"Date Range: {start_str} to {end_str}"
            summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Location count
        location_count = report_data.get('location_count', 0)
        summary_ws.merge_cells(f"A{row}:F{row}")
        summary_ws[f"A{row}"] = f"Locations Included: {location_count}"
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center")
        summary_ws[f"A{row}"].font = Font(bold=True)
        row += 2

        # === CONSOLIDATED TOTALS ===
        consolidated = report_data.get('consolidated', {})

        summary_ws.merge_cells(f"A{row}:B{row}")
        summary_ws[f"A{row}"] = "CONSOLIDATED FINANCIAL SUMMARY"
        summary_ws[f"A{row}"].font = section_header_font
        summary_ws[f"A{row}"].fill = section_header_fill
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Headers
        summary_ws[f"A{row}"] = "Metric"
        summary_ws[f"B{row}"] = "Total Value"
        summary_ws[f"A{row}"].font = subsection_font
        summary_ws[f"B{row}"].font = subsection_font
        summary_ws[f"A{row}"].fill = subsection_fill
        summary_ws[f"B{row}"].fill = subsection_fill
        row += 1

        # Financial metrics
        consolidated_metrics = [
            ("Total Revenue (All Locations)", consolidated.get('total_revenue', 0), currency_format),
            ("Net Revenue (All Locations)", consolidated.get('net_revenue', 0), currency_format),
            ("Total Orders (All Locations)", consolidated.get("total_orders", 0), number_format),
            ("Average Order Value (All Locations)", consolidated.get('avg_order_value', 0), currency_format),
            ("Total Items Sold (All Locations)", consolidated.get('total_items', 0), number_format),
            ("", "", ""),  # Empty row
            ("Total Subtotal", consolidated.get('total_subtotal', 0), currency_format),
            ("Total Tax Collected", consolidated.get('total_tax', 0), currency_format),
            ("Total Tips", consolidated.get('total_tips', 0), currency_format),
            ("Total Surcharges", consolidated.get('total_surcharges', 0), currency_format),
            ("Total Discounts Applied", -abs(consolidated.get('total_discounts', 0)), currency_format),
            ("Total Refunds", consolidated.get('total_refunds', 0), currency_format),
        ]

        for metric, value, fmt in consolidated_metrics:
            summary_ws[f"A{row}"] = metric
            summary_ws[f"B{row}"] = value if value != "" else ""
            if fmt and value != "":
                summary_ws[f"B{row}"].number_format = fmt
            summary_ws[f"A{row}"].border = thin_border
            summary_ws[f"B{row}"].border = thin_border
            # Color code negative values (discounts)
            if isinstance(value, (int, float)) and value < 0:
                summary_ws[f"B{row}"].font = Font(color="FF0000")
            row += 1

        row += 2

        # === COMPARISON TABLE ===
        summary_ws.merge_cells(f"A{row}:H{row}")
        summary_ws[f"A{row}"] = "LOCATION COMPARISON"
        summary_ws[f"A{row}"].font = section_header_font
        summary_ws[f"A{row}"].fill = section_header_fill
        summary_ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Comparison headers
        comparison_headers = ["Location", "Revenue", "Orders", "Avg Order", "Items", "Subtotal", "Tax", "Tips"]
        for col, header in enumerate(comparison_headers, 1):
            cell = summary_ws.cell(row=row, column=col, value=header)
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Location comparison rows
        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            comparison_data = [
                location_name,
                loc_report.get('total_revenue', 0),
                loc_report.get('total_orders', 0),
                loc_report.get('avg_order_value', 0),
                loc_report.get('total_items', 0),
                loc_report.get('total_subtotal', 0),
                loc_report.get('total_tax', 0),
                loc_report.get('total_tips', 0),
            ]

            for col, value in enumerate(comparison_data, 1):
                cell = summary_ws.cell(row=row, column=col, value=value)
                cell.border = thin_border

                # Apply number formats
                if col == 1:  # Location name
                    cell.alignment = Alignment(horizontal="left")
                elif col in [2, 4, 6, 7, 8]:  # Currency columns
                    cell.number_format = currency_format
                elif col in [3, 5]:  # Number columns
                    cell.number_format = number_format

            row += 1

        # Auto-adjust column widths for summary sheet
        for column in summary_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width

        # === CREATE INDIVIDUAL LOCATION SHEETS ===
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Create a new sheet for this location
            # Sanitize sheet name (Excel sheet names have restrictions)
            safe_sheet_name = location_name[:31]  # Excel sheet name max length is 31
            # Remove invalid characters
            for char in ['\\', '/', '*', '?', ':', '[', ']']:
                safe_sheet_name = safe_sheet_name.replace(char, '')

            location_ws = wb.create_sheet(title=safe_sheet_name)

            # Export this location's data using the existing single-location export logic
            SalesReportService._export_sales_to_xlsx(location_ws, loc_report, header_font, header_fill, header_alignment)

    @staticmethod
    def export_sales_to_pdf(report_data: Dict[str, Any], story, styles):
        """Export sales report to PDF format."""
        # Check if this is a multi-location report
        if report_data.get('is_multi_location', False):
            SalesReportService._export_multi_location_sales_to_pdf(story, report_data, styles)
        else:
            SalesReportService._export_sales_to_pdf(story, report_data, styles)

    @staticmethod
    def _export_sales_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export comprehensive sales report to PDF matching the CSV and XLSX formats"""
        from django.utils import timezone
        from datetime import datetime
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Spacer, PageBreak
        
        # Report header
        story.append(Paragraph("Sales Report", styles["Title"]))
        story.append(Spacer(1, 12))
        
        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))
        
        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")
        
        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))
        except:
            story.append(Paragraph(f"Date Range: {start_str} to {end_str}", styles["Normal"]))

        # Location info
        location_info = report_data.get("location_info", {})
        location_name = location_info.get("location_name", "All Locations")
        story.append(Paragraph(f"<b>Location: {location_name}</b>", styles["Normal"]))

        story.append(Spacer(1, 20))

        # === FINANCIAL SUMMARY ===
        story.append(Paragraph("Financial Summary", styles["Heading2"]))
        story.append(Spacer(1, 12))

        financial_data = [
            ["Metric", "Value"],
            ["Total Revenue (Collected)", f"${report_data.get('total_revenue', 0):,.2f}"],
            ["Net Revenue", f"${report_data.get('net_revenue', 0):,.2f}"],
            ["Total Orders", f"{report_data.get('total_orders', 0):,}"],
            ["Average Order Value", f"${report_data.get('avg_order_value', 0):,.2f}"],
            ["Total Items Sold", f"{report_data.get('total_items', 0):,}"],
        ]

        financial_table = Table(financial_data, colWidths=[3 * inch, 2.5 * inch])
        financial_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])
        )

        story.append(financial_table)
        story.append(Spacer(1, 20))

        # === REVENUE BREAKDOWN ===
        story.append(Paragraph("Revenue Breakdown", styles["Heading2"]))
        story.append(Spacer(1, 12))

        revenue_data = [
            ["Component", "Amount"],
            ["Subtotal", f"${report_data.get('total_subtotal', 0):,.2f}"],
            ["Tax Collected", f"${report_data.get('total_tax', 0):,.2f}"],
            ["Tips", f"${report_data.get('total_tips', 0):,.2f}"],
            ["Surcharges", f"${report_data.get('total_surcharges', 0):,.2f}"],
            ["Discounts Applied", f"-${report_data.get('total_discounts', 0):,.2f}"],
        ]

        revenue_table = Table(revenue_data, colWidths=[3 * inch, 2.5 * inch])
        revenue_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TEXTCOLOR", (1, -1), (1, -1), colors.red),  # Make discount red
            ])
        )

        story.append(revenue_table)
        story.append(Spacer(1, 20))

        # === SALES BY CATEGORY ===
        if 'sales_by_category' in report_data and report_data['sales_by_category']:
            story.append(Paragraph("Top Sales by Category", styles["Heading2"]))
            story.append(Spacer(1, 12))

            category_data = [["Category", "Revenue", "Quantity"]]
            for cat in report_data['sales_by_category'][:10]:  # Top 10 categories
                category_data.append([
                    cat['category'],
                    f"${cat['revenue']:,.2f}",
                    f"{cat['quantity']:,}"
                ])

            category_table = Table(category_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
            category_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ])
            )

            story.append(category_table)
            story.append(Spacer(1, 20))

        # === ORDER DETAILS ===
        story.append(PageBreak())
        story.append(Paragraph("Order Details", styles["Heading1"]))
        story.append(Spacer(1, 12))

        # Get individual order details
        try:
            start_date_obj = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end_date_obj = datetime.fromisoformat(end_str.replace('Z', '+00:00'))

            # Extract tenant and location filters
            tenant_id = report_data.get('tenant_id')
            location_id = report_data.get('location_info', {}).get('location_id')

            from orders.models import Order

            filters = {
                'status': Order.OrderStatus.COMPLETED,
                'completed_at__range': (start_date_obj, end_date_obj),
                'subtotal__gt': 0
            }

            # Add tenant filter
            if tenant_id:
                filters['tenant_id'] = tenant_id

            # Add location filter if specified
            if location_id is not None:
                filters['store_location_id'] = location_id

            orders = Order.objects.filter(**filters).select_related(
                'cashier', 'customer', 'payment_details'
            ).prefetch_related(
                'items__product'
            ).order_by('order_number')[:50]  # Limit to 50 orders for PDF readability

            if orders.exists():
                # Order details table - essential data for PDF readability
                order_data = [
                    ["Order #", "Date", "Time", "Type", "Status", "Pay Status", 
                     "Pay Method", "Subtotal", "Tax", "Tips", "Surcharges", "Collected", "Discounts"]
                ]

                for order in orders:
                    # Get payment method
                    payment_method = "N/A"
                    if hasattr(order, 'payment_details') and order.payment_details:
                        transactions = order.payment_details.transactions.filter(
                            status="SUCCESSFUL"
                        ).first()
                        if transactions:
                            method_map = {
                                "CARD_TERMINAL": "Card",
                                "CASH": "Cash",
                                "GIFT_CARD": "Gift Card"
                            }
                            payment_method = method_map.get(transactions.method, transactions.method)
                    
                    # Payment details - use safe access
                    total_tips = 0
                    total_surcharges = 0
                    total_collected = 0
                    if hasattr(order, 'payment_details') and order.payment_details:
                        total_tips = order.payment_details.total_tips or 0
                        total_surcharges = order.payment_details.total_surcharges or 0
                        total_collected = order.payment_details.total_collected or 0

                    order_row = [
                        order.order_number,
                        order.completed_at.strftime("%m/%d/%Y") if order.completed_at else "N/A",
                        order.completed_at.strftime("%H:%M") if order.completed_at else "N/A",
                        order.order_type,
                        order.status,
                        "PAID" if order.payment_status == "PAID" else order.payment_status,
                        payment_method,
                        f"${order.subtotal:.2f}",
                        f"${order.tax_total:.2f}",
                        f"${total_tips:.2f}",
                        f"${total_surcharges:.2f}",
                        f"${total_collected:.2f}",
                        f"${order.total_discounts_amount:.2f}",
                    ]
                    order_data.append(order_row)

                # Create table with more comfortable column widths for PDF readability
                order_table = Table(order_data, colWidths=[
                    0.8*inch,  # Order #
                    0.8*inch,  # Date
                    0.5*inch,  # Time
                    0.8*inch,  # Type
                    0.8*inch,  # Status
                    0.8*inch,  # Pay Status
                    0.8*inch,  # Pay Method
                    0.8*inch,  # Subtotal
                    0.6*inch,  # Tax
                    0.6*inch,  # Tips
                    0.7*inch,  # Surcharges
                    0.8*inch,  # Collected
                    0.7*inch,  # Discounts
                ])
                order_table.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 1), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ])
                )

                story.append(order_table)
                
                if orders.count() > 50:
                    story.append(Spacer(1, 12))
                    story.append(Paragraph(f"Note: Showing first 50 orders of {orders.count()} total orders.", styles["Normal"]))

            else:
                story.append(Paragraph("No completed orders found in the selected date range.", styles["Normal"]))
                
        except Exception as e:
            story.append(Paragraph("Error: Could not load detailed order data", styles["Normal"]))

        story.append(Spacer(1, 20))

    @staticmethod
    def _export_multi_location_sales_to_pdf(story, report_data: Dict[str, Any], styles):
        """Export multi-location sales report to PDF with multiple pages."""
        from django.utils import timezone
        from datetime import datetime
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Spacer, PageBreak, Table, TableStyle, Paragraph

        # === EXECUTIVE SUMMARY PAGE ===
        story.append(Paragraph("Sales Report - All Locations", styles["Title"]))
        story.append(Paragraph("Executive Summary", styles["Heading1"]))
        story.append(Spacer(1, 12))

        # Generated date and range
        generated_at = report_data.get('generated_at', timezone.now().isoformat())
        story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))

        date_range = report_data.get("date_range", {})
        start_str = date_range.get("start", "")
        end_str = date_range.get("end", "")

        try:
            start_date = datetime.fromisoformat(start_str.replace('Z', '+00:00')).date()
            end_date = datetime.fromisoformat(end_str.replace('Z', '+00:00')).date()
            story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))
        except:
            story.append(Paragraph(f"Date Range: {start_str} to {end_str}", styles["Normal"]))

        # Location count
        location_count = report_data.get('location_count', 0)
        story.append(Paragraph(f"<b>Locations Included: {location_count}</b>", styles["Normal"]))
        story.append(Spacer(1, 20))

        # === CONSOLIDATED FINANCIAL SUMMARY ===
        consolidated = report_data.get('consolidated', {})

        story.append(Paragraph("Consolidated Financial Summary (All Locations)", styles["Heading2"]))
        story.append(Spacer(1, 12))

        consolidated_summary_data = [
            ["Metric", "Total Value"],
            ["Total Revenue (All Locations)", f"${consolidated.get('total_revenue', 0):,.2f}"],
            ["Net Revenue (All Locations)", f"${consolidated.get('net_revenue', 0):,.2f}"],
            ["Total Orders (All Locations)", f"{consolidated.get('total_orders', 0):,}"],
            ["Average Order Value (All Locations)", f"${consolidated.get('avg_order_value', 0):,.2f}"],
            ["Total Items Sold (All Locations)", f"{consolidated.get('total_items', 0):,}"],
        ]

        summary_table = Table(consolidated_summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 20))

        # === CONSOLIDATED REVENUE BREAKDOWN ===
        story.append(Paragraph("Consolidated Revenue Breakdown", styles["Heading2"]))
        story.append(Spacer(1, 12))

        revenue_breakdown_data = [
            ["Component", "Amount"],
            ["Total Subtotal", f"${consolidated.get('total_subtotal', 0):,.2f}"],
            ["Total Tax Collected", f"${consolidated.get('total_tax', 0):,.2f}"],
            ["Total Tips", f"${consolidated.get('total_tips', 0):,.2f}"],
            ["Total Surcharges", f"${consolidated.get('total_surcharges', 0):,.2f}"],
            ["Total Discounts Applied", f"-${abs(consolidated.get('total_discounts', 0)):,.2f}"],
            ["Total Refunds", f"${consolidated.get('total_refunds', 0):,.2f}"],
        ]

        breakdown_table = Table(revenue_breakdown_data, colWidths=[4*inch, 2*inch])
        breakdown_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(breakdown_table)
        story.append(Spacer(1, 20))

        # === LOCATION COMPARISON TABLE ===
        story.append(Paragraph("Location Comparison", styles["Heading2"]))
        story.append(Spacer(1, 12))

        comparison_data = [
            ["Location", "Revenue", "Orders", "Avg Order", "Items", "Tax", "Tips"]
        ]

        locations = report_data.get('locations', [])
        for location_data in locations:
            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            comparison_data.append([
                location_name,
                f"${loc_report.get('total_revenue', 0):,.2f}",
                f"{loc_report.get('total_orders', 0):,}",
                f"${loc_report.get('avg_order_value', 0):,.2f}",
                f"{loc_report.get('total_items', 0):,}",
                f"${loc_report.get('total_tax', 0):,.2f}",
                f"${loc_report.get('total_tips', 0):,.2f}",
            ])

        comparison_table = Table(comparison_data, colWidths=[
            1.8*inch,  # Location
            1.2*inch,  # Revenue
            0.8*inch,  # Orders
            1*inch,    # Avg Order
            0.7*inch,  # Items
            0.9*inch,  # Tax
            0.9*inch,  # Tips
        ])
        comparison_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(comparison_table)
        story.append(Spacer(1, 20))

        # === INDIVIDUAL LOCATION PAGES ===
        for i, location_data in enumerate(locations, 1):
            # Page break before each location (except first, which follows summary)
            story.append(PageBreak())

            location_name = location_data.get('location_name', 'Unknown')
            loc_report = location_data.get('report_data', {})

            # Location header
            story.append(Paragraph(f"Location {i}: {location_name}", styles["Title"]))
            story.append(Paragraph("Detailed Sales Report", styles["Heading1"]))
            story.append(Spacer(1, 12))

            # Add date range for this location
            story.append(Paragraph(f"Date Range: {start_date} to {end_date}", styles["Normal"]))
            story.append(Spacer(1, 20))

            # Export this location's data using the existing single-location export logic
            # We'll just add the key sections without the header
            SalesReportService._export_single_location_section_to_pdf(story, loc_report, styles, include_header=False)

    @staticmethod
    def _export_single_location_section_to_pdf(story, report_data: Dict[str, Any], styles, include_header=True):
        """Helper method to export a single location's data section for multi-location PDF reports."""
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import Spacer, Table, TableStyle, Paragraph

        # === FINANCIAL SUMMARY ===
        story.append(Paragraph("Financial Summary", styles["Heading2"]))
        story.append(Spacer(1, 12))

        financial_data = [
            ["Metric", "Value"],
            ["Total Revenue (Collected)", f"${report_data.get('total_revenue', 0):,.2f}"],
            ["Net Revenue", f"${report_data.get('net_revenue', 0):,.2f}"],
            ["Total Orders", f"{report_data.get('total_orders', 0):,}"],
            ["Average Order Value", f"${report_data.get('avg_order_value', 0):,.2f}"],
            ["Total Items Sold", f"{report_data.get('total_items', 0):,}"],
        ]

        fin_table = Table(financial_data, colWidths=[4*inch, 2*inch])
        fin_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(fin_table)
        story.append(Spacer(1, 20))

        # === REVENUE BREAKDOWN ===
        story.append(Paragraph("Revenue Breakdown", styles["Heading2"]))
        story.append(Spacer(1, 12))

        revenue_data = [
            ["Component", "Amount"],
            ["Subtotal", f"${report_data.get('total_subtotal', 0):,.2f}"],
            ["Tax Collected", f"${report_data.get('total_tax', 0):,.2f}"],
            ["Tips", f"${report_data.get('total_tips', 0):,.2f}"],
            ["Surcharges", f"${report_data.get('total_surcharges', 0):,.2f}"],
            ["Discounts Applied", f"-${abs(report_data.get('total_discounts', 0)):,.2f}"],
        ]

        rev_table = Table(revenue_data, colWidths=[4*inch, 2*inch])
        rev_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(rev_table)
        story.append(Spacer(1, 20))

        # === SALES BY CATEGORY ===
        if 'sales_by_category' in report_data and report_data['sales_by_category']:
            story.append(Paragraph("Sales by Category (Top 10)", styles["Heading2"]))
            story.append(Spacer(1, 12))

            category_data = [["Category", "Revenue", "Quantity"]]
            categories_to_write = report_data['sales_by_category'][:10]
            for cat in categories_to_write:
                category_data.append([
                    cat['category'],
                    f"${cat['revenue']:,.2f}",
                    f"{cat['quantity']:,}"
                ])

            cat_table = Table(category_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            cat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(cat_table)
            story.append(Spacer(1, 20))